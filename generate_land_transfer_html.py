from __future__ import annotations

import base64
import html
import json
import re
import shutil
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


BASE_DIR = Path(r"D:\Codex\Tools")
SOURCE = Path(r"C:\Users\QUANGHUY\Downloads\Bieu_chu_chuyen_dat_dai_mau_cong_thuc.xlsx")
OUT = BASE_DIR / "public" / "index.html"
JSZIP = Path(r"C:\Users\QUANGHUY\.cache\codex-runtimes\codex-primary-runtime\dependencies\node\node_modules\jszip\dist\jszip.min.js")
LOGO = Path(r"C:\Users\QUANGHUY\Downloads\482087578_122221961630205345_1940337838885474762_n.jpg")
HOME_BACKGROUND = Path(r"C:\Users\QUANGHUY\Downloads\ChatGPT Image 12_00_04 30 thg 4, 2026.png")

LAND_NAME_FIXES = {
    "Đất côn trình thủy lợi": "Đất công trình thủy lợi",
}

STT_FIXES_BY_CODE = {
    "NKH": "1.10",
    "TIN": "2.10",
}

HEADER_ROW = 3
CURRENT_COL = 4
MATRIX_START_COL = 5
MATRIX_END_COL = 66
DECREASE_COL = 67
CHANGE_COL = 68
PLAN_COL = 69
PREVIOUS_PLAN_COL = 70
PREVIOUS_PLAN_CHANGE_COL = 71
PREVIOUS_PLAN_STRUCTURE_COL = 72
TOTAL_INCREASE_ROW = 67
PLAN_ROW = 68
PREVIOUS_PLAN_DIR = BASE_DIR / "Dulieu"
SAMPLE_DIR = BASE_DIR / "public" / "samples"
LEGACY_SAMPLE_DIR = BASE_DIR / "samples"
SAMPLE_FILES = [
    ("Dữ liệu hiện trạng mẫu.xlsx", "hien-trang-mau.xlsx", "Dữ liệu hiện trạng mẫu"),
    ("Kết quả thực hiện quy hoạch năm mẫu.xlsx", "ket-qua-quy-hoach-nam-mau.xlsx", "Kết quả thực hiện quy hoạch năm mẫu"),
    ("Bảng quy hoạch mẫu.xlsx", "bang-quy-hoach-mau.xlsx", "Bảng quy hoạch mẫu"),
]
WEBGIS_SAMPLE_DATA = {
    "type": "FeatureCollection",
    "features": [
        {
            "type": "Feature",
            "properties": {"layer": "administrative", "ten": "Ranh giới xã mẫu", "ma_dv": "XA-001", "ghi_chu": "Ranh giới hành chính phục vụ demo WebGIS"},
            "geometry": {"type": "Polygon", "coordinates": [[[105.8422, 21.0474], [105.8616, 21.0474], [105.8616, 21.0330], [105.8422, 21.0330], [105.8422, 21.0474]]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "parcels", "ma_thua": "TD-101", "chu_su_dung": "Nguyễn Văn A", "loai_dat": "ONT", "dien_tich": 1240.5, "muc_dich": "Đất ở tại nông thôn", "quy_hoach": "Đất ở", "dia_danh": "Thôn Đông", "ghi_chu": "Thửa đất mẫu"},
            "geometry": {"type": "Polygon", "coordinates": [[[105.8481, 21.0424], [105.8515, 21.0423], [105.8513, 21.0398], [105.8478, 21.0399], [105.8481, 21.0424]]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "parcels", "ma_thua": "TD-102", "chu_su_dung": "Trần Thị B", "loai_dat": "LUC", "dien_tich": 3560.0, "muc_dich": "Đất trồng lúa nước", "quy_hoach": "Đất nông nghiệp", "dia_danh": "Cánh đồng Bắc", "ghi_chu": "Giữ nguyên hiện trạng"},
            "geometry": {"type": "Polygon", "coordinates": [[[105.8517, 21.0422], [105.8562, 21.0421], [105.8560, 21.0395], [105.8515, 21.0397], [105.8517, 21.0422]]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "landuse", "ma_khoanh": "HT-01", "loai_dat": "LUC", "dien_tich": 6.42, "muc_dich": "Đất trồng lúa", "quy_hoach": "Một phần chuyển sang giao thông", "ghi_chu": "Hiện trạng sử dụng đất"},
            "geometry": {"type": "Polygon", "coordinates": [[[105.8460, 21.0384], [105.8542, 21.0381], [105.8540, 21.0346], [105.8458, 21.0349], [105.8460, 21.0384]]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "planning", "ma_khoanh": "QH-02", "loai_dat": "DGT", "dien_tich": 1.18, "muc_dich": "Đất giao thông", "quy_hoach": "Tuyến đường quy hoạch", "loai_quy_hoach": "Hạ tầng giao thông", "ghi_chu": "Vùng quy hoạch mẫu"},
            "geometry": {"type": "Polygon", "coordinates": [[[105.8450, 21.0400], [105.8600, 21.0395], [105.8600, 21.0387], [105.8450, 21.0392], [105.8450, 21.0400]]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "roads", "ten": "Đường trục xã", "loai_dat": "DGT", "dien_tich": 0.84, "muc_dich": "Giao thông", "quy_hoach": "Nâng cấp mở rộng", "ghi_chu": "Tuyến đường mẫu"},
            "geometry": {"type": "LineString", "coordinates": [[105.8442, 21.0436], [105.8490, 21.0411], [105.8548, 21.0390], [105.8608, 21.0362]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "water", "ten": "Kênh tiêu nội đồng", "loai_dat": "DTL", "dien_tich": 0.52, "muc_dich": "Thủy lợi", "quy_hoach": "Giữ nguyên", "ghi_chu": "Tuyến thủy hệ mẫu"},
            "geometry": {"type": "LineString", "coordinates": [[105.8435, 21.0360], [105.8497, 21.0375], [105.8552, 21.0370], [105.8610, 21.0350]]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "public", "ten": "Trụ sở UBND xã", "loai_dat": "TSC", "dien_tich": 0.32, "muc_dich": "Đất trụ sở cơ quan", "quy_hoach": "Giữ nguyên công trình công cộng", "ghi_chu": "Điểm công trình công cộng"},
            "geometry": {"type": "Point", "coordinates": [105.8520, 21.0437]},
        },
        {
            "type": "Feature",
            "properties": {"layer": "public", "ten": "Trường tiểu học", "loai_dat": "DGD", "dien_tich": 0.78, "muc_dich": "Đất giáo dục", "quy_hoach": "Mở rộng khuôn viên", "ghi_chu": "Điểm công trình công cộng"},
            "geometry": {"type": "Point", "coordinates": [105.8570, 21.0413]},
        },
    ],
}

WEBGIS_CSS = r"""
.webgis-page {
  display: flex;
  flex-direction: column;
  min-height: 0;
  overflow: hidden;
  background: linear-gradient(135deg, #f8fbff 0%, #eef6ff 48%, #f8fafc 100%);
}
.webgis-shell {
  display: flex;
  flex-direction: column;
  flex: 1 1 auto;
  height: 100%;
  min-height: 0;
  border-radius: 18px;
  overflow: hidden;
}
.webgis-topbar {
  flex: 0 0 auto;
  display: grid;
  grid-template-columns: minmax(260px, 0.92fr) minmax(320px, 560px) auto;
  gap: 12px;
  align-items: center;
  padding: 12px 16px;
  border-bottom: 1px solid #dbe7f3;
  background: linear-gradient(135deg, rgba(255,255,255,0.98), rgba(237,248,255,0.96));
}
.webgis-title {
  min-width: 0;
}
.webgis-title strong {
  display: block;
  color: #0f2f57;
  font-size: 18px;
  letter-spacing: 0;
}
.webgis-title span {
  display: block;
  margin-top: 3px;
  color: #55708d;
  font-size: 12px;
}
.webgis-search {
  position: relative;
  display: flex;
  gap: 8px;
  min-width: 0;
}
.webgis-search input,
.webgis-panel input,
.webgis-panel select,
.webgis-panel textarea,
.webgis-attr-tools input,
.webgis-attr-tools select {
  height: 36px;
  min-width: 0;
  border: 1px solid #c9d8e8;
  border-radius: 10px;
  padding: 7px 10px;
  background: #fff;
  color: #102033;
  font-size: 13px;
}
.webgis-search input {
  flex: 1;
}
.webgis-actions,
.webgis-map-tools,
.webgis-attr-tools {
  display: flex;
  align-items: center;
  flex-wrap: wrap;
  gap: 8px;
}
.webgis-actions {
  justify-content: flex-end;
}
.webgis-save-status {
  min-height: 26px;
  display: inline-flex;
  align-items: center;
  gap: 6px;
  border: 1px solid #bfdbfe;
  border-radius: 999px;
  padding: 4px 9px;
  background: rgba(239, 246, 255, 0.82);
  color: #1d4ed8;
  font-size: 12px;
  font-weight: 700;
  box-shadow: none;
}
.webgis-save-status::before {
  content: "";
  width: 8px;
  height: 8px;
  border-radius: 50%;
  background: #94a3b8;
  box-shadow: 0 0 0 3px rgba(148, 163, 184, 0.16);
}
.webgis-save-status.connected {
  border-color: #bbf7d0;
  background: #f0fdf4;
  color: #166534;
}
.webgis-save-status.connected::before {
  background: #22c55e;
  box-shadow: 0 0 0 3px rgba(34, 197, 94, 0.16);
}
.webgis-save-status.error {
  border-color: #fecaca;
  background: #fff1f2;
  color: #991b1b;
}
.webgis-save-status.error::before {
  background: #ef4444;
  box-shadow: 0 0 0 3px rgba(239, 68, 68, 0.15);
}
.webgis-stats {
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
  margin-top: 8px;
}
.webgis-stat {
  display: inline-flex;
  align-items: center;
  min-height: 24px;
  padding: 3px 8px;
  border: 1px solid #cfe0f2;
  border-radius: 999px;
  background: #f8fbff;
  color: #315d87;
  font-size: 12px;
  font-weight: 800;
}
.webgis-actions button,
.webgis-map-tools button,
.webgis-search button,
.webgis-panel button,
.webgis-attr-tools button {
  height: 36px;
  border: 1px solid #c1d3e6;
  border-radius: 10px;
  padding: 0 12px;
  background: #fff;
  color: #143452;
  font-size: 12px;
  font-weight: 800;
  cursor: pointer;
  box-shadow: 0 6px 16px rgba(15, 47, 87, 0.08);
}
.webgis-actions button {
  background: #f8fbff;
}
.webgis-actions .primary,
.webgis-search .primary,
.webgis-panel .primary {
  border-color: #0f766e;
  background: #0f766e;
  color: #fff;
}
.webgis-actions .admin-action {
  border-color: #0f766e;
  background: linear-gradient(135deg, #0f766e, #0d9488);
  color: #fff;
  box-shadow: 0 10px 22px rgba(15, 118, 110, 0.18);
}
.webgis-search-results {
  position: absolute;
  z-index: 540;
  top: calc(100% + 7px);
  left: 0;
  right: 0;
  max-height: 260px;
  overflow: auto;
  border: 1px solid #bdd0e3;
  border-radius: 8px;
  background: #fff;
  box-shadow: 0 18px 34px rgba(15, 47, 87, 0.18);
}
.webgis-search-results[hidden] {
  display: none;
}
.webgis-result-item {
  width: 100%;
  min-height: 42px;
  border: 0;
  border-bottom: 1px solid #edf2f7;
  padding: 8px 10px;
  background: #fff;
  color: #102033;
  text-align: left;
  cursor: pointer;
}
.webgis-result-item:hover {
  background: #eff6ff;
}
.webgis-result-item strong {
  display: block;
  font-size: 13px;
}
.webgis-result-item span {
  display: block;
  margin-top: 2px;
  color: #64748b;
  font-size: 12px;
}
.webgis-workspace {
  display: grid;
  grid-template-columns: 304px minmax(520px, 1fr) 340px;
  gap: 12px;
  align-items: stretch;
  min-height: 0;
  flex: 1 1 auto;
  padding: 12px;
  overflow: hidden;
  transition: grid-template-columns 0.24s ease;
}
.webgis-page.layers-collapsed .webgis-workspace {
  grid-template-columns: 46px minmax(520px, 1fr) 340px;
}
.webgis-page.info-collapsed .webgis-workspace {
  grid-template-columns: 304px minmax(520px, 1fr) 46px;
}
.webgis-page.layers-collapsed.info-collapsed .webgis-workspace {
  grid-template-columns: 46px minmax(520px, 1fr) 46px;
}
.webgis-sidebar,
.webgis-info {
  display: flex;
  flex-direction: column;
  gap: 10px;
  height: 100%;
  max-height: 100%;
  min-height: 0;
  min-width: 0;
  overflow: auto;
  transition: width 0.24s ease;
}
.webgis-page.layers-collapsed .webgis-sidebar,
.webgis-page.info-collapsed .webgis-info {
  overflow: hidden;
}
.webgis-page.layers-collapsed .webgis-sidebar .webgis-panel-body,
.webgis-page.layers-collapsed .webgis-sidebar .webgis-admin-panel,
.webgis-page.info-collapsed .webgis-info .webgis-panel-body,
.webgis-page.info-collapsed .webgis-info .webgis-panel:not(:first-child) {
  display: none;
}
.webgis-page.layers-collapsed .webgis-sidebar .webgis-panel,
.webgis-page.info-collapsed .webgis-info .webgis-panel {
  min-height: 0;
  flex: 0 0 auto;
}
.webgis-page.layers-collapsed .webgis-sidebar h2,
.webgis-page.info-collapsed .webgis-info h2 {
  writing-mode: vertical-rl;
  text-orientation: mixed;
  white-space: nowrap;
}
.webgis-page.layers-collapsed .webgis-sidebar .webgis-panel-head,
.webgis-page.info-collapsed .webgis-info .webgis-panel-head {
  min-height: 136px;
  flex-direction: column;
  justify-content: flex-start;
  padding: 8px 6px;
}
.webgis-page.layers-collapsed .webgis-sidebar .webgis-panel-actions {
  flex-direction: column;
}
.webgis-page.layers-collapsed #webgisFitAllBtn {
  display: none;
}
.webgis-page.layers-collapsed #webgisToggleSidebarBtn,
.webgis-page.info-collapsed #webgisToggleInfoBtn {
  display: inline-grid !important;
  width: 30px;
  min-width: 30px;
}
.webgis-collapse-btn {
  width: 28px;
  height: 28px !important;
  padding: 0 !important;
  border-radius: 9px !important;
  display: inline-grid;
  place-items: center;
  flex: 0 0 auto;
}
.webgis-panel {
  flex: 0 0 auto;
  border: 1px solid #d4e2ef;
  border-radius: 14px;
  background: rgba(255,255,255,0.96);
  box-shadow: 0 12px 30px rgba(15, 47, 87, 0.08);
  overflow: hidden;
}
.webgis-sidebar > .webgis-panel:first-child,
.webgis-info > .webgis-panel:first-child {
  flex: 1 1 auto;
  min-height: 0;
  display: flex;
  flex-direction: column;
}
.webgis-panel-head {
  flex: 0 0 auto;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 8px;
  padding: 10px 12px;
  border-bottom: 1px solid #e4edf6;
}
.webgis-panel-actions {
  display: inline-flex;
  align-items: center;
  gap: 6px;
}
.webgis-panel-head h2,
.webgis-panel-head h3 {
  margin: 0;
  color: #0f2f57;
  font-size: 14px;
}
.webgis-panel-body {
  min-height: 0;
  padding: 10px 12px;
}
.webgis-sidebar > .webgis-panel:first-child .webgis-panel-body,
.webgis-info > .webgis-panel:first-child .webgis-panel-body {
  flex: 1 1 auto;
  overflow: auto;
}
.webgis-layer-list {
  display: flex;
  flex-direction: column;
  gap: 7px;
}
.webgis-layer-item {
  display: grid;
  grid-template-columns: auto 1fr auto;
  gap: 7px;
  align-items: center;
  padding: 8px;
  border: 1px solid #e3edf6;
  border-radius: 12px;
  background: linear-gradient(135deg, #ffffff, #f8fbff);
  box-shadow: 0 8px 18px rgba(15, 47, 87, 0.05);
}
.webgis-layer-main {
  min-width: 0;
}
.webgis-layer-main label {
  display: flex;
  align-items: center;
  gap: 7px;
  color: #102033;
  font-size: 12px;
  font-weight: 800;
  line-height: 1.25;
}
.webgis-layer-count {
  display: inline-flex;
  margin-top: 3px;
  color: #64748b;
  font-size: 11px;
  font-weight: 700;
}
.webgis-layer-tools {
  grid-column: 1 / 4;
  display: flex;
  align-items: center;
  gap: 7px;
  color: #64748b;
  font-size: 11px;
  font-weight: 800;
}
.webgis-layer-tools input[type="range"] {
  flex: 1;
  accent-color: #2563eb;
}
.webgis-layer-actions {
  display: inline-flex;
  align-items: center;
  gap: 5px;
}
.webgis-icon-btn {
  width: 30px;
  height: 30px !important;
  min-width: 30px;
  padding: 0 !important;
  display: inline-grid;
  place-items: center;
  border-radius: 9px !important;
  font-size: 14px !important;
  line-height: 1;
}
.webgis-symbol {
  width: 17px;
  height: 17px;
  border: 1px solid rgba(15,47,87,0.20);
  border-radius: 5px;
  flex: 0 0 auto;
}
.webgis-map-panel {
  position: relative;
  height: 100%;
  min-height: 0;
  overflow: hidden;
  border: 1px solid #cbdcec;
  border-radius: 16px;
  background: #eaf2f8;
  box-shadow: inset 0 0 0 1px rgba(255,255,255,0.65), 0 12px 30px rgba(15, 47, 87, 0.12);
}
.webgis-map {
  width: 100%;
  height: 100%;
  min-height: 0;
}
.webgis-map-tools {
  position: absolute;
  z-index: 500;
  top: 14px;
  left: 50%;
  transform: translateX(-50%);
  max-width: calc(100% - 24px);
  justify-content: center;
  padding: 8px 10px;
  border: 1px solid rgba(183, 200, 218, 0.72);
  border-radius: 20px;
  background: rgba(255,255,255,0.94);
  box-shadow: 0 14px 30px rgba(15,47,87,0.14);
  backdrop-filter: blur(8px);
}
.webgis-tool-group {
  display: inline-flex;
  align-items: center;
  gap: 6px;
}
.webgis-tool-divider {
  width: 1px;
  height: 22px;
  background: rgba(148, 163, 184, 0.42);
}
.webgis-map-tools button {
  height: 32px;
  border-radius: 999px;
  padding: 0 11px;
  box-shadow: none;
}
.webgis-coordinate-bar,
.webgis-measure-badge {
  position: absolute;
  z-index: 500;
  right: 12px;
  padding: 7px 10px;
  border-radius: 8px;
  background: rgba(15, 47, 87, 0.86);
  color: #fff;
  font-size: 12px;
  box-shadow: 0 12px 28px rgba(15,47,87,0.16);
}
.webgis-coordinate-bar {
  bottom: 12px;
}
.webgis-measure-badge {
  left: 12px;
  right: auto;
  bottom: 12px;
  max-width: 420px;
}
.webgis-detail-empty {
  min-height: 140px;
  display: block;
  color: #64748b;
  font-size: 13px;
}
.webgis-empty-card {
  display: grid;
  gap: 10px;
  min-height: 190px;
  padding: 16px;
  border: 1px dashed #cbdcec;
  border-radius: 12px;
  background: linear-gradient(135deg, #f8fbff, #eef6ff);
}
.webgis-empty-icon {
  width: 44px;
  height: 44px;
  display: grid;
  place-items: center;
  border-radius: 14px;
  background: #dbeafe;
  color: #1d4ed8;
  font-size: 22px;
  font-weight: 800;
}
.webgis-empty-card strong {
  color: #0f2f57;
  font-size: 15px;
}
.webgis-empty-card ul {
  margin: 0;
  padding-left: 18px;
  line-height: 1.55;
}
.webgis-detail-table {
  width: 100%;
  border-collapse: collapse;
  font-size: 13px;
  overflow: hidden;
  border: 1px solid #e4edf6;
  border-radius: 12px;
}
.webgis-detail-table th,
.webgis-detail-table td {
  padding: 7px 5px;
  border-bottom: 1px solid #edf2f7;
  text-align: left;
  vertical-align: top;
}
.webgis-detail-table th {
  width: 112px;
  color: #64748b;
  font-weight: 700;
  background: #f8fbff;
}
.webgis-detail-actions {
  display: flex;
  flex-wrap: wrap;
  gap: 8px;
  margin-bottom: 10px;
}
.webgis-detail-title {
  margin: 0 0 10px;
  color: #0f2f57;
  font-size: 15px;
  font-weight: 900;
}
.webgis-admin-panel[hidden],
.webgis-attr-panel[hidden] {
  display: none;
}
.webgis-admin-grid {
  display: grid;
  gap: 8px;
}
.webgis-admin-head-actions {
  display: flex;
  align-items: center;
  gap: 6px;
}
.webgis-admin-session {
  min-height: 24px;
  display: inline-flex;
  align-items: center;
  border: 1px solid #99f6e4;
  border-radius: 999px;
  padding: 2px 8px;
  background: #f0fdfa;
  color: #115e59;
  font-size: 11px;
  font-weight: 800;
}
.webgis-admin-grid label {
  display: grid;
  gap: 4px;
  color: #334155;
  font-size: 12px;
  font-weight: 700;
}
.webgis-admin-note {
  margin: 0;
  color: #64748b;
  font-size: 12px;
  line-height: 1.4;
}
.webgis-admin-status {
  margin: 0;
  border: 1px solid #bbf7d0;
  border-radius: 8px;
  padding: 8px 9px;
  background: #f0fdf4;
  color: #166534;
  font-size: 12px;
  line-height: 1.4;
}
.webgis-admin-status.error {
  border-color: #fecaca;
  background: #fff1f2;
  color: #991b1b;
}
.webgis-admin-layer-list {
  display: grid;
  gap: 8px;
}
.webgis-admin-layer-card {
  display: grid;
  gap: 7px;
  border: 1px solid #dbe7f3;
  border-radius: 10px;
  padding: 9px;
  background: #f8fbff;
}
.webgis-admin-layer-title {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 8px;
  color: #0f2f57;
  font-size: 13px;
  font-weight: 800;
}
.webgis-admin-layer-actions {
  display: inline-flex;
  align-items: center;
  gap: 7px;
  flex-shrink: 0;
}
.webgis-admin-layer-actions button {
  height: 28px;
  border-color: #fecaca;
  background: #fff1f2;
  color: #991b1b;
}
.webgis-admin-layer-grid {
  display: grid;
  grid-template-columns: repeat(2, minmax(0, 1fr));
  gap: 7px;
}
.webgis-admin-layer-grid label {
  display: grid;
  gap: 4px;
  color: #334155;
  font-size: 11px;
  font-weight: 700;
}
.webgis-admin-layer-grid label.switch-line {
  display: flex;
  align-items: center;
  gap: 6px;
}
.webgis-admin-layer-grid input[type="number"],
.webgis-admin-layer-grid input[type="text"] {
  width: 100%;
}
.webgis-field-config,
.webgis-field-config-empty {
  display: grid;
  gap: 7px;
  border-top: 1px dashed #cbdcec;
  padding-top: 8px;
}
.webgis-field-config-head {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 8px;
  color: #0f2f57;
  font-size: 12px;
  font-weight: 800;
}
.webgis-field-config-head button,
.webgis-field-config-empty button {
  height: 28px;
  padding: 0 9px;
}
.webgis-field-list {
  display: grid;
  grid-template-columns: repeat(2, minmax(0, 1fr));
  gap: 5px 8px;
  max-height: 132px;
  overflow: auto;
}
.webgis-field-list label {
  display: flex;
  align-items: center;
  gap: 5px;
  min-width: 0;
  color: #334155;
  font-size: 11px;
  font-weight: 700;
}
.webgis-field-list label > span:first-of-type {
  min-width: 0;
  flex: 1;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
}
.webgis-field-order-actions {
  display: inline-flex;
  gap: 2px;
  margin-left: auto;
}
.webgis-field-order-actions button {
  width: 22px;
  height: 22px;
  padding: 0;
  border-radius: 7px;
  font-size: 12px;
  line-height: 1;
}
.webgis-field-order-actions button:disabled {
  opacity: 0.35;
  cursor: not-allowed;
}
.webgis-attr-panel {
  margin: 0 10px 10px;
  border: 1px solid #cbdcec;
  border-radius: 8px;
  background: #fff;
  box-shadow: 0 14px 32px rgba(15,47,87,0.12);
}
.webgis-attr-tools {
  padding: 10px 12px;
  border-bottom: 1px solid #e4edf6;
}
.webgis-attr-wrap {
  max-height: 260px;
  overflow: auto;
}
.webgis-attr-table {
  width: 100%;
  border-collapse: collapse;
  min-width: 900px;
  font-size: 12px;
}
.webgis-attr-table th,
.webgis-attr-table td {
  padding: 8px;
  border: 1px solid #e4edf6;
  background: #fff;
  text-align: left;
  white-space: nowrap;
}
.webgis-attr-table th {
  position: sticky;
  top: 0;
  z-index: 1;
  background: #eaf4ff;
  color: #0f2f57;
  cursor: pointer;
}
.webgis-attr-field-actions {
  display: inline-flex;
  gap: 3px;
  margin-left: 8px;
  vertical-align: middle;
}
.webgis-attr-field-actions button {
  width: 22px;
  height: 22px;
  padding: 0;
  border-radius: 7px;
  font-size: 12px;
  line-height: 1;
}
.webgis-attr-field-actions button:disabled {
  opacity: 0.35;
  cursor: not-allowed;
}
.webgis-attr-table tr:hover td,
.webgis-attr-table tr.selected td {
  background: #fff7ed;
}
.webgis-popup {
  min-width: 220px;
  font-size: 13px;
}
.webgis-popup strong {
  display: block;
  margin-bottom: 6px;
  color: #0f2f57;
}
body.webgis-mode .leaflet-control-layers {
  border-radius: 8px;
  border-color: #b7c8da;
  box-shadow: 0 12px 28px rgba(15,47,87,0.15);
}
@media print {
  body.webgis-mode .appbar,
  body.webgis-mode .webgis-sidebar,
  body.webgis-mode .webgis-info,
  body.webgis-mode .webgis-topbar,
  body.webgis-mode .webgis-attr-panel,
  body.webgis-mode .webgis-map-tools {
    display: none !important;
  }
  body.webgis-mode .webgis-page,
  body.webgis-mode .webgis-shell,
  body.webgis-mode .webgis-workspace,
  body.webgis-mode .webgis-map-panel,
  body.webgis-mode .webgis-map {
    display: block !important;
    margin: 0 !important;
    width: 100% !important;
    height: 100vh !important;
    min-height: 100vh !important;
    box-shadow: none !important;
  }
}
@media (max-width: 1180px) {
  .webgis-workspace {
    grid-template-columns: 280px minmax(320px, 1fr);
  }
  .webgis-info {
    grid-column: 1 / -1;
    display: grid;
    grid-template-columns: repeat(2, minmax(0, 1fr));
    height: auto;
    max-height: 240px;
  }
}
@media (max-width: 820px) {
  .webgis-topbar {
    grid-template-columns: 1fr;
  }
  .webgis-workspace {
    grid-template-columns: 1fr;
  }
  .webgis-sidebar,
  .webgis-info {
    max-height: 220px;
  }
  .webgis-info {
    display: flex;
  }
}
"""

WEBGIS_HTML = r"""
<main id="webgisPage" class="webgis-page" aria-label="WebGis">
  <section class="webgis-shell">
    <header class="webgis-topbar">
      <div class="webgis-title">
        <strong>WEBGIS QUẢN LÝ DỮ LIỆU ĐẤT ĐAI</strong>
        <span>Hiển thị, tra cứu và quản lý dữ liệu bản đồ hiện trạng/quy hoạch</span>
        <div class="webgis-stats" aria-label="Thống kê dữ liệu WebGIS">
          <span id="webgisLayerCount" class="webgis-stat">0 lớp</span>
          <span id="webgisFeatureCount" class="webgis-stat">0 đối tượng</span>
          <span id="webgisVisibleCount" class="webgis-stat">0 đang bật</span>
        </div>
      </div>
      <div class="webgis-search">
        <input id="webgisSearchInput" type="search" placeholder="Tìm mã thửa, chủ sử dụng, mã đất, địa danh, quy hoạch">
        <button id="webgisSearchBtn" class="primary" type="button">Tìm</button>
        <div id="webgisSearchResults" class="webgis-search-results" hidden></div>
      </div>
      <div class="webgis-actions">
        <span id="webgisSaveStatus" class="webgis-save-status">Chưa nạp dữ liệu</span>
        <button id="webgisAiBtn" type="button">Trợ lý AI</button>
        <button id="webgisOpenTableBtn" type="button">Bảng thuộc tính</button>
        <button id="webgisAdminBtn" class="admin-action" type="button">Quản trị dữ liệu</button>
      </div>
    </header>
    <div class="webgis-workspace">
      <aside class="webgis-sidebar">
        <section class="webgis-panel">
          <div class="webgis-panel-head">
            <h2>Lớp bản đồ</h2>
            <div class="webgis-panel-actions">
              <button id="webgisFitAllBtn" type="button">Toàn bộ</button>
              <button id="webgisToggleSidebarBtn" class="webgis-collapse-btn" type="button" title="Thu gọn cột lớp bản đồ" aria-label="Thu gọn cột lớp bản đồ">&lsaquo;</button>
            </div>
          </div>
          <div class="webgis-panel-body">
            <div id="webgisLayerList" class="webgis-layer-list"></div>
          </div>
        </section>
        <section id="webgisAdminLoginPanel" class="webgis-panel webgis-admin-panel" hidden>
          <div class="webgis-panel-head">
            <h3>&#272;&#259;ng nh&#7853;p admin WebGIS</h3>
            <button id="webgisCloseAdminLoginBtn" type="button">&#272;&#243;ng</button>
          </div>
          <div class="webgis-panel-body webgis-admin-grid">
            <p class="webgis-admin-note">T&#224;i kho&#7843;n n&#224;y ch&#7881; d&#224;nh cho qu&#7843;n tr&#7883; WebGIS: upload GeoJSON, s&#7917;a thu&#7897;c t&#237;nh v&#224; l&#432;u d&#7919; li&#7879;u b&#7843;n &#273;&#7891;.</p>
            <label>T&#224;i kho&#7843;n
              <input id="webgisAdminUsername" type="text" autocomplete="username">
            </label>
            <label>M&#7853;t kh&#7849;u
              <input id="webgisAdminPassword" type="password" autocomplete="current-password">
            </label>
            <button id="webgisAdminLoginSubmit" class="primary" type="button">&#272;&#259;ng nh&#7853;p</button>
            <p id="webgisAdminLoginStatus" class="webgis-admin-status" hidden></p>
          </div>
        </section>
        <section id="webgisAdminPanel" class="webgis-panel webgis-admin-panel" hidden>
          <div class="webgis-panel-head">
            <h3>Quản trị dữ liệu</h3>
            <div class="webgis-admin-head-actions">
              <span id="webgisAdminSessionBadge" class="webgis-admin-session" hidden>Admin</span>
              <button id="webgisAdminLogoutBtn" type="button">&#272;&#259;ng xu&#7845;t</button>
              <button id="webgisCloseAdminBtn" type="button">Đóng</button>
            </div>
          </div>
          <div class="webgis-panel-body webgis-admin-grid">
            <p class="webgis-admin-note">Bản demo xử lý GeoJSON trên trình duyệt. Khi nâng cấp backend có thể lưu vào PostGIS và cấp quyền admin.</p>
            <label>Tên layer mới
              <input id="webgisNewLayerName" type="text" placeholder="Ví dụ: Quy hoạch khu dân cư">
            </label>
            <label>Màu ký hiệu
              <input id="webgisNewLayerColor" type="color" value="#2563eb">
            </label>
            <div>
              <strong>Quan ly hien thi layer</strong>
              <div id="webgisAdminLayerList" class="webgis-admin-layer-list"></div>
            </div>
            <label>Nhom layer
              <input id="webgisNewLayerCategory" type="text" placeholder="Vi du: Quy hoach">
            </label>
            <label>File GeoJSON
              <input id="webgisImportInput" type="file" accept=".geojson,.json,application/geo+json,application/json">
            </label>
            <button id="webgisImportBtn" class="primary" type="button">Thêm layer GeoJSON</button>
            <label>Thuộc tính đối tượng đang chọn
              <textarea id="webgisFeatureEditor" placeholder="Chọn một đối tượng trên bản đồ để sửa thuộc tính JSON"></textarea>
            </label>
            <button id="webgisSaveFeatureBtn" type="button">Lưu thuộc tính</button>
          </div>
        </section>
      </aside>
      <section class="webgis-map-panel">
        <div id="webgisMap" class="webgis-map" role="application" aria-label="Bản đồ WebGIS"></div>
        <div class="webgis-map-tools" aria-label="Công cụ bản đồ">
          <span class="webgis-tool-group"><button id="webgisLocateBtn" type="button">Vị trí</button></span>
          <span class="webgis-tool-divider" aria-hidden="true"></span>
          <span class="webgis-tool-group">
            <button id="webgisMeasureDistanceBtn" type="button">Đo dài</button>
            <button id="webgisMeasureAreaBtn" type="button">Đo diện tích</button>
            <button id="webgisClearMeasureBtn" type="button">Xóa đo</button>
          </span>
          <span class="webgis-tool-divider" aria-hidden="true"></span>
          <span class="webgis-tool-group">
            <button id="webgisPrintBtn" type="button">In</button>
            <button id="webgisShotBtn" type="button">Chụp ảnh</button>
          </span>
          <span class="webgis-tool-divider" aria-hidden="true"></span>
          <span class="webgis-tool-group"><button id="webgisFullscreenBtn" type="button">Toàn màn hình</button></span>
        </div>
        <div id="webgisMeasureBadge" class="webgis-measure-badge">Sẵn sàng tra cứu bản đồ</div>
        <div id="webgisCoordinateBar" class="webgis-coordinate-bar">Tọa độ: --</div>
      </section>
      <aside class="webgis-info">
        <section class="webgis-panel">
          <div class="webgis-panel-head">
            <h2>Thông tin đối tượng</h2>
            <button id="webgisToggleInfoBtn" class="webgis-collapse-btn" type="button" title="Thu gọn cột thông tin" aria-label="Thu gọn cột thông tin">&rsaquo;</button>
          </div>
          <div id="webgisFeatureDetail" class="webgis-panel-body">
            <div class="webgis-detail-empty">
              <div class="webgis-empty-card">
                <div class="webgis-empty-icon">◎</div>
                <strong>Chưa chọn đối tượng bản đồ</strong>
                <ul>
                  <li>Bấm vào thửa đất, vùng quy hoạch, tuyến hoặc điểm công trình để xem thông tin.</li>
                  <li>Dùng ô tìm kiếm để tra theo mã thửa, mã đất, chủ sử dụng hoặc địa danh.</li>
                </ul>
              </div>
            </div>
          </div>
        </section>
        <section class="webgis-panel">
          <div class="webgis-panel-head">
            <h3>Hướng dẫn nhanh</h3>
          </div>
          <div class="webgis-panel-body">
            <p class="webgis-admin-note">Bật/tắt lớp ở sidebar, dùng thanh trong suốt để so sánh nền bản đồ và dữ liệu. Ô tìm kiếm hỗ trợ mã thửa, mã đất, chủ sử dụng, địa danh và loại quy hoạch.</p>
          </div>
        </section>
      </aside>
    </div>
    <section id="webgisAttributePanel" class="webgis-attr-panel" hidden>
      <div class="webgis-attr-tools">
        <strong>Bảng thuộc tính</strong>
        <select id="webgisAttrLayer"></select>
        <input id="webgisAttrSearch" type="search" placeholder="Lọc thuộc tính">
        <button id="webgisCloseTableBtn" type="button">Đóng bảng</button>
      </div>
      <div class="webgis-attr-wrap">
        <table id="webgisAttrTable" class="webgis-attr-table"></table>
      </div>
    </section>
  </section>
</main>
"""

WEBGIS_JS = r"""
const webgisLayerDefs = [
  { id: 'administrative', label: 'Ranh giới hành chính', color: '#2563eb', visible: true },
  { id: 'landuse', label: 'Hiện trạng sử dụng đất', color: '#22c55e', visible: true },
  { id: 'planning', label: 'Quy hoạch sử dụng đất', color: '#f59e0b', visible: true },
  { id: 'roads', label: 'Giao thông', color: '#6b7280', visible: true },
  { id: 'water', label: 'Thủy hệ', color: '#0ea5e9', visible: true },
  { id: 'parcels', label: 'Thửa đất', color: '#ef4444', visible: true },
  { id: 'public', label: 'Công trình công cộng', color: '#8b5cf6', visible: true }
];

// Bang mau duoc trich tu bo ky hieu TT08: Chucnangsudungdat.style.
const webgisLandColors = {
  LUC: '#ffff00',
  LUK: '#ffff00',
  HNK: '#f5a623',
  CLN: '#f5a623',
  RDD: '#2ca25f',
  RPH: '#2ca25f',
  RSX: '#2ca25f',
  NTS: '#1e88e5',
  CNT: '#ffff00',
  LMU: '#ffffff',
  NKH: '#2ca25f',
  ONT: '#f4a3c1',
  ODT: '#f4a3c1',
  TSC: '#ff0000',
  CQP: '#ff0000',
  CAN: '#ff0000',
  DVH: '#ff0000',
  DXH: '#ff0000',
  DYT: '#ff0000',
  DGD: '#ff0000',
  DTT: '#ff0000',
  DKH: '#ff0000',
  DMT: '#ff0000',
  DKT: '#ff0000',
  DNG: '#ff0000',
  DSK: '#ff0000',
  SCC: '#ff0000',
  SKK: '#ff0000',
  SKN: '#ff0000',
  SCT: '#ff0000',
  TMD: '#ff0000',
  SKC: '#ff0000',
  SKS: '#ff66b3',
  DGT: '#f5a623',
  DTL: '#1e88e5',
  DPC: '#ff0000',
  DDD: '#ff0000',
  DRA: '#ffffff',
  DNL: '#ff0000',
  DBV: '#ff0000',
  DCH: '#ff0000',
  DKV: '#ff0000',
  TON: '#ff0000',
  TIN: '#ff0000',
  NTD: '#9ca3af',
  MNC: '#1e88e5',
  SON: '#1e88e5',
  PNK: '#ff0000',
  CGT: '#ffffff',
  BCS: '#ffffff',
  DCS: '#ffffff',
  NCS: '#ffff00',
  MCS: '#1e88e5'
};

const webgisLandCodeFields = [
  'loai_dat',
  'loaidat',
  'Loaidat',
  'loaiDat',
  'LOAI_DAT',
  'LOAIDAT',
  'ma_loai_dat',
  'maloaidat',
  'MaLoaiDat',
  'ma_dat',
  'madat',
  'MaDat',
  'ky_hieu',
  'kyhieu',
  'KyHieu',
  'muc_dich_sd',
  'MDSD',
  'mdsd'
];

let webgisState = {
  initialized: false,
  initializing: null,
  map: null,
  layerDefs: [],
  overlayLayers: new Map(),
  featureLayers: new Map(),
  featureCache: new Map(),
  loadedLayerIds: new Set(),
  features: [],
  selectedFeatureId: null,
  selectedVector: null,
  attrSortKey: '',
  attrSortDir: 1,
  measureMode: null,
  measurePoints: [],
  measureLayer: null,
  resizeObserver: null,
  deletedLayerIds: new Set(),
  layerPatchTimers: new Map()
};

const webgisStorageKey = 'webgis-state-v1';
const webgisProjectId = 'webgis-default';
const webgisApiBase = '/api/webgis';
const webgisAdminTokenKey = 'webgis-admin-token';
const webgisAdminUserKey = 'webgis-admin-user';
let webgisSaveTimer = 0;
let webgisAdminToken = localStorage.getItem(webgisAdminTokenKey) || '';
let webgisAdminUser = localStorage.getItem(webgisAdminUserKey) || '';

function webgisEl(id) {
  return document.getElementById(id);
}

function webgisInvalidateSize(delay = 80) {
  window.setTimeout(() => webgisState.map?.invalidateSize(), delay);
}

function webgisEscape(value) {
  return String(value ?? '').replace(/[&<>"']/g, ch => ({
    '&': '&amp;',
    '<': '&lt;',
    '>': '&gt;',
    '"': '&quot;',
    "'": '&#039;'
  }[ch]));
}

function webgisFeatureTitle(feature) {
  const props = feature.properties || {};
  const def = webgisLayerDefById(props.layer);
  const category = String(def?.category || '').trim();
  if (category && category.toLowerCase() !== 'chung') return category;
  return webgisLayerLabel(props.layer) || 'Doi tuong ban do';
}

function webgisLayerLabel(id) {
  return webgisState.layerDefs.find(layer => layer.id === id)?.label || id || 'Khác';
}

function webgisNormalizeLandCode(value) {
  const text = String(value ?? '').trim().toUpperCase();
  if (!text) return '';
  const compact = text.replace(/\s+/g, '');
  if (webgisLandColors[compact]) return compact;
  const match = text.match(/\b[A-Z]{2,4}\b/);
  return match && webgisLandColors[match[0]] ? match[0] : '';
}

function webgisFeatureLandCode(feature) {
  const props = feature?.properties || {};
  for (const key of webgisLandCodeFields) {
    const code = webgisNormalizeLandCode(props[key]);
    if (code) return code;
  }
  const lowerEntries = Object.entries(props).map(([key, value]) => [String(key).toLowerCase().replace(/[_\s-]/g, ''), value]);
  for (const [key, value] of lowerEntries) {
    if (!/(loaidat|maloaidat|madat|kyhieu|mdsd)/.test(key)) continue;
    const code = webgisNormalizeLandCode(value);
    if (code) return code;
  }
  for (const value of Object.values(props)) {
    const code = webgisNormalizeLandCode(value);
    if (code) return code;
  }
  return '';
}

function webgisLayerColor(id, feature) {
  const code = webgisFeatureLandCode(feature);
  return webgisLandColors[code] || webgisState.layerDefs.find(layer => layer.id === id)?.color || '#2563eb';
}

function webgisNormalizeOpacity(value) {
  const numeric = Number(value);
  if (!Number.isFinite(numeric)) return 1;
  const normalized = numeric > 1 ? numeric / 100 : numeric;
  return Math.max(0, Math.min(1, normalized));
}

function webgisNormalizeBoolean(value, defaultValue) {
  if (value === undefined || value === null) return defaultValue;
  if (typeof value === 'string') return !['false', '0', 'no', 'off'].includes(value.trim().toLowerCase());
  return Boolean(value);
}

const webgisInternalFields = new Set(['__id', 'layer']);

function webgisNormalizeFieldList(value) {
  if (value === undefined || value === null) return null;
  const values = Array.isArray(value) ? value : String(value).split(',');
  return Array.from(new Set(values.map(item => String(item || '').trim()).filter(Boolean)));
}

function webgisLayerSort(a, b) {
  return Number(a.sort_order || 0) - Number(b.sort_order || 0) || String(a.label || '').localeCompare(String(b.label || ''), 'vi');
}

function webgisEmptyDetailHtml() {
  return `
    <div class="webgis-detail-empty">
      <div class="webgis-empty-card">
        <div class="webgis-empty-icon">◎</div>
        <strong>Chưa chọn đối tượng bản đồ</strong>
        <ul>
          <li>Bấm vào thửa đất, vùng quy hoạch, tuyến hoặc điểm công trình để xem thông tin.</li>
          <li>Dùng ô tìm kiếm để tra theo mã thửa, mã đất, chủ sử dụng hoặc địa danh.</li>
        </ul>
      </div>
    </div>
  `;
}

function webgisLayerFeatureCount(layerId) {
  const cached = webgisState.featureCache.get(layerId);
  if (cached) return cached.length;
  const def = webgisState.layerDefs.find(layer => layer.id === layerId);
  return Number(def?.feature_count || 0);
}

function webgisUpdateStats() {
  const publicLayers = webgisState.layerDefs.filter(def => def.is_public !== false);
  const layerCount = publicLayers.length;
  const featureCount = webgisState.layerDefs.reduce((sum, def) => sum + webgisLayerFeatureCount(def.id), 0);
  const visibleFeatureCount = publicLayers
    .filter(def => def.visible === true)
    .reduce((sum, def) => sum + webgisLayerFeatureCount(def.id), 0);
  const layerEl = webgisEl('webgisLayerCount');
  const featureEl = webgisEl('webgisFeatureCount');
  const visibleEl = webgisEl('webgisVisibleCount');
  if (layerEl) layerEl.textContent = `${layerCount} lớp`;
  if (featureEl) featureEl.textContent = `${featureCount} đối tượng`;
  if (visibleEl) visibleEl.textContent = `${visibleFeatureCount} đang bật`;
}

function webgisNormalizeFeatures(collection, defaultLayer = '') {
  const features = Array.isArray(collection?.features) ? collection.features : [];
  return features
    .filter(feature => feature && feature.geometry)
    .map((feature, index) => {
      const props = { ...(feature.properties || {}) };
      props.layer = props.layer || defaultLayer || 'imported';
      props.__id = props.__id || `${props.layer}-${Date.now()}-${index}-${Math.random().toString(36).slice(2, 7)}`;
      return { ...feature, properties: props };
    });
}

function webgisNormalizeLayerDefs(savedDefs = [], deletedLayerIds = []) {
  const deletedSet = new Set((Array.isArray(deletedLayerIds) ? deletedLayerIds : []).map(id => String(id || '').trim()).filter(Boolean));
  webgisState.deletedLayerIds = new Set(deletedSet);
  const savedById = new Map(
    (Array.isArray(savedDefs) ? savedDefs : [])
      .filter(def => def && def.id && !deletedSet.has(String(def.id)))
      .map(def => [String(def.id), def])
  );
  const defaultIds = new Set(webgisLayerDefs.map(def => def.id));
  const defaults = webgisLayerDefs.filter(def => !deletedSet.has(def.id)).map(def => {
    const saved = savedById.get(def.id) || {};
    const hasDefaultVisible = Object.prototype.hasOwnProperty.call(saved, 'default_visible');
    const hasLegacyVisible = Object.prototype.hasOwnProperty.call(saved, 'visible');
    const defaultVisible = hasDefaultVisible ? webgisNormalizeBoolean(saved.default_visible, false) : (hasLegacyVisible ? saved.visible !== false : false);
    return {
      ...def,
      ...saved,
      id: def.id,
      label: saved.label || def.label,
      color: saved.color || def.color,
      is_public: webgisNormalizeBoolean(saved.is_public, true),
      default_visible: defaultVisible,
      allow_user_toggle: webgisNormalizeBoolean(saved.allow_user_toggle, true),
      opacity: webgisNormalizeOpacity(saved.opacity ?? 1),
      sort_order: Number.isFinite(Number(saved.sort_order)) ? Number(saved.sort_order) : webgisLayerDefs.findIndex(item => item.id === def.id) + 1,
      category: String(saved.category || def.category || 'Chung'),
      visible_fields: webgisNormalizeFieldList(saved.visible_fields),
      visible: defaultVisible,
      feature_count: Number(saved.feature_count || 0)
    };
  });
  const custom = Array.from(savedById.values())
    .filter(def => !defaultIds.has(String(def.id)))
    .map(def => ({
      id: String(def.id),
      label: String(def.label || def.id),
      color: String(def.color || '#2563eb'),
      is_public: webgisNormalizeBoolean(def.is_public, true),
      default_visible: webgisNormalizeBoolean(def.default_visible, def.visible === true),
      allow_user_toggle: webgisNormalizeBoolean(def.allow_user_toggle, true),
      visible: webgisNormalizeBoolean(def.default_visible, def.visible === true),
      opacity: webgisNormalizeOpacity(def.opacity ?? 1),
      sort_order: Number.isFinite(Number(def.sort_order)) ? Number(def.sort_order) : defaults.length + 1,
      category: String(def.category || 'Chung'),
      visible_fields: webgisNormalizeFieldList(def.visible_fields),
      feature_count: Number(def.feature_count || 0),
      custom: true
    }));
  return [...defaults, ...custom].sort(webgisLayerSort);
}

function webgisAllCachedFeatures() {
  const byId = new Map();
  webgisState.features.forEach(feature => byId.set(feature.properties?.__id || JSON.stringify(feature.geometry), feature));
  webgisState.featureCache.forEach(features => {
    features.forEach(feature => byId.set(feature.properties?.__id || JSON.stringify(feature.geometry), feature));
  });
  return Array.from(byId.values()).filter(feature => !webgisState.deletedLayerIds.has(String(feature.properties?.layer || '')));
}

function webgisStatePayload() {
  return {
    version: 2,
    savedAt: new Date().toISOString(),
    deletedLayerIds: Array.from(webgisState.deletedLayerIds),
    layerDefs: webgisState.layerDefs.map(def => ({
      id: def.id,
      label: def.label,
      color: def.color,
      is_public: def.is_public !== false,
      default_visible: def.default_visible === true,
      allow_user_toggle: def.allow_user_toggle !== false,
      opacity: webgisNormalizeOpacity(def.opacity),
      sort_order: Number(def.sort_order || 0),
      category: def.category || 'Chung',
      visible_fields: webgisNormalizeFieldList(def.visible_fields),
      feature_count: webgisLayerFeatureCount(def.id),
      visible: def.visible === true,
      custom: Boolean(def.custom)
    })),
    // Du lieu hinh hoc GeoJSON duoc luu rieng theo tung layer de tranh qua tai bo nho backend.
    features: []
  };
}

function webgisValidPayload(data) {
  return data && typeof data === 'object' && Array.isArray(data.layerDefs) && (!data.features || Array.isArray(data.features));
}

function webgisSetSaveStatus(text, isError = false) {
  const status = webgisEl('webgisSaveStatus');
  if (!status) return;
  status.textContent = text;
  status.classList.toggle('error', Boolean(isError));
  status.classList.toggle('connected', !isError && /kết nối|Supabase|dữ liệu đã lưu|nạp dữ liệu/i.test(text));
}

function webgisRestoreConnectedStatus() {
  const status = webgisEl('webgisSaveStatus');
  if (status?.classList.contains('error')) return;
  webgisSetSaveStatus('Đã kết nối dữ liệu');
}

function webgisSetPanelCollapsed(panel, collapsed) {
  const page = webgisEl('webgisPage');
  if (!page) return;
  const className = panel === 'layers' ? 'layers-collapsed' : 'info-collapsed';
  page.classList.toggle(className, Boolean(collapsed));
  const button = panel === 'layers' ? webgisEl('webgisToggleSidebarBtn') : webgisEl('webgisToggleInfoBtn');
  if (button) {
    const isCollapsed = page.classList.contains(className);
    button.innerHTML = panel === 'layers' ? (isCollapsed ? '&rsaquo;' : '&lsaquo;') : (isCollapsed ? '&lsaquo;' : '&rsaquo;');
    button.setAttribute('aria-expanded', String(!isCollapsed));
  }
  webgisInvalidateSize(120);
}

function webgisExportSelectedFeatureInfo() {
  const feature = webgisAllCachedFeatures().find(item => item.properties.__id === webgisState.selectedFeatureId);
  if (!feature) return;
  const rows = webgisVisiblePropertyEntries(feature);
  const title = webgisFeatureTitle(feature);
  const text = [
    title,
    '',
    ...rows.map(([key, value]) => `${key}: ${value}`)
  ].join('\n');
  const blob = new Blob([text], { type: 'text/plain;charset=utf-8' });
  const link = document.createElement('a');
  link.href = URL.createObjectURL(blob);
  link.download = `${String(title || 'webgis-thong-tin').replace(/[^\w\-]+/g, '_')}.txt`;
  link.click();
  URL.revokeObjectURL(link.href);
}

function webgisAdminHeaders() {
  return webgisAdminToken ? { Authorization: `Bearer ${webgisAdminToken}` } : {};
}

function webgisSetAdminStatus(text, isError = false) {
  const status = webgisEl('webgisAdminLoginStatus');
  if (!status) return;
  status.hidden = !text;
  status.textContent = text || '';
  status.classList.toggle('error', Boolean(isError));
}

function webgisUpdateAdminUi() {
  const logged = Boolean(webgisAdminToken);
  const adminPanel = webgisEl('webgisAdminPanel');
  const loginPanel = webgisEl('webgisAdminLoginPanel');
  const adminBtn = webgisEl('webgisAdminBtn');
  const badge = webgisEl('webgisAdminSessionBadge');
  if (adminBtn) adminBtn.textContent = logged ? 'Quản trị dữ liệu' : 'Đăng nhập admin';
  if (badge) {
    badge.hidden = !logged;
    badge.textContent = logged ? `Admin${webgisAdminUser ? ': ' + webgisAdminUser : ''}` : 'Admin';
  }
  if (!logged && adminPanel) adminPanel.hidden = true;
  if (logged && loginPanel) loginPanel.hidden = true;
}

function webgisShowAdminLogin(message = '') {
  const panel = webgisEl('webgisAdminLoginPanel');
  if (!panel) return;
  panel.hidden = false;
  webgisEl('webgisAdminPanel').hidden = true;
  webgisSetAdminStatus(message, Boolean(message));
  webgisInvalidateSize(80);
  window.setTimeout(() => webgisEl('webgisAdminUsername')?.focus(), 60);
}

function webgisRequireAdmin(message = 'Vui lòng đăng nhập admin WebGIS để quản trị dữ liệu.') {
  if (webgisAdminToken) return true;
  webgisShowAdminLogin(message);
  return false;
}

function webgisSetAdminSession(payload) {
  webgisAdminToken = payload.token || '';
  webgisAdminUser = payload.username || '';
  if (webgisAdminToken) {
    localStorage.setItem(webgisAdminTokenKey, webgisAdminToken);
    localStorage.setItem(webgisAdminUserKey, webgisAdminUser);
  }
  webgisUpdateAdminUi();
}

function webgisClearAdminSession() {
  webgisAdminToken = '';
  webgisAdminUser = '';
  localStorage.removeItem(webgisAdminTokenKey);
  localStorage.removeItem(webgisAdminUserKey);
  webgisUpdateAdminUi();
}

async function webgisLoginAdmin() {
  const username = webgisEl('webgisAdminUsername').value.trim();
  const password = webgisEl('webgisAdminPassword').value;
  if (!username || !password) {
    webgisSetAdminStatus('Vui lòng nhập tài khoản và mật khẩu admin WebGIS.', true);
    return;
  }
  webgisSetAdminStatus('Đang đăng nhập...');
  const response = await fetch('/api/webgis/admin/login', {
    method: 'POST',
    headers: { 'Content-Type': 'application/json' },
    body: JSON.stringify({ username, password })
  });
  const payload = await response.json().catch(() => ({}));
  if (!response.ok) {
    webgisSetAdminStatus(payload.error || 'Không đăng nhập được admin WebGIS.', true);
    return;
  }
  webgisSetAdminSession(payload);
  const loadedLayerIds = Array.from(webgisState.loadedLayerIds);
  loadedLayerIds.forEach(layerId => {
    webgisState.loadedLayerIds.delete(layerId);
    webgisState.featureCache.delete(layerId);
  });
  await Promise.all(loadedLayerIds.map(layerId => webgisEnsureLayerLoaded(layerId)));
  webgisRebuildOverlays();
  webgisEl('webgisAdminPassword').value = '';
  webgisEl('webgisAdminPanel').hidden = false;
  webgisSetAdminStatus('Đã đăng nhập admin WebGIS thành công.');
  webgisSetSaveStatus('Admin WebGIS đã đăng nhập');
  webgisInvalidateSize(80);
}

function webgisSaveLocal(data) {
  try {
    localStorage.setItem(webgisStorageKey, JSON.stringify(data));
    return true;
  } catch (error) {
    return false;
  }
}

function webgisLoadLocal() {
  try {
    const raw = localStorage.getItem(webgisStorageKey);
    if (!raw) return null;
    const data = JSON.parse(raw);
    return webgisValidPayload(data) ? data : null;
  } catch (error) {
    return null;
  }
}

function webgisSetLayerFeatureCache(layerId, features) {
  const normalizedLayerId = String(layerId || '');
  if (webgisState.deletedLayerIds.has(normalizedLayerId)) return [];
  const normalized = webgisNormalizeFeatures({ type: 'FeatureCollection', features }, normalizedLayerId)
    .map(feature => ({ ...feature, properties: { ...feature.properties, layer: normalizedLayerId } }));
  webgisState.featureCache.set(normalizedLayerId, normalized);
  webgisState.loadedLayerIds.add(normalizedLayerId);
  const otherLoaded = webgisState.features.filter(feature => String(feature.properties?.layer || '') !== normalizedLayerId);
  webgisState.features = [...otherLoaded, ...normalized];
  const def = webgisState.layerDefs.find(layer => layer.id === normalizedLayerId);
  if (def) def.feature_count = normalized.length;
  return normalized;
}

function webgisPrimeFeatureCache(features) {
  const grouped = new Map();
  webgisNormalizeFeatures({ type: 'FeatureCollection', features: features || [] }).forEach(feature => {
    const layerId = String(feature.properties?.layer || 'imported');
    if (webgisState.deletedLayerIds.has(layerId)) return;
    if (!grouped.has(layerId)) grouped.set(layerId, []);
    grouped.get(layerId).push(feature);
  });
  grouped.forEach((items, layerId) => webgisSetLayerFeatureCache(layerId, items));
}

async function webgisEnsureLayerLoaded(layerId) {
  const normalizedLayerId = String(layerId || '');
  if (webgisState.loadedLayerIds.has(normalizedLayerId)) return webgisState.featureCache.get(normalizedLayerId) || [];
  try {
    const response = await fetch(`${webgisApiBase}/${encodeURIComponent(webgisProjectId)}/layers/${encodeURIComponent(normalizedLayerId)}/features`, {
      cache: 'no-store',
      headers: webgisAdminHeaders()
    });
    if (!response.ok) throw new Error(await response.text());
    const payload = await response.json();
    return webgisSetLayerFeatureCache(normalizedLayerId, payload.features || []);
  } catch (error) {
    webgisSetLayerFeatureCache(normalizedLayerId, []);
    webgisSetSaveStatus(`Khong nap duoc du lieu layer ${normalizedLayerId}`, true);
    return [];
  }
}

async function webgisLoadSavedData() {
  const localData = webgisLoadLocal();
  try {
    const response = await fetch(`${webgisApiBase}/${encodeURIComponent(webgisProjectId)}?metadata=1`, {
      cache: 'no-store',
      headers: webgisAdminHeaders()
    });
    if (response.ok) {
      const payload = await response.json();
      if (webgisValidPayload(payload.data)) {
        if (Array.isArray(payload.data.features) && payload.data.features.length) webgisSaveLocal(payload.data);
        webgisSetSaveStatus(payload.storage === 'supabase' || payload.storage === 'supabase-migrated' ? 'Đã kết nối dữ liệu' : 'Đã nạp dữ liệu đã lưu');
        return payload.data;
      }
    }
    if (response.status !== 404) throw new Error('Không nạp được dữ liệu WebGIS từ server.');
  } catch (error) {
    if (localData) webgisSetSaveStatus('Đang dùng bản lưu tạm', true);
  }
  if (localData) return localData;
  webgisSetSaveStatus('Đang dùng dữ liệu mẫu');
  return null;
}

async function webgisSaveLayerFeatures(layerId) {
  if (!webgisAdminToken) return null;
  const normalizedLayerId = String(layerId || '');
  if (!normalizedLayerId) return null;
  const cached = webgisState.featureCache.get(normalizedLayerId);
  const features = cached || webgisAllCachedFeatures().filter(feature => String(feature.properties?.layer || '') === normalizedLayerId);
  const response = await fetch(`${webgisApiBase}/${encodeURIComponent(webgisProjectId)}/layers/${encodeURIComponent(normalizedLayerId)}/features`, {
    method: 'PUT',
    headers: { 'Content-Type': 'application/json', ...webgisAdminHeaders() },
    body: JSON.stringify({ features })
  });
  const payload = await response.json().catch(() => ({}));
  if (response.status === 401) {
    webgisClearAdminSession();
    webgisSetSaveStatus('Phien admin WebGIS da het han. Dang nhap admin de luu layer.', true);
    return null;
  }
  if (!response.ok) throw new Error(payload.error || 'Khong luu duoc du lieu layer WebGIS.');
  const def = webgisState.layerDefs.find(layer => layer.id === normalizedLayerId);
  if (def) def.feature_count = Number(payload.feature_count ?? features.length);
  webgisSetSaveStatus('Da luu du lieu layer');
  return payload;
}

async function webgisSaveNow() {
  if (!webgisState.initialized) return;
  if (!webgisAdminToken) return;
  const data = webgisStatePayload();
  const savedLocal = webgisSaveLocal(data);
  webgisSetSaveStatus('Đang tự lưu...');
  try {
    const response = await fetch(`${webgisApiBase}/${encodeURIComponent(webgisProjectId)}`, {
      method: 'PUT',
      headers: { 'Content-Type': 'application/json', ...webgisAdminHeaders() },
      body: JSON.stringify({ data })
    });
    if (response.status === 401) {
      webgisClearAdminSession();
      webgisSetSaveStatus('Phiên admin WebGIS đã hết hạn. Bấm Đăng nhập admin khi cần quản trị.', true);
      return;
    }
    if (!response.ok) throw new Error(await response.text());
    const result = await response.json().catch(() => ({}));
    const target = result.storage === 'supabase' ? 'Supabase' : 'server';
    webgisSetSaveStatus(`Đã tự lưu ${target} ${new Date().toLocaleTimeString('vi-VN', { hour: '2-digit', minute: '2-digit' })}`);
  } catch (error) {
    webgisSetSaveStatus(savedLocal ? 'Đã lưu tạm trên trình duyệt' : 'Không lưu được dữ liệu', true);
  }
}

function webgisScheduleSave() {
  if (!webgisState.initialized) return;
  if (!webgisAdminToken) return;
  clearTimeout(webgisSaveTimer);
  webgisSaveTimer = setTimeout(() => webgisSaveNow(), 600);
}

function webgisStyle(feature) {
  const layerId = feature?.properties?.layer;
  const color = webgisLayerColor(layerId, feature);
  const layerDef = webgisState.layerDefs.find(def => def.id === layerId);
  const opacity = webgisNormalizeOpacity(layerDef?.opacity ?? webgisEl(`webgisOpacity_${layerId}`)?.value ?? 1);
  const isLine = ['LineString', 'MultiLineString'].includes(feature.geometry?.type);
  return {
    color,
    weight: layerId === 'administrative' ? 3 : isLine ? 4 : 1.6,
    dashArray: layerId === 'administrative' ? '8 5' : '',
    opacity: Math.max(0.15, opacity),
    fillColor: color,
    fillOpacity: isLine ? 0 : Math.min(0.55, opacity * 0.55)
  };
}

function webgisLayerDefById(layerId) {
  return webgisState.layerDefs.find(def => def.id === layerId);
}

function webgisIsVisibleValue(value) {
  return value !== undefined && value !== null && String(value).trim() !== '';
}

const webgisPreferredPropertyOrder = [
  'Quy hoạch',
  'quy_hoach',
  'muc_dich_quy_hoach',
  'Địa điểm quy hoạch',
  'Dia diem quy hoach',
  'Mã quy hoạch',
  'ma_quy_hoach',
  'Loaidat',
  'loai_dat',
  'Mã đất',
  'ma_dat',
  'Diện tích',
  'dien_tich',
  'SHAPE_Area',
  'Shape_Area',
  'shape_area',
  'ma_thua',
  'ma_khoanh',
  'OBJECTID',
  'objectid',
  'STT',
  'stt',
  'ghi_chu'
];

function webgisOrderPropertyKeys(keys, configuredOrder = null) {
  const unique = Array.from(new Set((keys || []).filter(Boolean)));
  if (Array.isArray(configuredOrder)) {
    const allowed = new Set(unique);
    return [
      ...configuredOrder.filter(key => allowed.has(key)),
      ...unique.filter(key => !configuredOrder.includes(key))
    ];
  }
  const priority = new Map(webgisPreferredPropertyOrder.map((key, index) => [key, index]));
  return unique.sort((a, b) => {
    const pa = priority.has(a) ? priority.get(a) : 1000;
    const pb = priority.has(b) ? priority.get(b) : 1000;
    return pa - pb || String(a).localeCompare(String(b), 'vi');
  });
}

function webgisLayerPropertyKeys(layerId) {
  const keys = new Set();
  const def = webgisLayerDefById(layerId);
  const configuredOrder = webgisNormalizeFieldList(def?.visible_fields);
  configuredOrder?.forEach(key => keys.add(key));
  (webgisState.featureCache.get(layerId) || []).forEach(feature => {
    Object.keys(feature.properties || {}).forEach(key => {
      if (!webgisInternalFields.has(key)) keys.add(key);
    });
  });
  return webgisOrderPropertyKeys(Array.from(keys), configuredOrder);
}

function webgisVisiblePropertyEntries(feature, forceAll = false) {
  const props = feature?.properties || {};
  const def = webgisLayerDefById(props.layer);
  const visibleFields = forceAll ? null : webgisNormalizeFieldList(def?.visible_fields);
  const entries = Object.entries(props)
    .filter(([key, value]) => !webgisInternalFields.has(key) && webgisIsVisibleValue(value))
    .filter(([key]) => !Array.isArray(visibleFields) || visibleFields.includes(key));
  if (Array.isArray(visibleFields)) {
    const byKey = new Map(entries);
    return visibleFields.filter(key => byKey.has(key)).map(key => [key, byKey.get(key)]);
  }
  const orderedKeys = webgisOrderPropertyKeys(entries.map(([key]) => key));
  const byKey = new Map(entries);
  return orderedKeys.map(key => [key, byKey.get(key)]);
}

function webgisRenderAdminFieldList(def) {
  const keys = webgisLayerPropertyKeys(def.id);
  const configured = Array.isArray(def.visible_fields);
  const selected = new Set(webgisNormalizeFieldList(def.visible_fields) || keys);
  if (!keys.length) {
    return `
      <div class="webgis-field-config-empty">
        <span>Chua nap danh sach thuoc tinh cua layer nay.</span>
        <button type="button" data-webgis-layer-load-fields="${webgisEscape(def.id)}">Nap thuoc tinh</button>
      </div>
    `;
  }
  return `
    <div class="webgis-field-config">
      <div class="webgis-field-config-head">
        <span>Thuoc tinh hien thi</span>
        <button type="button" data-webgis-layer-fields-all="${webgisEscape(def.id)}">Tat ca</button>
      </div>
      <div class="webgis-field-list">
        ${keys.map((key, index) => `
          <label>
            <input type="checkbox" data-webgis-field-toggle="${webgisEscape(def.id)}" value="${webgisEscape(key)}" ${(!configured || selected.has(key)) ? 'checked' : ''}>
            <span>${webgisEscape(key)}</span>
            <span class="webgis-field-order-actions">
              <button type="button" title="Đưa thuộc tính lên trên" data-webgis-field-move="${webgisEscape(def.id)}" data-field="${webgisEscape(key)}" data-direction="up" ${index === 0 ? 'disabled' : ''}>↑</button>
              <button type="button" title="Đưa thuộc tính xuống dưới" data-webgis-field-move="${webgisEscape(def.id)}" data-field="${webgisEscape(key)}" data-direction="down" ${index === keys.length - 1 ? 'disabled' : ''}>↓</button>
            </span>
          </label>
        `).join('')}
      </div>
    </div>
  `;
}

function webgisMoveLayerField(layerId, field, direction, selectedFields = null) {
  const def = webgisState.layerDefs.find(layer => layer.id === layerId);
  if (!def) return false;
  const keys = webgisLayerPropertyKeys(layerId);
  const index = keys.indexOf(field);
  const target = index + (direction === 'down' ? 1 : -1);
  if (index < 0 || target < 0 || target >= keys.length) return false;
  [keys[index], keys[target]] = [keys[target], keys[index]];
  const selected = selectedFields ? new Set(selectedFields) : new Set(webgisNormalizeFieldList(def.visible_fields) || keys);
  def.visible_fields = keys.filter(key => selected.has(key));
  webgisRefreshSelectedFeatureDisplay();
  if (!webgisEl('webgisAttributePanel')?.hidden) webgisRenderAttributeTable();
  webgisScheduleLayerMetadataPatch(layerId, 120);
  return true;
}

function webgisPopupHtml(feature) {
  const rows = webgisVisiblePropertyEntries(feature);
  return `<div class="webgis-popup"><strong>${webgisEscape(webgisFeatureTitle(feature))}</strong>${
    rows.length
      ? rows.map(([key, value]) => `<div><b>${webgisEscape(key)}:</b> ${webgisEscape(value)}</div>`).join('')
      : '<div>Chua co thuoc tinh duoc phep hien thi.</div>'
  }</div>`;
}

function webgisRenderFeatureDetail(feature) {
  const detail = webgisEl('webgisFeatureDetail');
  if (!feature) {
    detail.innerHTML = webgisEmptyDetailHtml();
    webgisEl('webgisFeatureEditor').value = '';
    return;
  }
  const props = feature.properties || {};
  const rows = webgisVisiblePropertyEntries(feature);
  detail.innerHTML = rows.length
    ? `
      <h3 class="webgis-detail-title">${webgisEscape(webgisFeatureTitle(feature))}</h3>
      <div class="webgis-detail-actions">
        <button type="button" data-webgis-detail-zoom="1">Phóng tới thửa</button>
        <button type="button" data-webgis-detail-export="1">Xuất thông tin</button>
      </div>
      <table class="webgis-detail-table"><tbody>${
      rows.map(([key, value]) => `<tr><th>${webgisEscape(key)}</th><td>${webgisEscape(value)}</td></tr>`).join('')
    }</tbody></table>`
    : '<div class="webgis-detail-empty">Chua co thuoc tinh duoc phep hien thi.</div>';
  const editableRows = Object.entries(props).filter(([key]) => key !== '__id');
  webgisEl('webgisFeatureEditor').value = JSON.stringify(Object.fromEntries(editableRows), null, 2);
}

function webgisSelectFeature(feature, vectorLayer, openPopup = true) {
  if (!feature) return;
  if (webgisState.selectedVector?.setStyle && webgisState.selectedFeatureId) {
    const previousFeature = webgisState.features.find(item => item.properties.__id === webgisState.selectedFeatureId);
    webgisState.selectedVector.setStyle(webgisStyle(previousFeature));
  }
  webgisState.selectedFeatureId = feature.properties.__id;
  webgisState.selectedVector = vectorLayer || webgisState.featureLayers.get(feature.properties.__id);
  if (webgisState.selectedVector?.setStyle) {
    webgisState.selectedVector.setStyle({ color: '#f97316', weight: 4, fillOpacity: 0.48 });
    webgisState.selectedVector.bringToFront?.();
  }
  webgisRenderFeatureDetail(feature);
  if (openPopup && webgisState.selectedVector?.bindPopup) {
    webgisState.selectedVector.bindPopup(webgisPopupHtml(feature)).openPopup();
  }
  webgisHighlightAttrRow(feature.properties.__id);
}

function webgisBuildOverlayLayer(def) {
  const collection = {
    type: 'FeatureCollection',
    features: webgisState.featureCache.get(def.id) || []
  };
  return L.geoJSON(collection, {
    style: webgisStyle,
    pointToLayer(feature, latlng) {
      const color = webgisLayerColor(def.id, feature);
      return L.circleMarker(latlng, { radius: 8, color, weight: 2, fillColor: color, fillOpacity: 0.82 });
    },
    onEachFeature(feature, vectorLayer) {
      webgisState.featureLayers.set(feature.properties.__id, vectorLayer);
      vectorLayer.on('click', () => webgisSelectFeature(feature, vectorLayer));
      vectorLayer.bindTooltip(webgisFeatureTitle(feature), { sticky: true });
    }
  });
}

function webgisRebuildOverlays() {
  if (!webgisState.map) return;
  webgisState.overlayLayers.forEach(layer => webgisState.map.removeLayer(layer));
  webgisState.overlayLayers.clear();
  webgisState.featureLayers.clear();
  webgisState.layerDefs
    .filter(def => def.is_public !== false && def.visible === true && webgisState.loadedLayerIds.has(def.id))
    .forEach(def => {
    const layer = webgisBuildOverlayLayer(def);
    webgisState.overlayLayers.set(def.id, layer);
    layer.addTo(webgisState.map);
  });
  if (webgisState.selectedFeatureId && !webgisAllCachedFeatures().some(feature => feature.properties.__id === webgisState.selectedFeatureId)) {
    webgisState.selectedFeatureId = null;
    webgisState.selectedVector = null;
    webgisRenderFeatureDetail(null);
  }
  webgisRenderLayerList();
  webgisRenderAdminLayerList();
  webgisPopulateAttrLayerSelect();
  webgisUpdateStats();
  webgisInvalidateSize(60);
}

function webgisRenderLayerList() {
  const root = webgisEl('webgisLayerList');
  root.innerHTML = webgisState.layerDefs.filter(def => def.is_public !== false).map(def => `
    <div class="webgis-layer-item" data-layer="${webgisEscape(def.id)}">
      <span class="webgis-symbol" style="background:${webgisEscape(def.color)}"></span>
      <div class="webgis-layer-main">
        <label><input type="checkbox" data-webgis-layer-toggle="${webgisEscape(def.id)}" ${def.visible === true ? 'checked' : ''} ${def.allow_user_toggle === false ? 'disabled' : ''}> ${webgisEscape(def.label)}</label>
        <span class="webgis-layer-count">${webgisLayerFeatureCount(def.id)} đối tượng</span>
      </div>
      <div class="webgis-layer-actions">
        <button class="webgis-icon-btn" type="button" data-webgis-layer-zoom="${webgisEscape(def.id)}" title="Phóng tới layer" aria-label="Phóng tới layer">&#8981;</button>
      </div>
      <div class="webgis-layer-tools">
        <span>Độ mờ</span>
        <input id="webgisOpacity_${webgisEscape(def.id)}" type="range" min="0" max="1" step="0.05" value="${webgisNormalizeOpacity(def.opacity)}" data-webgis-layer-opacity="${webgisEscape(def.id)}">
      </div>
    </div>
  `).join('');
}

function webgisRenderAdminLayerList() {
  const root = webgisEl('webgisAdminLayerList');
  if (!root) return;
  root.innerHTML = webgisState.layerDefs.map(def => `
    <div class="webgis-admin-layer-card" data-admin-layer="${webgisEscape(def.id)}">
      <div class="webgis-admin-layer-title">
        <span><span class="webgis-symbol" style="background:${webgisEscape(def.color)}"></span> ${webgisEscape(def.label)}</span>
        <span class="webgis-admin-layer-actions">
          <span>${webgisLayerFeatureCount(def.id)} doi tuong</span>
          <button type="button" data-webgis-layer-delete="${webgisEscape(def.id)}">Xoa</button>
        </span>
      </div>
      <div class="webgis-admin-layer-grid">
        <label class="switch-line"><input type="checkbox" data-webgis-admin-field="is_public" data-layer="${webgisEscape(def.id)}" ${def.is_public !== false ? 'checked' : ''}> Hien thi cho nguoi dung</label>
        <label class="switch-line"><input type="checkbox" data-webgis-admin-field="default_visible" data-layer="${webgisEscape(def.id)}" ${def.default_visible === true ? 'checked' : ''}> Bat mac dinh</label>
        <label class="switch-line"><input type="checkbox" data-webgis-admin-field="allow_user_toggle" data-layer="${webgisEscape(def.id)}" ${def.allow_user_toggle !== false ? 'checked' : ''}> Cho phep bat/tat</label>
        <label>Do trong suot
          <input type="range" min="0" max="1" step="0.05" value="${webgisNormalizeOpacity(def.opacity)}" data-webgis-admin-field="opacity" data-layer="${webgisEscape(def.id)}">
        </label>
        <label>Thu tu
          <input type="number" value="${Number(def.sort_order || 0)}" data-webgis-admin-field="sort_order" data-layer="${webgisEscape(def.id)}">
        </label>
        <label>Nhom layer
          <input type="text" value="${webgisEscape(def.category || 'Chung')}" data-webgis-admin-field="category" data-layer="${webgisEscape(def.id)}">
        </label>
      </div>
      ${webgisRenderAdminFieldList(def)}
    </div>
  `).join('');
}

function webgisLayerMetadataPayload(def) {
  return {
    is_public: def.is_public !== false,
    default_visible: def.default_visible === true,
    allow_user_toggle: def.allow_user_toggle !== false,
    opacity: webgisNormalizeOpacity(def.opacity),
    sort_order: Number(def.sort_order || 0),
    category: def.category || 'Chung',
    visible_fields: webgisNormalizeFieldList(def.visible_fields)
  };
}

function webgisScheduleLayerMetadataPatch(layerId, delay = 350) {
  if (!webgisAdminToken) return;
  clearTimeout(webgisState.layerPatchTimers.get(layerId));
  webgisState.layerPatchTimers.set(layerId, setTimeout(() => {
    webgisPatchLayerMetadata(layerId).catch(error => webgisSetSaveStatus(error.message || String(error), true));
  }, delay));
}

function webgisApplyLayerOpacity(layerId, rawValue, persistDelay = 0) {
  const def = webgisState.layerDefs.find(layer => layer.id === layerId);
  if (def) def.opacity = Number(rawValue);
  webgisUpdateLayerStyle(layerId);
  if (webgisAdminToken && persistDelay >= 0) webgisScheduleLayerMetadataPatch(layerId, persistDelay);
}

async function webgisPatchLayerMetadata(layerId) {
  if (!webgisRequireAdmin()) return;
  const def = webgisState.layerDefs.find(layer => layer.id === layerId);
  if (!def) return;
  const response = await fetch(`${webgisApiBase}/${encodeURIComponent(webgisProjectId)}/layers/${encodeURIComponent(layerId)}`, {
    method: 'PATCH',
    headers: { 'Content-Type': 'application/json', ...webgisAdminHeaders() },
    body: JSON.stringify(webgisLayerMetadataPayload(def))
  });
  const payload = await response.json().catch(() => ({}));
  if (response.status === 401) {
    webgisClearAdminSession();
    webgisShowAdminLogin('Phien admin WebGIS da het han. Vui long dang nhap lai.');
  }
  if (!response.ok) throw new Error(payload.error || 'Khong cap nhat duoc layer WebGIS.');
  webgisSetSaveStatus('Da cap nhat hien thi layer');
}

function webgisRemoveLayerLocally(layerId) {
  const normalizedLayerId = String(layerId || '');
  webgisState.deletedLayerIds.add(normalizedLayerId);
  clearTimeout(webgisState.layerPatchTimers.get(normalizedLayerId));
  webgisState.layerPatchTimers.delete(normalizedLayerId);
  const layer = webgisState.overlayLayers.get(normalizedLayerId);
  if (layer && webgisState.map) webgisState.map.removeLayer(layer);
  webgisState.overlayLayers.delete(normalizedLayerId);
  webgisState.featureCache.delete(normalizedLayerId);
  webgisState.loadedLayerIds.delete(normalizedLayerId);
  webgisState.features = webgisState.features.filter(feature => String(feature.properties?.layer || '') !== normalizedLayerId);
  webgisState.layerDefs = webgisState.layerDefs.filter(def => def.id !== normalizedLayerId);
  if (webgisState.selectedFeatureId && !webgisAllCachedFeatures().some(feature => feature.properties.__id === webgisState.selectedFeatureId)) {
    webgisState.selectedFeatureId = null;
    webgisState.selectedVector = null;
    webgisRenderFeatureDetail(null);
  }
  webgisRebuildOverlays();
  if (!webgisEl('webgisAttributePanel')?.hidden) webgisRenderAttributeTable();
}

async function webgisDeleteLayer(layerId) {
  if (!webgisRequireAdmin()) return;
  const def = webgisState.layerDefs.find(layer => layer.id === layerId);
  if (!def) return;
  const count = webgisLayerFeatureCount(layerId);
  if (!confirm(`Xoa layer "${def.label}" va ${count} doi tuong cua layer nay?`)) return;
  const response = await fetch(`${webgisApiBase}/${encodeURIComponent(webgisProjectId)}/layers/${encodeURIComponent(layerId)}`, {
    method: 'DELETE',
    headers: webgisAdminHeaders()
  });
  const payload = await response.json().catch(() => ({}));
  if (response.status === 401) {
    webgisClearAdminSession();
    webgisShowAdminLogin('Phien admin WebGIS da het han. Vui long dang nhap lai.');
  }
  if (!response.ok) throw new Error(payload.error || 'Khong xoa duoc layer WebGIS.');
  webgisRemoveLayerLocally(layerId);
  webgisSaveLocal(webgisStatePayload());
  webgisSetSaveStatus(`Da xoa layer ${def.label}`);
}

function webgisRefreshSelectedFeatureDisplay() {
  if (!webgisState.selectedFeatureId) return;
  const feature = webgisAllCachedFeatures().find(item => item.properties.__id === webgisState.selectedFeatureId);
  if (!feature) return;
  webgisRenderFeatureDetail(feature);
  if (webgisState.selectedVector?.bindPopup) {
    webgisState.selectedVector.bindPopup(webgisPopupHtml(feature));
    if (webgisState.selectedVector.isPopupOpen?.()) {
      webgisState.selectedVector.setPopupContent(webgisPopupHtml(feature));
    }
  }
  if (!webgisEl('webgisAttributePanel')?.hidden) webgisRenderAttributeTable();
}

async function webgisFitLayer(layerId) {
  const existingLayer = webgisState.overlayLayers.get(layerId);
  if (existingLayer) {
    const existingBounds = existingLayer.getBounds?.();
    if (existingBounds?.isValid?.()) {
      webgisState.map.fitBounds(existingBounds.pad(0.14));
      webgisInvalidateSize(40);
      return;
    }
  }
  const def = webgisState.layerDefs.find(layer => layer.id === layerId);
  const wasVisible = def?.visible === true;
  if (def && !wasVisible) def.visible = true;
  if (!webgisState.loadedLayerIds.has(layerId)) {
    webgisSetSaveStatus(`Đang nạp layer ${def?.label || layerId}...`);
    await webgisEnsureLayerLoaded(layerId);
    webgisRestoreConnectedStatus();
  }
  if (!wasVisible || !webgisState.overlayLayers.has(layerId)) webgisRebuildOverlays();
  const layer = webgisState.overlayLayers.get(layerId);
  if (!layer) {
    webgisRestoreConnectedStatus();
    return;
  }
  const bounds = layer.getBounds?.();
  if (bounds?.isValid?.()) {
    webgisState.map.fitBounds(bounds.pad(0.14));
    webgisInvalidateSize(40);
  }
  webgisRestoreConnectedStatus();
}

function webgisFitAll() {
  const group = L.featureGroup(Array.from(webgisState.overlayLayers.values()));
  const bounds = group.getBounds();
  if (bounds.isValid()) webgisState.map.fitBounds(bounds.pad(0.12));
}

function webgisUpdateLayerStyle(layerId) {
  const layer = webgisState.overlayLayers.get(layerId);
  if (!layer) return;
  layer.eachLayer(child => {
    if (child.feature && child.setStyle) child.setStyle(webgisStyle(child.feature));
  });
}

function webgisTextForFeature(feature) {
  const props = feature.properties || {};
  return [props.ma_thua, props.ma_khoanh, props.ma_dv, props.ten, props.chu_su_dung, webgisFeatureLandCode(feature), props.loai_dat, props.Loaidat, props.muc_dich, props.quy_hoach, props.loai_quy_hoach, props.dia_danh, props.ghi_chu, webgisLayerLabel(props.layer)].join(' ').toLowerCase();
}

function webgisSearch() {
  const query = webgisEl('webgisSearchInput').value.trim().toLowerCase();
  const box = webgisEl('webgisSearchResults');
  if (!query) {
    box.hidden = true;
    box.innerHTML = '';
    return;
  }
  const results = webgisState.features.filter(feature => webgisTextForFeature(feature).includes(query)).slice(0, 30);
  if (!results.length) {
    box.innerHTML = '<div class="webgis-result-item"><strong>Không tìm thấy dữ liệu phù hợp</strong><span>Thử mã đất, mã thửa hoặc tên địa danh khác.</span></div>';
    box.hidden = false;
    return;
  }
  box.innerHTML = results.map(feature => `
    <button type="button" class="webgis-result-item" data-feature-id="${webgisEscape(feature.properties.__id)}">
      <strong>${webgisEscape(webgisFeatureTitle(feature))}</strong>
      <span>${webgisEscape(webgisLayerLabel(feature.properties.layer))} - ${webgisEscape(webgisFeatureLandCode(feature) || feature.properties.loai_dat || feature.properties.Loaidat || feature.properties.quy_hoach || '')}</span>
    </button>
  `).join('');
  box.hidden = false;
  webgisZoomToFeature(results[0].properties.__id);
}

function webgisZoomToFeature(featureId) {
  const feature = webgisState.features.find(item => item.properties.__id === featureId);
  const vector = webgisState.featureLayers.get(featureId);
  if (!feature || !vector) return;
  if (vector.getBounds) {
    const bounds = vector.getBounds();
    if (bounds.isValid()) webgisState.map.fitBounds(bounds.pad(0.35));
  } else if (vector.getLatLng) {
    webgisState.map.setView(vector.getLatLng(), 17);
  }
  webgisSelectFeature(feature, vector);
}

function webgisAllPropertyKeys(features, layerId = '') {
  const def = webgisLayerDefById(layerId);
  const visibleFields = webgisNormalizeFieldList(def?.visible_fields);
  if (Array.isArray(visibleFields)) return visibleFields;
  const keys = new Set(['ma_thua', 'ma_khoanh', 'ten', 'loai_dat', 'Loaidat', 'dien_tich', 'chu_su_dung', 'muc_dich', 'quy_hoach', 'ghi_chu']);
  features.forEach(feature => Object.keys(feature.properties || {}).forEach(key => {
    if (!webgisInternalFields.has(key)) keys.add(key);
  }));
  return webgisOrderPropertyKeys(Array.from(keys).filter(key => features.some(feature => webgisIsVisibleValue(feature.properties?.[key]))));
}

function webgisPopulateAttrLayerSelect() {
  const select = webgisEl('webgisAttrLayer');
  select.innerHTML = webgisState.layerDefs
    .filter(def => def.is_public !== false || webgisAdminToken)
    .map(def => `<option value="${webgisEscape(def.id)}">${webgisEscape(def.label)}</option>`).join('');
}

function webgisRenderAttributeTable() {
  const layerId = webgisEl('webgisAttrLayer').value || webgisState.layerDefs[0]?.id || '';
  const query = webgisEl('webgisAttrSearch').value.trim().toLowerCase();
  let features = webgisState.features.filter(feature => feature.properties.layer === layerId);
  if (query) features = features.filter(feature => JSON.stringify(feature.properties).toLowerCase().includes(query));
  if (webgisState.attrSortKey) {
    const key = webgisState.attrSortKey;
    features = features.slice().sort((a, b) => String(a.properties[key] ?? '').localeCompare(String(b.properties[key] ?? ''), 'vi') * webgisState.attrSortDir);
  }
  const keys = webgisAllPropertyKeys(features, layerId);
  const table = webgisEl('webgisAttrTable');
  table.innerHTML = `
    <thead><tr>${keys.map((key, index) => `<th data-webgis-sort="${webgisEscape(key)}">
      <span>${webgisEscape(key)}</span>
      ${webgisAdminToken ? `<span class="webgis-attr-field-actions">
        <button type="button" title="Đưa cột sang trái" data-webgis-attr-field-move="${webgisEscape(layerId)}" data-field="${webgisEscape(key)}" data-direction="up" ${index === 0 ? 'disabled' : ''}>←</button>
        <button type="button" title="Đưa cột sang phải" data-webgis-attr-field-move="${webgisEscape(layerId)}" data-field="${webgisEscape(key)}" data-direction="down" ${index === keys.length - 1 ? 'disabled' : ''}>→</button>
      </span>` : ''}
    </th>`).join('')}</tr></thead>
    <tbody>${features.map(feature => `
      <tr data-feature-id="${webgisEscape(feature.properties.__id)}">${keys.map(key => `<td>${webgisEscape(feature.properties[key] ?? '')}</td>`).join('')}</tr>
    `).join('')}</tbody>
  `;
  webgisHighlightAttrRow(webgisState.selectedFeatureId);
}

function webgisHighlightAttrRow(featureId) {
  const table = webgisEl('webgisAttrTable');
  if (!table) return;
  table.querySelectorAll('tr.selected').forEach(row => row.classList.remove('selected'));
  if (!featureId || !window.CSS?.escape) return;
  table.querySelector(`tr[data-feature-id="${CSS.escape(featureId)}"]`)?.classList.add('selected');
}

function webgisFormatDistance(meters) {
  return meters >= 1000 ? `${(meters / 1000).toFixed(2)} km` : `${meters.toFixed(1)} m`;
}

function webgisFormatArea(squareMeters) {
  return squareMeters >= 10000 ? `${(squareMeters / 10000).toFixed(2)} ha` : `${squareMeters.toFixed(1)} m²`;
}

function webgisSphericalArea(latlngs) {
  if (latlngs.length < 3) return 0;
  const radius = 6378137;
  const rad = Math.PI / 180;
  let area = 0;
  for (let i = 0; i < latlngs.length; i += 1) {
    const p1 = latlngs[i];
    const p2 = latlngs[(i + 1) % latlngs.length];
    area += (p2.lng - p1.lng) * rad * (2 + Math.sin(p1.lat * rad) + Math.sin(p2.lat * rad));
  }
  return Math.abs(area * radius * radius / 2);
}

function webgisSetMeasureMode(mode) {
  webgisState.measureMode = mode;
  webgisState.measurePoints = [];
  webgisClearMeasure(false);
  webgisEl('webgisMeasureBadge').textContent = mode === 'distance' ? 'Đo khoảng cách: bấm các điểm trên bản đồ.' : 'Đo diện tích: bấm các đỉnh vùng trên bản đồ.';
}

function webgisClearMeasure(resetMode = true) {
  if (webgisState.measureLayer) {
    webgisState.map.removeLayer(webgisState.measureLayer);
    webgisState.measureLayer = null;
  }
  webgisState.measurePoints = [];
  if (resetMode) {
    webgisState.measureMode = null;
    webgisEl('webgisMeasureBadge').textContent = 'Sẵn sàng tra cứu bản đồ';
  }
}

function webgisHandleMeasureClick(latlng) {
  if (!webgisState.measureMode) return;
  webgisState.measurePoints.push(latlng);
  if (webgisState.measureLayer) webgisState.map.removeLayer(webgisState.measureLayer);
  if (webgisState.measureMode === 'distance') {
    webgisState.measureLayer = L.polyline(webgisState.measurePoints, { color: '#f97316', weight: 4 }).addTo(webgisState.map);
    const total = webgisState.measurePoints.slice(1).reduce((sum, point, index) => sum + webgisState.map.distance(webgisState.measurePoints[index], point), 0);
    webgisEl('webgisMeasureBadge').textContent = `Chiều dài: ${webgisFormatDistance(total)}. Bấm Xóa đo để kết thúc.`;
  } else {
    webgisState.measureLayer = L.polygon(webgisState.measurePoints, { color: '#f97316', weight: 3, fillOpacity: 0.18 }).addTo(webgisState.map);
    webgisEl('webgisMeasureBadge').textContent = `Diện tích: ${webgisFormatArea(webgisSphericalArea(webgisState.measurePoints))}. Bấm Xóa đo để kết thúc.`;
  }
}

async function webgisTakeScreenshot() {
  if (!window.html2canvas) {
    alert('Chưa tải được thư viện chụp ảnh bản đồ. Vui lòng thử lại sau.');
    return;
  }
  try {
    const canvas = await window.html2canvas(webgisEl('webgisMap'), { useCORS: true, backgroundColor: '#eef6ff' });
    const link = document.createElement('a');
    link.download = 'webgis-ban-do.png';
    link.href = canvas.toDataURL('image/png');
    link.click();
  } catch (error) {
    alert('Không chụp được ảnh do trình duyệt chặn ảnh nền bản đồ từ nguồn ngoài. Có thể dùng công cụ In bản đồ để lưu PDF.');
  }
}

async function webgisImportGeoJson() {
  if (!webgisRequireAdmin()) return;
  const file = webgisEl('webgisImportInput').files?.[0];
  const name = webgisEl('webgisNewLayerName').value.trim() || file?.name?.replace(/\.(geojson|json)$/i, '') || 'Layer GeoJSON';
  if (!file) {
    alert('Hãy chọn file GeoJSON trước.');
    return;
  }
  const data = JSON.parse(await file.text());
  const layerId = `custom_${Date.now()}`;
  const color = webgisEl('webgisNewLayerColor').value || '#2563eb';
  const category = webgisEl('webgisNewLayerCategory')?.value.trim() || 'Tu them';
  const features = webgisNormalizeFeatures(data, layerId).map(feature => ({ ...feature, properties: { ...feature.properties, layer: layerId } }));
  if (!features.length) {
    alert('File GeoJSON không có đối tượng hợp lệ.');
    return;
  }
  const nextSort = Math.max(0, ...webgisState.layerDefs.map(def => Number(def.sort_order || 0))) + 1;
  webgisState.layerDefs.push({
    id: layerId,
    label: name,
    color,
    is_public: true,
    default_visible: false,
    allow_user_toggle: true,
    visible: true,
    opacity: 1,
    sort_order: nextSort,
    category,
    feature_count: features.length,
    custom: true
  });
  webgisState.layerDefs.sort(webgisLayerSort);
  webgisSetLayerFeatureCache(layerId, features);
  webgisRebuildOverlays();
  await webgisFitLayer(layerId);
  await webgisSaveLayerFeatures(layerId);
  webgisScheduleSave();
  alert(`Đã thêm ${features.length} đối tượng vào layer "${name}".`);
}

function webgisSaveSelectedFeatureProps() {
  if (!webgisRequireAdmin()) return;
  const id = webgisState.selectedFeatureId;
  if (!id) {
    alert('Hãy chọn một đối tượng trên bản đồ trước.');
    return;
  }
  let props;
  try {
    props = JSON.parse(webgisEl('webgisFeatureEditor').value || '{}');
  } catch (error) {
    alert('Nội dung thuộc tính phải là JSON hợp lệ.');
    return;
  }
  const feature = webgisState.features.find(item => item.properties.__id === id);
  if (!feature) return;
  feature.properties = { ...props, layer: props.layer || feature.properties.layer, __id: id };
  const layerFeatures = webgisState.featureCache.get(feature.properties.layer) || [];
  const cacheIndex = layerFeatures.findIndex(item => item.properties.__id === id);
  if (cacheIndex >= 0) layerFeatures[cacheIndex] = feature;
  webgisState.featureCache.set(feature.properties.layer, layerFeatures);
  webgisRebuildOverlays();
  webgisZoomToFeature(id);
  webgisRenderAttributeTable();
  webgisSaveLayerFeatures(feature.properties.layer).catch(error => webgisSetSaveStatus(error.message || String(error), true));
  webgisScheduleSave();
}

function webgisBindEvents() {
  webgisEl('webgisLayerList').addEventListener('change', async event => {
    const toggleId = event.target?.dataset?.webgisLayerToggle;
    const opacityId = event.target?.dataset?.webgisLayerOpacity;
    if (toggleId) {
      const def = webgisState.layerDefs.find(layer => layer.id === toggleId);
      if (def?.allow_user_toggle === false) {
        event.target.checked = def.visible === true;
        return;
      }
      if (def) def.visible = event.target.checked;
      if (event.target.checked) {
        await webgisEnsureLayerLoaded(toggleId);
        webgisRestoreConnectedStatus();
      }
      webgisRebuildOverlays();
      webgisScheduleSave();
    }
    if (opacityId) {
      webgisApplyLayerOpacity(opacityId, event.target.value, 250);
    }
  });
  webgisEl('webgisLayerList').addEventListener('input', event => {
    const opacityId = event.target?.dataset?.webgisLayerOpacity;
    if (!opacityId) return;
    webgisApplyLayerOpacity(opacityId, event.target.value, 900);
  });
  webgisEl('webgisLayerList').addEventListener('click', event => {
    const layerId = event.target?.dataset?.webgisLayerZoom;
    if (layerId) webgisFitLayer(layerId).catch(error => webgisSetSaveStatus(error.message || String(error), true));
  });
  webgisEl('webgisSearchBtn').addEventListener('click', webgisSearch);
  webgisEl('webgisSearchInput').addEventListener('keydown', event => {
    if (event.key === 'Enter') webgisSearch();
  });
  webgisEl('webgisSearchResults').addEventListener('click', event => {
    const button = event.target.closest('[data-feature-id]');
    if (!button) return;
    webgisZoomToFeature(button.dataset.featureId);
    webgisEl('webgisSearchResults').hidden = true;
  });
  webgisEl('webgisFitAllBtn').addEventListener('click', webgisFitAll);
  webgisEl('webgisToggleSidebarBtn').addEventListener('click', () => {
    const page = webgisEl('webgisPage');
    webgisSetPanelCollapsed('layers', !page.classList.contains('layers-collapsed'));
  });
  webgisEl('webgisToggleInfoBtn').addEventListener('click', () => {
    const page = webgisEl('webgisPage');
    webgisSetPanelCollapsed('info', !page.classList.contains('info-collapsed'));
  });
  webgisEl('webgisFeatureDetail').addEventListener('click', event => {
    if (event.target?.dataset?.webgisDetailZoom) webgisZoomToFeature(webgisState.selectedFeatureId);
    if (event.target?.dataset?.webgisDetailExport) webgisExportSelectedFeatureInfo();
  });
  webgisEl('webgisHomeBtn')?.addEventListener('click', showHomePage);
  webgisEl('webgisAiBtn').addEventListener('click', () => openAiAssistant('webgis'));
  webgisEl('webgisOpenTableBtn').addEventListener('click', () => {
    webgisEl('webgisAttributePanel').hidden = false;
    webgisRenderAttributeTable();
  });
  webgisEl('webgisCloseTableBtn').addEventListener('click', () => webgisEl('webgisAttributePanel').hidden = true);
  webgisEl('webgisAttrLayer').addEventListener('change', webgisRenderAttributeTable);
  webgisEl('webgisAttrSearch').addEventListener('input', webgisRenderAttributeTable);
  webgisEl('webgisAttrTable').addEventListener('click', event => {
    const moveLayerId = event.target?.dataset?.webgisAttrFieldMove;
    if (moveLayerId) {
      if (webgisMoveLayerField(moveLayerId, event.target.dataset.field, event.target.dataset.direction)) {
        webgisRenderAdminLayerList();
      }
      event.stopPropagation();
      return;
    }
    const sortKey = event.target?.dataset?.webgisSort;
    if (sortKey) {
      webgisState.attrSortDir = webgisState.attrSortKey === sortKey ? -webgisState.attrSortDir : 1;
      webgisState.attrSortKey = sortKey;
      webgisRenderAttributeTable();
      return;
    }
    const row = event.target.closest('tr[data-feature-id]');
    if (row) webgisZoomToFeature(row.dataset.featureId);
  });
  webgisEl('webgisAdminBtn').addEventListener('click', () => {
    if (!webgisRequireAdmin()) return;
    webgisEl('webgisAdminPanel').hidden = !webgisEl('webgisAdminPanel').hidden;
    webgisEl('webgisAdminLoginPanel').hidden = true;
    webgisRenderAdminLayerList();
    webgisInvalidateSize(80);
  });
  webgisEl('webgisCloseAdminBtn').addEventListener('click', () => webgisEl('webgisAdminPanel').hidden = true);
  webgisEl('webgisAdminLayerList').addEventListener('click', async event => {
    const layerId = event.target?.dataset?.webgisLayerDelete;
    if (layerId) {
      webgisDeleteLayer(layerId).catch(error => webgisSetSaveStatus(error.message || String(error), true));
      return;
    }
    const loadFieldsId = event.target?.dataset?.webgisLayerLoadFields;
    if (loadFieldsId) {
      await webgisEnsureLayerLoaded(loadFieldsId);
      webgisRestoreConnectedStatus();
      webgisRenderAdminLayerList();
      webgisSetSaveStatus('Da nap danh sach thuoc tinh layer');
      return;
    }
    const allFieldsId = event.target?.dataset?.webgisLayerFieldsAll;
    if (allFieldsId) {
      const def = webgisState.layerDefs.find(layer => layer.id === allFieldsId);
      if (!def) return;
      def.visible_fields = null;
      webgisRenderAdminLayerList();
      webgisRefreshSelectedFeatureDisplay();
      webgisScheduleLayerMetadataPatch(allFieldsId, 50);
      return;
    }
    const moveLayerId = event.target?.dataset?.webgisFieldMove;
    if (moveLayerId) {
      const card = event.target.closest('[data-admin-layer]');
      const selected = new Set(Array.from(card?.querySelectorAll('[data-webgis-field-toggle]') || [])
        .filter(input => input.checked)
        .map(input => input.value));
      if (webgisMoveLayerField(moveLayerId, event.target.dataset.field, event.target.dataset.direction, selected)) {
        webgisRenderAdminLayerList();
      }
    }
  });
  webgisEl('webgisAdminLayerList').addEventListener('change', async event => {
    const fieldLayerId = event.target?.dataset?.webgisFieldToggle;
    if (fieldLayerId) {
      const def = webgisState.layerDefs.find(layer => layer.id === fieldLayerId);
      if (!def) return;
      const card = event.target.closest('[data-admin-layer]');
      const checked = Array.from(card?.querySelectorAll('[data-webgis-field-toggle]') || [])
        .filter(input => input.checked)
        .map(input => input.value);
      def.visible_fields = checked;
      webgisRefreshSelectedFeatureDisplay();
      webgisScheduleLayerMetadataPatch(fieldLayerId, 120);
      return;
    }
    const layerId = event.target?.dataset?.layer;
    const field = event.target?.dataset?.webgisAdminField;
    if (!layerId || !field) return;
    const def = webgisState.layerDefs.find(layer => layer.id === layerId);
    if (!def) return;
    if (['is_public', 'default_visible', 'allow_user_toggle'].includes(field)) {
      def[field] = event.target.checked;
      if (field === 'is_public' && !def.is_public) def.visible = false;
      if (field === 'default_visible') def.visible = event.target.checked;
      if (field === 'default_visible' && event.target.checked) await webgisEnsureLayerLoaded(layerId);
    } else if (field === 'sort_order') {
      def.sort_order = Number(event.target.value || 0);
      webgisState.layerDefs.sort(webgisLayerSort);
    } else if (field === 'category') {
      def.category = event.target.value.trim() || 'Chung';
    } else if (field === 'opacity') {
      def.opacity = webgisNormalizeOpacity(event.target.value);
      webgisUpdateLayerStyle(layerId);
    }
    webgisRenderLayerList();
    webgisRenderAdminLayerList();
    webgisRebuildOverlays();
    webgisScheduleLayerMetadataPatch(layerId, ['category', 'sort_order', 'opacity'].includes(field) ? 500 : 50);
  });
  webgisEl('webgisAdminLayerList').addEventListener('input', event => {
    const layerId = event.target?.dataset?.layer;
    const field = event.target?.dataset?.webgisAdminField;
    if (!layerId || !['opacity', 'sort_order', 'category'].includes(field)) return;
    const def = webgisState.layerDefs.find(layer => layer.id === layerId);
    if (!def) return;
    if (field === 'opacity') {
      def.opacity = webgisNormalizeOpacity(event.target.value);
      webgisUpdateLayerStyle(layerId);
    } else if (field === 'sort_order') {
      def.sort_order = Number(event.target.value || 0);
    } else {
      def.category = event.target.value.trim() || 'Chung';
    }
    webgisScheduleLayerMetadataPatch(layerId, 600);
  });
  webgisEl('webgisCloseAdminLoginBtn').addEventListener('click', () => webgisEl('webgisAdminLoginPanel').hidden = true);
  webgisEl('webgisAdminLoginSubmit').addEventListener('click', () => webgisLoginAdmin().catch(error => webgisSetAdminStatus(error.message || String(error), true)));
  webgisEl('webgisAdminPassword').addEventListener('keydown', event => {
    if (event.key === 'Enter') webgisLoginAdmin().catch(error => webgisSetAdminStatus(error.message || String(error), true));
  });
  webgisEl('webgisAdminLogoutBtn').addEventListener('click', () => {
    webgisClearAdminSession();
    webgisSetSaveStatus('Đã đăng xuất admin WebGIS');
    webgisShowAdminLogin('Đã đăng xuất admin WebGIS.');
  });
  webgisEl('webgisImportBtn').addEventListener('click', () => webgisImportGeoJson().catch(error => alert(error.message || String(error))));
  webgisEl('webgisSaveFeatureBtn').addEventListener('click', webgisSaveSelectedFeatureProps);
  webgisEl('webgisLocateBtn').addEventListener('click', () => webgisState.map.locate({ setView: true, maxZoom: 17 }));
  webgisEl('webgisMeasureDistanceBtn').addEventListener('click', () => webgisSetMeasureMode('distance'));
  webgisEl('webgisMeasureAreaBtn').addEventListener('click', () => webgisSetMeasureMode('area'));
  webgisEl('webgisClearMeasureBtn').addEventListener('click', () => webgisClearMeasure(true));
  webgisEl('webgisPrintBtn').addEventListener('click', () => window.print());
  webgisEl('webgisShotBtn').addEventListener('click', () => webgisTakeScreenshot());
  webgisEl('webgisFullscreenBtn').addEventListener('click', () => webgisEl('webgisPage').requestFullscreen?.());
  window.addEventListener('resize', () => webgisInvalidateSize(90), { passive: true });
  document.addEventListener('fullscreenchange', () => webgisInvalidateSize(120));
  if (window.ResizeObserver && !webgisState.resizeObserver) {
    webgisState.resizeObserver = new ResizeObserver(() => webgisInvalidateSize(40));
    webgisState.resizeObserver.observe(webgisEl('webgisMap'));
  }
}

async function initializeWebGIS() {
  if (webgisState.initialized) {
    webgisInvalidateSize(80);
    return;
  }
  if (webgisState.initializing) {
    await webgisState.initializing;
    webgisInvalidateSize(80);
    return;
  }
  if (!window.L) {
    webgisEl('webgisMap').innerHTML = '<div class="webgis-detail-empty">Không tải được Leaflet.js. Vui lòng kiểm tra kết nối mạng hoặc CDN.</div>';
    return;
  }
  webgisState.initializing = (async () => {
    try {
      const sample = JSON.parse(document.getElementById('webgisSampleData').textContent);
      const savedData = await webgisLoadSavedData();
      webgisState.layerDefs = webgisNormalizeLayerDefs(savedData?.layerDefs, savedData?.deletedLayerIds);
      webgisState.features = [];
      webgisState.featureCache.clear();
      webgisState.loadedLayerIds.clear();
      const initialFeatures = Array.isArray(savedData?.features) && savedData.features.length ? savedData.features : (!savedData ? sample.features : []);
      webgisPrimeFeatureCache(initialFeatures);
      const map = L.map('webgisMap', { zoomControl: false, preferCanvas: true }).setView([21.0405, 105.8520], 15);
      webgisState.map = map;
      L.control.zoom({ position: 'bottomright' }).addTo(map);
      const osm = L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', { maxZoom: 20, attribution: '&copy; OpenStreetMap contributors' }).addTo(map);
      const satellite = L.tileLayer('https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}', { maxZoom: 19, attribution: 'Tiles &copy; Esri' });
      const terrain = L.tileLayer('https://{s}.tile.opentopomap.org/{z}/{x}/{y}.png', { maxZoom: 17, attribution: '&copy; OpenTopoMap contributors' });
      L.control.layers({ OpenStreetMap: osm, 'Ảnh vệ tinh': satellite, 'Địa hình': terrain }, null, { position: 'topright', collapsed: true }).addTo(map);
      await Promise.all(webgisState.layerDefs.filter(def => def.is_public !== false && def.visible === true).map(def => webgisEnsureLayerLoaded(def.id)));
      webgisRebuildOverlays();
      webgisBindEvents();
      webgisUpdateAdminUi();
      map.on('mousemove', event => {
        webgisEl('webgisCoordinateBar').textContent = `Tọa độ: ${event.latlng.lat.toFixed(6)}, ${event.latlng.lng.toFixed(6)}`;
      });
      map.on('click', event => webgisHandleMeasureClick(event.latlng));
      map.on('locationfound', event => {
        L.circleMarker(event.latlng, { radius: 8, color: '#0f766e', fillColor: '#14b8a6', fillOpacity: 0.85 }).addTo(map)
          .bindPopup('Vị trí hiện tại của bạn').openPopup();
      });
      map.on('locationerror', () => alert('Không xác định được vị trí. Hãy cho phép trình duyệt truy cập vị trí nếu cần.'));
      webgisFitAll();
      webgisState.initialized = true;
    } finally {
      webgisState.initializing = null;
    }
  })();
  await webgisState.initializing;
  webgisInvalidateSize(80);
  webgisInvalidateSize(260);
}
"""


def normalize_key(value) -> str:
    text = unicodedata.normalize("NFD", str(value or "").strip().lower())
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d")
    return re.sub(r"\s+", " ", text)


def parse_number(value) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def format_ha(value: float | None) -> str:
    if value is None:
        return ""
    return f"{value:.2f}".replace(".", ",")


def read_previous_plan_values() -> dict[str, float]:
    files = [p for p in PREVIOUS_PLAN_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
        return {}
    wb = openpyxl.load_workbook(files[0], data_only=True)
    ws = wb[wb.sheetnames[0]]
    code_col = None
    area_col = None
    for row in range(1, min(ws.max_row, 25) + 1):
        for col in range(1, ws.max_column + 1):
            text = normalize_key(ws.cell(row, col).value)
            if text in {"mã", "ma", "mã đất", "ma dat"}:
                code_col = col
            if "diện tích" in text or "dien tich" in text:
                area_col = col
    if not area_col:
        for row in range(1, min(ws.max_row, 25) + 1):
            for col in range(1, ws.max_column + 1):
                if "quy hoạch" in normalize_key(ws.cell(row, col).value) and col <= ws.max_column:
                    area_col = col
                    break
            if area_col:
                break
    if not code_col or not area_col:
        return {}

    values: dict[str, float] = {}
    for row in range(1, ws.max_row + 1):
        code = str(ws.cell(row, code_col).value or "").strip().upper()
        name = normalize_key(ws.cell(row, max(1, code_col - 1)).value)
        if not code and "tổng diện tích tự nhiên" in name:
            code = "DTTN"
        area = parse_number(ws.cell(row, area_col).value)
        if code and area is not None:
            values[code] = area
    return values


def read_previous_plan_values_clean() -> dict[str, float]:
    files = [p for p in PREVIOUS_PLAN_DIR.glob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
        return {}
    wb = openpyxl.load_workbook(files[0], data_only=True)
    ws = wb[wb.sheetnames[0]]
    code_col = None
    area_col = None
    for row in range(1, min(ws.max_row, 25) + 1):
        for col in range(1, ws.max_column + 1):
            text = normalize_key(ws.cell(row, col).value)
            if text in {"ma", "ma dat"}:
                code_col = col
            if "dien tich" in text and (code_col is None or col > code_col):
                area_col = col
    if not area_col:
        for row in range(1, min(ws.max_row, 25) + 1):
            for col in range(1, ws.max_column + 1):
                if "quy hoach" in normalize_key(ws.cell(row, col).value):
                    area_col = col
                    break
            if area_col:
                break
    if not code_col or not area_col:
        return {}

    values: dict[str, float] = {}
    for row in range(1, ws.max_row + 1):
        code = str(ws.cell(row, code_col).value or "").strip().upper()
        name = normalize_key(ws.cell(row, max(1, code_col - 1)).value)
        if not code and "tong dien tich tu nhien" in name:
            code = "DTTN"
        area = parse_number(ws.cell(row, area_col).value)
        if code and area is not None:
            values[code] = area
    return values
def color(value) -> str | None:
    if not value:
        return None
    if getattr(value, "type", None) == "rgb" and value.rgb:
        rgb = value.rgb[-6:]
        if rgb == "000000" and str(value.rgb).startswith("00"):
            return None
        return f"#{rgb}"
    return None


def border_css(side) -> str:
    if side is None or side.style is None:
        return "1px solid #c8d0d9"
    width = "2px" if side.style in {"medium", "thick", "double"} else "1px"
    clr = color(side.color) or "#2f3640"
    return f"{width} solid {clr}"


def style_key(cell) -> str:
    fill = None
    if cell.fill and cell.fill.fill_type == "solid":
        fill = color(cell.fill.fgColor)
    font_color = color(cell.font.color)
    horizontal = cell.alignment.horizontal or "center"
    if horizontal == "centerContinuous":
        horizontal = "center"
    if horizontal in {"general", "distributed", "justify"}:
        horizontal = "left"
    parts = [
        f"background:{fill}" if fill else "",
        f"font-weight:{'700' if cell.font.bold else '400'}",
        f"font-style:{'italic' if cell.font.italic else 'normal'}",
        f"font-size:{int(cell.font.sz or 11)}pt",
        f"color:{font_color or '#17202a'}",
        f"text-align:{horizontal}",
        f"vertical-align:{cell.alignment.vertical or 'middle'}",
        f"white-space:{'normal' if cell.alignment.wrap_text else 'nowrap'}",
        f"border-top:{border_css(cell.border.top)}",
        f"border-right:{border_css(cell.border.right)}",
        f"border-bottom:{border_css(cell.border.bottom)}",
        f"border-left:{border_css(cell.border.left)}",
    ]
    return ";".join(p for p in parts if p)


def display_value(value, code: str = "", col: int | None = None) -> str:
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        return ""
    if col == 1 and code in STT_FIXES_BY_CODE:
        return STT_FIXES_BY_CODE[code]
    text = str(value)
    return LAND_NAME_FIXES.get(text.strip(), text)


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE, data_only=False)
    ws = wb["Sheet1"]
    previous_plan_values = {}
    total_columns = ws.max_column + 3

    merged_parent = {}
    merged_skip = set()
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = rng.bounds
        merged_parent[(min_row, min_col)] = (max_row - min_row + 1, max_col - min_col + 1)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                if (row, col) != (min_row, min_col):
                    merged_skip.add((row, col))

    code_rows = {}
    for row in range(1, ws.max_row + 1):
        code = ws.cell(row, 3).value
        if code is not None:
            code_rows[str(code).strip()] = row

    code_cols = {}
    for col in range(MATRIX_START_COL, MATRIX_END_COL + 1):
        code = ws.cell(HEADER_ROW, col).value
        if code is not None:
            code_cols[str(code).strip()] = col

    direct_children: dict[str, list[str]] = {}
    for code, row in code_rows.items():
        value = ws.cell(row, CURRENT_COL).value
        if not (isinstance(value, str) and value.startswith("=")):
            continue
        child_codes = []
        for child_row in [int(x) for x in re.findall(r"D(\d+)", value)]:
            child_code = ws.cell(child_row, 3).value
            if child_code is not None:
                child_codes.append(str(child_code).strip())
        if child_codes:
            direct_children[code] = child_codes

    all_data_codes = [
        str(ws.cell(row, 3).value).strip()
        for row in range(5, TOTAL_INCREASE_ROW)
        if ws.cell(row, 3).value is not None
    ]
    parent_codes = set(direct_children)
    input_codes = [code for code in all_data_codes if code not in parent_codes and code in code_cols]
    missing_codes: list[str] = []

    styles = {}
    style_names = {}
    css_rules = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            key = style_key(ws.cell(row, col))
            if key not in style_names:
                name = f"xl{len(style_names) + 1}"
                style_names[key] = name
                css_rules.append(f".{name}{{{key}}}")

    colgroup = []
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = ws.column_dimensions[letter].width or 8
        px = int(width * 8)
        colgroup.append(f'<col style="width:{px}px;min-width:{px}px">')
    colgroup.append('<col style="width:112px;min-width:112px">')
    colgroup.append('<col style="width:112px;min-width:112px">')
    colgroup.append('<col style="width:92px;min-width:92px">')

    rows_html = []
    for row in range(1, ws.max_row + 1):
        height = ws.row_dimensions[row].height or 30
        if row == 1:
            rows_html.append(
                f'<tr style="height:{max(height, 42)}px">'
                f'<td class="sheet-title" data-addr="A1" data-row="1" data-col="1" colspan="{total_columns}">'
                'BẢNG CHU CHUYỂN ĐẤT ĐAI'
                '</td></tr>'
            )
            continue
        cells = []
        for col in range(1, ws.max_column + 1):
            if (row, col) in merged_skip:
                continue
            cell = ws.cell(row, col)
            rowspan, colspan = merged_parent.get((row, col), (1, 1))
            cls = style_names[style_key(cell)]
            addr = f"{get_column_letter(col)}{row}"
            code = str(ws.cell(row, 3).value or "").strip()
            col_code = str(ws.cell(HEADER_ROW, col).value or "").strip()
            is_current_input = col == CURRENT_COL and code in input_codes
            is_matrix_input = row in [code_rows[c] for c in input_codes] and col in [code_cols[c] for c in input_codes]
            is_input = is_current_input or is_matrix_input
            attrs = [
                f'class="{cls}"',
                f'data-addr="{addr}"',
                f'data-row="{row}"',
                f'data-col="{col}"',
            ]
            if rowspan > 1:
                attrs.append(f'rowspan="{rowspan}"')
            if colspan > 1:
                attrs.append(f'colspan="{colspan}"')
            if code:
                attrs.append(f'data-code="{html.escape(code)}"')
            if col_code:
                attrs.append(f'data-col-code="{html.escape(col_code)}"')

            text = html.escape(display_value(cell.value, code, col))
            if is_input:
                attrs.append('data-input="1"')
                value = "" if cell.value is None or (isinstance(cell.value, str) and cell.value.startswith("=")) else html.escape(str(cell.value))
                content = f'<input inputmode="decimal" value="{value}" aria-label="{html.escape(addr)}">'
            elif col >= CURRENT_COL and row >= 4:
                attrs.append('data-auto="1"')
                content = f'<span class="value">{text}</span>'
            else:
                content = text
            cells.append(f"<td {' '.join(attrs)}>{content}</td>")
        if row == 2:
            cells.append(
                f'<td class="xl3" data-addr="{get_column_letter(PREVIOUS_PLAN_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_COL}" rowspan="2">'
                'Quy hoạch kỳ trước</td>'
            )
            cells.append(
                f'<td class="xl3" data-addr="{get_column_letter(PREVIOUS_PLAN_CHANGE_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_CHANGE_COL}" colspan="2">'
                'T&#259;ng (+) gi&#7843;m (-)</td>'
            )
        elif row == 3:
            cells.append(
                f'<td class="xl3" data-addr="{get_column_letter(PREVIOUS_PLAN_CHANGE_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_CHANGE_COL}">Di&#7879;n t&#237;ch</td>'
            )
            cells.append(
                f'<td class="xl3" data-addr="{get_column_letter(PREVIOUS_PLAN_STRUCTURE_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_STRUCTURE_COL}">C&#417; c&#7845;u (%)</td>'
            )
        elif row >= 4:
            previous_code = str(ws.cell(row, 3).value or "").strip().upper()
            if row == 4:
                previous_code = "DTTN"
            previous_text = html.escape(format_ha(previous_plan_values.get(previous_code)))
            cells.append(
                f'<td class="xl7" data-addr="{get_column_letter(PREVIOUS_PLAN_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_COL}" data-previous-plan="1" data-auto="1">'
                f'<span class="value">{previous_text}</span></td>'
            )
            cells.append(
                f'<td class="xl7" data-addr="{get_column_letter(PREVIOUS_PLAN_CHANGE_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_CHANGE_COL}" data-previous-change="1" data-auto="1">'
                '<span class="value"></span></td>'
            )
            cells.append(
                f'<td class="xl7" data-addr="{get_column_letter(PREVIOUS_PLAN_STRUCTURE_COL)}{row}" '
                f'data-row="{row}" data-col="{PREVIOUS_PLAN_STRUCTURE_COL}" data-previous-structure="1" data-auto="1">'
                '<span class="value"></span></td>'
            )
        rows_html.append(f'<tr style="height:{height}px">{"".join(cells)}</tr>')

    meta = {
        "inputCodes": input_codes,
        "missingCodes": missing_codes,
        "directChildren": direct_children,
        "codeRows": code_rows,
        "codeCols": code_cols,
        "dttnRow": 4,
        "currentCol": CURRENT_COL,
        "matrixStartCol": MATRIX_START_COL,
        "matrixEndCol": MATRIX_END_COL,
        "decreaseCol": DECREASE_COL,
        "changeCol": CHANGE_COL,
        "planCol": PLAN_COL,
        "previousPlanCol": PREVIOUS_PLAN_COL,
        "previousPlanChangeCol": PREVIOUS_PLAN_CHANGE_COL,
        "previousPlanStructureCol": PREVIOUS_PLAN_STRUCTURE_COL,
        "totalIncreaseRow": TOTAL_INCREASE_ROW,
        "planRow": PLAN_ROW,
        "tolerance": 0.0001,
    }

    meta_json = json.dumps(meta, ensure_ascii=False).replace("</", "<\\/")

    jszip_js = JSZIP.read_text(encoding="utf-8")
    logo_data_url = ""
    if LOGO.exists():
        logo_data_url = "data:image/jpeg;base64," + base64.b64encode(LOGO.read_bytes()).decode("ascii")
    home_bg_data_url = ""
    if HOME_BACKGROUND.exists():
        home_bg_data_url = "data:image/png;base64," + base64.b64encode(HOME_BACKGROUND.read_bytes()).decode("ascii")
    SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    LEGACY_SAMPLE_DIR.mkdir(parents=True, exist_ok=True)
    sample_links = []
    for source_name, public_name, label in SAMPLE_FILES:
        source_path = PREVIOUS_PLAN_DIR / source_name
        if not source_path.exists():
            continue
        shutil.copy2(source_path, SAMPLE_DIR / public_name)
        shutil.copy2(source_path, LEGACY_SAMPLE_DIR / public_name)
        payload = base64.b64encode(source_path.read_bytes()).decode("ascii")
        href = f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{payload}"
        sample_links.append(
            f'<a href="{href}" download="{html.escape(public_name, quote=True)}">{html.escape(label)}</a>'
        )
    sample_links_html = "\n      ".join(sample_links)
    webgis_data_dir = OUT.parent / "webgis"
    webgis_data_dir.mkdir(parents=True, exist_ok=True)
    webgis_sample_json_pretty = json.dumps(WEBGIS_SAMPLE_DATA, ensure_ascii=False, indent=2)
    (webgis_data_dir / "sample-land-data.geojson").write_text(webgis_sample_json_pretty, encoding="utf-8")
    webgis_sample_json = json.dumps(WEBGIS_SAMPLE_DATA, ensure_ascii=False).replace("</", "<\\/")
    doc = f"""<!doctype html>
<html lang="vi">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Phần mềm quản lý dữ liệu đất đai</title>
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css">
<style>
:root {{
  --bg: #eef5fb;
  --panel: #ffffff;
  --ink: #0f172a;
  --muted: #64748b;
  --line: #d8e3ee;
  --accent: #0f766e;
  --accent-2: #2563eb;
  --surface: rgba(255, 255, 255, 0.88);
  --warn: #b42318;
  --input: #fff8d9;
  --diagonal: #dcfce7;
  --auto: #f8fafc;
  --header: #e8f1f7;
  --locked: #f7f9fc;
}}
* {{ box-sizing: border-box; }}
html,
body,
#root {{
  height: 100%;
}}
html {{
  min-height: 100%;
}}
body {{
  margin: 0;
  min-height: 100vh;
  overflow-x: hidden;
  background:
    radial-gradient(circle at 18% 12%, rgba(37, 99, 235, 0.14), transparent 28%),
    radial-gradient(circle at 84% 10%, rgba(14, 165, 233, 0.12), transparent 30%),
    linear-gradient(135deg, #eef5fb 0%, #f8fbff 48%, #edf4fa 100%);
  background-attachment: fixed;
  color: var(--ink);
  font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, Helvetica, sans-serif;
}}
body.webgis-mode {{
  height: 100vh;
  overflow: hidden;
}}
.appbar {{
  position: sticky;
  top: 0;
  z-index: 70;
  display: flex;
  flex-wrap: wrap;
  align-items: flex-start;
  gap: 10px 14px;
  min-height: 74px;
  padding: 10px 16px 10px 108px;
  background: linear-gradient(135deg, rgba(255, 255, 255, 0.96), rgba(239, 246, 255, 0.94));
  border-bottom: 1px solid rgba(148, 163, 184, 0.42);
  box-shadow: 0 10px 28px rgba(15, 23, 42, 0.10);
  backdrop-filter: blur(12px);
}}
.title {{
  font-size: 16px;
  font-weight: 800;
  text-transform: uppercase;
  color: #0f2f57;
  letter-spacing: 0;
}}
.subtitle {{
  color: #315d87;
  font-size: 12px;
  font-weight: 700;
}}
.brand {{
  display: flex;
  align-items: center;
  gap: 9px;
  flex: 0 1 310px;
  min-width: 250px;
}}
.brand-logo {{
  width: 40px;
  height: 40px;
  border-radius: 50%;
  object-fit: cover;
  border: 2px solid #ffffff;
  box-shadow: 0 4px 12px rgba(15, 23, 42, 0.20);
  flex: 0 0 auto;
}}
.brand-text {{
  display: flex;
  flex-direction: column;
  gap: 1px;
  line-height: 1.15;
}}
.designer {{
  color: #64748b;
  font-size: 11px;
  font-weight: 600;
}}
.module-label {{
  display: none;
  width: max-content;
  max-width: 420px;
  margin-top: 3px;
  padding: 4px 10px;
  border: 1px solid #bfdbfe;
  border-radius: 999px;
  background: #eff6ff;
  color: #1d4ed8;
  font-size: 12px;
  font-weight: 800;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}}
body:not(.home-mode) .subtitle,
body:not(.home-mode) .designer {{
  display: none;
}}
body:not(.home-mode) .module-label {{
  display: inline-flex;
  align-items: center;
}}
body:not(.home-mode) .appbar {{
  min-height: 62px;
  align-items: center;
}}
body.webgis-mode .appbar {{
  height: 62px;
  min-height: 62px;
  overflow: hidden;
  flex-wrap: nowrap;
}}
.app-sidebar {{
  position: fixed;
  top: 10px;
  left: 12px;
  bottom: 12px;
  z-index: 90;
  width: 76px;
  display: flex;
  flex-direction: column;
  gap: 8px;
  padding: 10px 8px;
  border: 1px solid rgba(148, 163, 184, 0.34);
  border-radius: 18px;
  background: rgba(255, 255, 255, 0.92);
  box-shadow: 0 18px 42px rgba(15, 23, 42, 0.14);
  backdrop-filter: blur(14px);
}}
.side-logo {{
  display: grid;
  place-items: center;
  width: 44px;
  height: 44px;
  margin: 0 auto 4px;
  border-radius: 14px;
  color: #ffffff;
  background: linear-gradient(135deg, #2563eb, #0ea5e9);
  font-size: 20px;
  font-weight: 800;
  box-shadow: 0 10px 22px rgba(37, 99, 235, 0.24);
}}
.side-nav {{
  display: grid;
  gap: 7px;
}}
.side-nav button {{
  width: 58px;
  min-height: 58px;
  display: grid;
  place-items: center;
  gap: 3px;
  border: 1px solid transparent;
  border-radius: 14px;
  padding: 6px 4px;
  background: transparent;
  color: #475569;
  font-size: 10px;
  font-weight: 700;
  line-height: 1.1;
  cursor: pointer;
  box-shadow: none;
}}
.side-nav button span:last-child {{
  display: block;
  max-width: 52px;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: normal;
  text-align: center;
}}
.side-icon {{
  font-size: 19px;
  line-height: 1;
}}
.side-icon svg {{
  display: block;
  width: 21px;
  height: 21px;
  stroke: currentColor;
  fill: none;
  stroke-width: 2;
  stroke-linecap: round;
  stroke-linejoin: round;
}}
.side-nav button:hover {{
  background: #eff6ff;
  color: #1d4ed8;
  filter: none;
}}
body.home-mode #sideHomeBtn,
body.module-mode #sideLandTransferBtn,
body.docs-mode #sideLibraryBtn,
body.webgis-mode #sideWebGisBtn {{
  border-color: #bfdbfe;
  background: linear-gradient(135deg, #eff6ff, #dbeafe);
  color: #1d4ed8;
  box-shadow: inset 0 0 0 1px rgba(255,255,255,0.72);
}}
body.webgis-mode #sideLandTransferBtn,
body.webgis-mode #sideLibraryBtn,
body.webgis-mode #sideWebGisBtn {{
  display: none;
}}
.status {{
  display: flex;
  flex: 1 1 260px;
  min-width: 220px;
  gap: 8px;
  flex-wrap: wrap;
  align-items: center;
  color: var(--muted);
  font-size: 12px;
}}
.quick-save {{
  flex: 0 0 auto;
  margin-left: auto;
}}
.home-page {{
  min-height: calc(100vh - 74px);
  margin: 14px 14px 14px 110px;
  border: 1px solid rgba(148, 163, 184, 0.24);
  border-radius: 18px;
  background:
    linear-gradient(135deg, rgba(11, 39, 89, 0.72), rgba(14, 116, 144, 0.42)),
    url("{home_bg_data_url}") center / cover no-repeat,
    linear-gradient(135deg, #0752b7, #52c7e8);
  box-shadow: inset 0 0 0 1px rgba(255, 255, 255, 0.18), 0 20px 50px rgba(15, 23, 42, 0.18);
  overflow: hidden;
}}
.dashboard {{
  min-height: calc(100vh - 104px);
  display: flex;
  flex-direction: column;
  justify-content: space-between;
  gap: 30px;
  padding: clamp(22px, 4vw, 52px);
}}
.dashboard-hero {{
  max-width: 880px;
  color: #ffffff;
  text-shadow: 0 2px 12px rgba(15, 23, 42, 0.22);
}}
.dashboard-eyebrow {{
  display: inline-flex;
  align-items: center;
  min-height: 30px;
  padding: 4px 11px;
  border: 1px solid rgba(255,255,255,0.36);
  border-radius: 999px;
  background: rgba(255,255,255,0.16);
  font-size: 12px;
  font-weight: 800;
  letter-spacing: 0.03em;
  text-transform: uppercase;
  backdrop-filter: blur(10px);
}}
.dashboard h1 {{
  margin: 18px 0 10px;
  font-size: clamp(34px, 5vw, 58px);
  line-height: 1.04;
  letter-spacing: 0;
}}
.dashboard-subtitle {{
  margin: 0;
  max-width: 760px;
  color: rgba(255,255,255,0.92);
  font-size: clamp(16px, 2vw, 20px);
  line-height: 1.55;
}}
.dashboard-metrics {{
  display: flex;
  flex-wrap: wrap;
  gap: 10px;
  margin-top: 20px;
}}
.dashboard-metric {{
  display: inline-flex;
  align-items: center;
  min-height: 38px;
  padding: 7px 12px;
  border: 1px solid rgba(255,255,255,0.35);
  border-radius: 999px;
  background: rgba(255,255,255,0.16);
  color: #ffffff;
  font-size: 13px;
  font-weight: 800;
  backdrop-filter: blur(10px);
}}
.module-grid {{
  display: grid;
  grid-template-columns: repeat(3, minmax(220px, 1fr));
  gap: 16px;
  align-items: stretch;
}}
.module-card {{
  display: grid;
  grid-template-rows: auto auto 1fr auto;
  min-height: 238px;
  padding: 20px;
  border: 1px solid rgba(255,255,255,0.44);
  border-radius: 18px;
  background: rgba(255,255,255,0.92);
  color: #0f172a;
  box-shadow: 0 20px 44px rgba(15, 23, 42, 0.20);
  backdrop-filter: blur(14px);
}}
.module-card-icon {{
  width: 52px;
  height: 52px;
  display: grid;
  place-items: center;
  border-radius: 16px;
  background: linear-gradient(135deg, #dbeafe, #eff6ff);
  color: #1d4ed8;
  font-size: 24px;
}}
.module-card-tag {{
  justify-self: start;
  margin-top: 14px;
  min-height: 26px;
  padding: 4px 9px;
  border: 1px solid #bfdbfe;
  border-radius: 999px;
  background: #eff6ff;
  color: #1d4ed8;
  font-size: 12px;
  font-weight: 800;
}}
.module-card h2 {{
  margin: 16px 0 8px;
  color: #0f2f57;
  font-size: 21px;
  line-height: 1.25;
}}
.module-card p {{
  margin: 0;
  color: #475569;
  font-size: 14px;
  line-height: 1.55;
}}
.module-card button {{
  justify-self: start;
  margin-top: 18px;
  min-height: 38px;
  border-radius: 11px;
  padding: 0 15px;
}}
.dashboard-footer {{
  color: rgba(255,255,255,0.78);
  font-size: 12px;
  text-align: right;
}}
body.home-mode .module-only,
body.home-mode .table-wrap,
body.home-mode #importLog {{
  display: none;
}}
body.docs-mode .module-only,
body.webgis-mode .module-only,
body.docs-mode .table-wrap,
body.webgis-mode .table-wrap,
body.docs-mode #importLog,
body.webgis-mode #importLog {{
  display: none;
}}
body.module-mode .home-page,
body.module-mode .docs-page,
body.module-mode .webgis-page,
body.home-mode .docs-page,
body.home-mode .webgis-page,
body.docs-mode .home-page,
body.docs-mode .webgis-page,
body.webgis-mode .home-page,
body.webgis-mode .docs-page {{
  display: none;
}}
.docs-page {{
  min-height: calc(100vh - 74px);
  margin: 14px 14px 14px 110px;
  border: 1px solid rgba(148, 163, 184, 0.28);
  border-radius: 18px;
  background: rgba(255, 255, 255, 0.92);
  box-shadow: 0 18px 42px rgba(15, 23, 42, 0.12);
  overflow: hidden;
}}
.webgis-page {{
  height: calc(100vh - 82px);
  min-height: 0;
  margin: 10px 14px 10px 110px;
  border: 1px solid rgba(148, 163, 184, 0.28);
  border-radius: 18px;
  background: #ffffff;
  box-shadow: 0 18px 42px rgba(15, 23, 42, 0.10);
  overflow: hidden;
}}
{WEBGIS_CSS}
.library-shell {{
  min-height: calc(100vh - 104px);
  padding: 18px;
}}
.library-head {{
  display: flex;
  align-items: flex-start;
  justify-content: space-between;
  gap: 14px;
  padding: 16px;
  border: 1px solid rgba(148, 163, 184, 0.28);
  border-radius: 10px;
  background: linear-gradient(135deg, rgba(240, 253, 250, 0.98), rgba(239, 246, 255, 0.98));
}}
.library-head h1 {{
  margin: 0;
  color: #0f172a;
  font-size: 23px;
  line-height: 1.2;
}}
.library-head p {{
  margin: 6px 0 0;
  max-width: 760px;
  color: #475569;
  font-size: 13px;
}}
.library-head-actions,
.reader-tools,
.library-admin-actions {{
  display: flex;
  align-items: center;
  flex-wrap: wrap;
  gap: 8px;
}}
.library-session-badge {{
  display: inline-flex;
  align-items: center;
  min-height: 28px;
  padding: 4px 9px;
  border: 1px solid #99f6e4;
  border-radius: 999px;
  color: #115e59;
  background: #f0fdfa;
  font-size: 12px;
  font-weight: 700;
}}
.library-session-badge.admin {{
  border-color: #99f6e4;
  color: #115e59;
  background: #f0fdfa;
}}
.library-session-badge.guest {{
  border-color: #bfdbfe;
  color: #1d4ed8;
  background: #eff6ff;
}}
.library-session-hint {{
  display: inline-flex;
  align-items: center;
  min-height: 28px;
  padding: 4px 9px;
  border: 1px solid #e2e8f0;
  border-radius: 999px;
  background: #ffffff;
  color: #475569;
  font-size: 12px;
  font-weight: 700;
}}
.library-controls {{
  display: grid;
  grid-template-columns: minmax(220px, 1fr) minmax(150px, 220px) minmax(130px, 180px) auto;
  gap: 9px;
  margin: 14px 0;
  padding: 10px;
  border: 1px solid rgba(148, 163, 184, 0.28);
  border-radius: 10px;
  background: #fff;
}}
.library-controls input,
.library-controls select,
.library-access input,
.library-admin input,
.library-admin select,
.library-admin textarea,
.reader-page-input {{
  height: 34px;
  border: 1px solid #cbd5e1;
  border-radius: 7px;
  padding: 5px 9px;
  color: #0f172a;
  background: #fff;
  font-size: 13px;
}}
.library-admin textarea {{
  min-height: 72px;
  resize: vertical;
}}
.library-grid {{
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(238px, 260px));
  gap: 14px;
  align-items: stretch;
}}
.library-card {{
  display: flex;
  flex-direction: column;
  width: 100%;
  height: 388px;
  min-width: 0;
  overflow: hidden;
  border: 1px solid rgba(148, 163, 184, 0.30);
  border-radius: 10px;
  background: #fff;
  box-shadow: 0 12px 28px rgba(15, 23, 42, 0.10);
}}
.library-cover {{
  position: relative;
  height: 158px;
  min-height: 158px;
  max-height: 158px;
  overflow: hidden;
  background: linear-gradient(135deg, #e0f2fe, #dcfce7);
  display: grid;
  place-items: center;
}}
.library-cover img {{
  display: block;
  width: 100%;
  height: 100%;
  object-fit: cover;
}}
.library-cover::after {{
  content: "";
  position: absolute;
  inset: 0;
  pointer-events: none;
  background: linear-gradient(180deg, rgba(255,255,255,0.04), rgba(15,23,42,0.08));
}}
.library-cover-placeholder {{
  max-width: 100%;
  padding: 12px;
  color: #0f172a;
  font-size: 15px;
  font-weight: 700;
  line-height: 1.3;
  text-align: center;
  overflow: hidden;
  overflow-wrap: anywhere;
  display: -webkit-box;
  -webkit-line-clamp: 3;
  -webkit-box-orient: vertical;
}}
.library-card-body {{
  display: flex;
  flex-direction: column;
  gap: 7px;
  flex: 1;
  min-width: 0;
  padding: 12px;
}}
.library-card h3 {{
  margin: 0;
  color: #0f172a;
  font-size: 15px;
  line-height: 1.3;
  min-height: 39px;
  overflow: hidden;
  overflow-wrap: anywhere;
  display: -webkit-box;
  -webkit-line-clamp: 2;
  -webkit-box-orient: vertical;
}}
.library-meta {{
  display: flex;
  flex-wrap: wrap;
  gap: 5px;
  color: #475569;
  font-size: 12px;
  min-width: 0;
}}
.library-author {{
  display: block;
  white-space: nowrap;
  overflow: hidden;
  text-overflow: ellipsis;
}}
.library-pill {{
  display: inline-flex;
  align-items: center;
  max-width: 100%;
  min-height: 22px;
  padding: 2px 7px;
  border: 1px solid #cbd5e1;
  border-radius: 999px;
  background: #f8fafc;
  overflow: hidden;
  text-overflow: ellipsis;
  white-space: nowrap;
}}
.library-description {{
  flex: 1;
  color: #475569;
  font-size: 13px;
  line-height: 1.45;
  overflow: hidden;
  overflow-wrap: anywhere;
  display: -webkit-box;
  -webkit-line-clamp: 3;
  -webkit-box-orient: vertical;
}}
.library-read-btn {{
  margin-top: auto;
}}
.library-empty {{
  padding: 28px;
  border: 1px dashed #cbd5e1;
  border-radius: 10px;
  color: #64748b;
  text-align: center;
  background: #f8fafc;
}}
.library-admin,
.library-access,
.pdf-reader {{
  position: fixed;
  inset: 88px 18px 18px;
  z-index: 120;
  overflow: auto;
  border: 1px solid rgba(100, 116, 139, 0.34);
  border-radius: 12px;
  background: #ffffff;
  box-shadow: 0 24px 70px rgba(15, 23, 42, 0.30);
}}
.library-admin[hidden],
.library-access[hidden],
.pdf-reader[hidden] {{
  display: none;
}}
.library-access {{
  z-index: 130;
  display: grid;
  place-items: center;
  background: rgba(15, 23, 42, 0.22);
}}
.library-access-card {{
  width: min(420px, calc(100vw - 32px));
  padding: 18px;
  border: 1px solid rgba(148, 163, 184, 0.36);
  border-radius: 12px;
  background: #ffffff;
  box-shadow: 0 24px 70px rgba(15, 23, 42, 0.26);
}}
.library-access-card h2 {{
  margin: 0;
  color: #0f172a;
  font-size: 18px;
}}
.library-access-card p {{
  margin: 8px 0 12px;
  color: #475569;
  font-size: 13px;
  line-height: 1.45;
}}
.library-access-form {{
  display: grid;
  gap: 10px;
}}
.library-access-form label {{
  display: grid;
  gap: 4px;
  color: #475569;
  font-size: 12px;
}}
.library-access-hint {{
  padding: 10px 12px;
  border: 1px solid #dbeafe;
  border-radius: 8px;
  background: #eff6ff;
}}
.library-admin-inner {{
  display: grid;
  grid-template-columns: minmax(300px, 420px) minmax(0, 1fr);
  gap: 14px;
  padding: 14px;
}}
.library-card-status {{
  display: inline-flex;
  align-self: flex-start;
  min-height: 23px;
  padding: 3px 8px;
  border: 1px solid #cbd5e1;
  border-radius: 999px;
  background: #f8fafc;
  color: #475569;
  font-size: 11px;
  font-weight: 800;
}}
.library-admin-toolbar {{
  position: sticky;
  top: 0;
  z-index: 2;
  display: flex;
  align-items: center;
  gap: 12px;
  padding: 14px;
  border-bottom: 1px solid #e2e8f0;
  background: rgba(255, 255, 255, 0.96);
}}
.library-admin-toolbar h2 {{
  flex: 1;
  margin: 0;
  color: #0f172a;
  font-size: 18px;
}}
.library-admin-card {{
  border: 1px solid rgba(148, 163, 184, 0.30);
  border-radius: 10px;
  padding: 12px;
  background: #f8fafc;
}}
.library-admin-card[hidden] {{
  display: none;
}}
.library-admin-status {{
  padding: 10px 12px;
  border: 1px solid #bbf7d0;
  border-radius: 8px;
  color: #166534;
  background: #f0fdf4;
  font-size: 13px;
  line-height: 1.45;
}}
.library-admin-status.error {{
  border-color: #fecaca;
  color: #991b1b;
  background: #fff1f2;
}}
.library-admin-card h2,
.library-admin-card h3 {{
  margin: 0 0 10px;
  color: #0f172a;
  font-size: 16px;
}}
.library-admin-form {{
  display: grid;
  gap: 9px;
}}
.library-admin-form label {{
  display: grid;
  gap: 4px;
  color: #475569;
  font-size: 12px;
}}
.library-admin-table {{
  width: 100%;
  border-collapse: collapse;
  font-size: 12px;
}}
.library-admin-table th,
.library-admin-table td {{
  border-bottom: 1px solid #e2e8f0;
  padding: 8px 6px;
  text-align: left;
  vertical-align: top;
}}
.library-admin-table th {{
  color: #334155;
  background: #f1f5f9;
}}
.pdf-reader {{
  display: flex;
  flex-direction: column;
  background: #f8fafc;
}}
.reader-topbar {{
  position: sticky;
  top: 0;
  z-index: 2;
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 10px;
  padding: 10px;
  border-bottom: 1px solid #cbd5e1;
  background: rgba(255, 255, 255, 0.96);
}}
.reader-title {{
  min-width: 180px;
  color: #0f172a;
  font-size: 14px;
  font-weight: 700;
}}
.reader-page-input {{
  width: 72px;
  text-align: center;
}}
.reader-notice {{
  padding: 9px 14px;
  color: #7a271a;
  border-bottom: 1px solid #fed7aa;
  background: #fff7ed;
  font-size: 13px;
}}
.pdf-stage {{
  flex: 1;
  overflow: auto;
  display: grid;
  place-items: start center;
  padding: 18px;
  user-select: none;
  -webkit-user-select: none;
}}
.pdf-canvas-wrap {{
  position: relative;
  max-width: 100%;
  padding: 10px;
  border: 1px solid #cbd5e1;
  border-radius: 8px;
  background: #fff;
  box-shadow: 0 16px 40px rgba(15, 23, 42, 0.18);
}}
#pdfCanvas {{
  display: block;
  max-width: 100%;
  height: auto;
  user-select: none;
  -webkit-user-select: none;
}}
@media (max-width: 820px) {{
  .library-head {{
    flex-direction: column;
  }}
  .library-controls {{
    grid-template-columns: 1fr;
  }}
  .library-admin,
  .library-access,
  .pdf-reader {{
    inset: 74px 8px 8px;
  }}
  .library-admin-inner {{
    grid-template-columns: 1fr;
  }}
  .reader-topbar {{
    align-items: flex-start;
    flex-direction: column;
  }}
}}
.badge {{
  display: inline-flex;
  align-items: center;
  min-height: 24px;
  padding: 4px 8px;
  border: 1px solid var(--line);
  background: rgba(248, 250, 252, 0.88);
  border-radius: 6px;
}}
.badge.warn {{
  color: #7a271a;
  border-color: #f4b0a1;
  background: #fff1ed;
}}
.actions {{
  display: flex;
  flex: 1 1 100%;
  flex-wrap: wrap;
  justify-content: flex-start;
  gap: 8px;
  align-items: flex-start;
}}
.tool-group {{
  position: relative;
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 0;
  min-height: 36px;
  padding: 0;
  border: 1px solid rgba(148, 163, 184, 0.36);
  border-radius: 8px;
  background: rgba(255, 255, 255, 0.72);
}}
.tool-group-title {{
  height: 34px;
  border: 0;
  border-radius: 8px;
  background: transparent;
  color: #334155;
  font-size: 12px;
  font-weight: 700;
  line-height: 34px;
  text-transform: uppercase;
  padding: 0 11px;
  cursor: pointer;
  box-shadow: none;
}}
.tool-group-title::after {{
  content: "▾";
  margin-left: 7px;
  font-size: 10px;
  color: #64748b;
}}
.tool-group.open .tool-group-title::after {{
  content: "▴";
}}
.tool-items {{
  position: absolute;
  top: calc(100% + 6px);
  left: 0;
  z-index: 85;
  display: none;
  flex-wrap: wrap;
  gap: 6px;
  min-width: 260px;
  max-width: min(520px, calc(100vw - 28px));
  padding: 8px;
  border: 1px solid rgba(100, 116, 139, 0.34);
  border-radius: 8px;
  background: rgba(255, 255, 255, 0.98);
  box-shadow: 0 18px 36px rgba(15, 23, 42, 0.18);
}}
.tool-group.open .tool-items {{
  display: flex;
}}
.tool-group:nth-last-child(-n+2) .tool-items {{
  left: auto;
  right: 0;
}}
.project-items {{
  min-width: min(420px, calc(100vw - 28px));
  gap: 10px;
}}
.project-section {{
  display: grid;
  grid-template-columns: 1fr;
  gap: 7px;
  width: 100%;
  padding: 8px;
  border: 1px solid rgba(226, 232, 240, 0.9);
  border-radius: 8px;
  background: #f8fafc;
}}
.project-section strong {{
  color: #0f172a;
  font-size: 12px;
}}
.project-field {{
  display: grid;
  grid-template-columns: 135px minmax(0, 1fr);
  gap: 8px;
  align-items: center;
  color: #475569;
  font-size: 12px;
}}
.project-field input {{
  min-width: 0;
  width: 100%;
  height: 30px;
  border: 1px solid #cbd5e1;
  border-radius: 6px;
  padding: 4px 8px;
  background: #fff;
  color: #0f172a;
  font-size: 12px;
}}
.project-actions {{
  display: flex;
  justify-content: flex-end;
  align-items: center;
  gap: 8px;
  width: 100%;
}}
.project-actions button {{
  min-width: 112px;
}}
.project-db-status {{
  flex: 1 1 auto;
  min-width: 160px;
  color: #64748b;
  font-size: 12px;
}}
.sample-downloads {{
  position: relative;
  display: flex;
  flex: 0 0 auto;
  justify-content: flex-start;
  flex-wrap: nowrap;
  gap: 0;
  align-items: center;
  font-size: 12px;
  color: #475569;
}}
.sample-downloads > span {{
  height: 34px;
  line-height: 34px;
  padding: 0 11px;
  border: 1px solid rgba(148, 163, 184, 0.36);
  border-radius: 8px;
  background: rgba(255, 255, 255, 0.72);
  color: #334155;
  font-size: 12px;
  font-weight: 700;
  text-transform: uppercase;
  cursor: pointer;
}}
.sample-downloads > span::after {{
  content: "▾";
  margin-left: 7px;
  font-size: 10px;
  color: #64748b;
}}
.sample-downloads.open > span::after {{
  content: "▴";
}}
.sample-items {{
  position: absolute;
  top: calc(100% + 6px);
  right: 0;
  z-index: 85;
  display: none;
  flex-wrap: wrap;
  gap: 6px;
  min-width: 280px;
  max-width: min(520px, calc(100vw - 28px));
  padding: 8px;
  border: 1px solid rgba(100, 116, 139, 0.34);
  border-radius: 8px;
  background: rgba(255, 255, 255, 0.98);
  box-shadow: 0 18px 36px rgba(15, 23, 42, 0.18);
}}
.sample-downloads.open .sample-items {{
  display: flex;
}}
.sample-downloads a {{
  color: #0f766e;
  text-decoration: none;
  border: 1px solid rgba(15, 118, 110, 0.24);
  background: rgba(240, 253, 250, 0.82);
  padding: 4px 8px;
}}
.sample-downloads a:hover {{
  border-color: rgba(15, 118, 110, 0.52);
  background: #ccfbf1;
}}
.search-box {{
  display: flex;
  align-items: center;
  gap: 4px;
  padding: 0;
  border: 1px solid rgba(100, 116, 139, 0.42);
  background: rgba(255, 255, 255, 0.68);
  border-radius: 6px;
}}
.search-box input {{
  width: 96px;
  height: 28px;
  min-height: 28px;
  border: 0;
  background: #ffffff;
  padding: 0 8px;
  text-align: left;
  text-transform: uppercase;
}}
.search-box button {{
  height: 28px;
  padding: 0 8px;
  border-top: 0;
  border-right: 0;
  border-bottom: 0;
}}
.import-options {{
  display: flex;
  align-items: center;
  gap: 6px;
  font-size: 12px;
  color: var(--muted);
}}
.import-options input,
.view-options input,
.report-option input {{
  width: auto;
  height: auto;
  min-height: 0;
}}
select {{
  height: 32px;
  border: 1px solid rgba(100, 116, 139, 0.62);
  border-radius: 6px;
  background: rgba(255, 255, 255, 0.92);
  color: #0f172a;
  padding: 0 8px;
  font-size: 13px;
}}
button {{
  height: 32px;
  border: 1px solid rgba(100, 116, 139, 0.62);
  border-radius: 6px;
  background: linear-gradient(180deg, #ffffff, #f8fafc);
  color: #0f172a;
  padding: 0 11px;
  font-size: 13px;
  cursor: pointer;
  box-shadow: 0 1px 2px rgba(15, 23, 42, 0.08);
}}
button.primary {{
  border-color: var(--accent);
  background: linear-gradient(180deg, #158176, #0f766e);
  color: #ffffff;
  font-weight: 700;
  box-shadow: 0 6px 14px rgba(15, 118, 110, 0.20);
}}
button:hover {{ filter: brightness(0.97); }}
.table-toolbar {{
  display: flex;
  flex-wrap: wrap;
  justify-content: space-between;
  align-items: center;
  gap: 8px 14px;
  margin: 12px 14px -4px 110px;
  padding: 8px 10px;
  border: 1px solid rgba(148, 163, 184, 0.34);
  border-radius: 8px;
  background: rgba(255, 255, 255, 0.86);
  box-shadow: 0 8px 22px rgba(15, 23, 42, 0.08);
}}
.legend {{
  display: flex;
  flex-wrap: wrap;
  align-items: center;
  gap: 8px 12px;
  color: #334155;
  font-size: 12px;
}}
.legend-item {{
  display: inline-flex;
  align-items: center;
  gap: 5px;
  white-space: nowrap;
}}
.swatch {{
  width: 16px;
  height: 16px;
  border: 1px solid #cbd5e1;
  border-radius: 4px;
}}
.swatch.input {{ background: var(--input); }}
.swatch.diagonal {{ background: var(--diagonal); }}
.swatch.auto {{ background: var(--auto); }}
.swatch.locked {{ background: #ffffff; }}
.swatch.plan-alert {{ background: #fee2e2; border-color: #fca5a5; }}
.view-options {{
  display: flex;
  align-items: center;
  gap: 8px;
  color: #334155;
  font-size: 12px;
}}
.table-wrap {{
  height: calc(100vh - 184px);
  min-height: 420px;
  overflow: auto;
  margin: 14px 14px 14px 110px;
  border: 1px solid rgba(148, 163, 184, 0.44);
  border-radius: 14px;
  background:
    linear-gradient(rgba(255, 255, 255, 0.94), rgba(255, 255, 255, 0.94)),
    repeating-linear-gradient(135deg, rgba(15, 118, 110, 0.05) 0 12px, rgba(37, 99, 235, 0.04) 12px 24px);
  box-shadow: 0 18px 40px rgba(15, 23, 42, 0.14);
  scrollbar-color: #94a3b8 #e2e8f0;
  scrollbar-width: thin;
}}
.table-wrap::-webkit-scrollbar {{
  width: 14px;
  height: 14px;
}}
.table-wrap::-webkit-scrollbar-track {{
  background: #e2e8f0;
}}
.table-wrap::-webkit-scrollbar-thumb {{
  background: #94a3b8;
  border: 3px solid #e2e8f0;
  border-radius: 999px;
}}
table {{
  border-collapse: collapse;
  table-layout: fixed;
  width: max-content;
  background: #ffffff;
}}
.sheet-title {{
  height: 42px;
  background: #f8fafc;
  color: #0f3d31;
  font-size: 18pt;
  font-weight: 700;
  text-align: center;
  vertical-align: middle;
  letter-spacing: 0;
  border: 1px solid #c8d0d9;
}}
td {{
  position: relative;
  background: #ffffff;
  padding: 4px 6px;
  line-height: 1.25;
  overflow: hidden;
  font-size: 12px;
}}
td[data-row="2"], td[data-row="3"] {{
  position: sticky;
  top: 0;
  z-index: 12;
  font-weight: 700;
  background: var(--header) !important;
  background-clip: padding-box;
}}
td[data-row="2"] {{ top: 0; }}
td[data-row="3"] {{ top: 30px; }}
td[data-col="1"], td[data-col="2"], td[data-col="3"], td[data-col="4"] {{
  position: sticky;
  z-index: 14;
  background: #ffffff;
  background-clip: padding-box;
}}
td[data-col="1"] {{ left: 0; }}
td[data-col="2"] {{ left: 48px; }}
td[data-col="3"] {{ left: 336px; }}
td[data-col="4"] {{ left: 400px; }}
td[data-row="2"][data-col="1"], td[data-row="2"][data-col="2"], td[data-row="2"][data-col="3"], td[data-row="2"][data-col="4"],
td[data-row="3"][data-col="1"], td[data-row="3"][data-col="2"], td[data-row="3"][data-col="3"], td[data-row="3"][data-col="4"] {{
  z-index: 30;
}}
td input {{
  width: 100%;
  height: 100%;
  min-height: 24px;
  border: 0;
  outline: 1px solid transparent;
  background: var(--input);
  text-align: right;
  font: inherit;
  color: #111827;
}}
td input:focus {{
  outline: 2px solid var(--accent);
  background: #ffffff;
}}
td[data-input="1"] {{ background: var(--input) !important; }}
td[data-auto="1"] {{ background-color: var(--locked); }}
td[data-auto="1"][style*="background"], td[data-input="1"] {{
  background-clip: padding-box;
}}
td.diagonal {{
  background: var(--diagonal) !important;
}}
td.diagonal input {{
  background: var(--diagonal) !important;
  font-weight: 700;
}}
td.diagonal input:focus {{
  background: #f1fff4 !important;
}}
.value {{
  display: block;
  text-align: right;
}}
body.hide-zero td.zero-cell .value {{
  visibility: hidden;
}}
body.hide-zero td.zero-cell input:not(:focus) {{
  color: transparent;
  caret-color: #111827;
}}
body.compact-zero-cols col.compact-hidden,
body.compact-zero-cols td.compact-hidden {{
  display: none !important;
}}
td.hover-row::after,
td.hover-col::after {{
  content: "";
  position: absolute;
  inset: 0;
  pointer-events: none;
  background: rgba(37, 99, 235, 0.055);
}}
td.hover-cell {{
  outline: 2px solid rgba(37, 99, 235, 0.75);
  outline-offset: -2px;
}}
td.warn {{
  background: #ffe4e6 !important;
  color: #7f1d1d !important;
}}
td.plan-target-alert {{
  background: #fee2e2 !important;
  color: #991b1b !important;
  font-weight: 700;
}}
td.plan-target-alert .value {{
  color: #991b1b !important;
}}
td.search-hit {{
  outline: 3px solid #2563eb !important;
  outline-offset: -3px;
  box-shadow: inset 0 0 0 2px rgba(255, 255, 255, 0.86), 0 0 0 2px rgba(37, 99, 235, 0.24);
  z-index: 35;
}}
.hidden-input {{ display: none; }}
.import-log {{
  border-bottom: 1px solid var(--line);
  background: rgba(248, 250, 252, 0.92);
  color: #17202a;
  padding: 8px 14px;
  font-size: 13px;
  line-height: 1.4;
}}
.import-log strong {{ font-weight: 700; }}
.import-log ul {{
  margin: 4px 0 0;
  padding-left: 18px;
}}
.report-panel {{
  position: fixed;
  inset: 64px 18px auto auto;
  z-index: 60;
  width: min(620px, calc(100vw - 36px));
}}
.report-card {{
  border: 1px solid rgba(100, 116, 139, 0.45);
  border-radius: 8px;
  background: #ffffff;
  box-shadow: 0 24px 48px rgba(15, 23, 42, 0.20);
  overflow: hidden;
}}
.report-head {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  gap: 10px;
  padding: 10px 12px;
  border-bottom: 1px solid var(--line);
  background: #f8fafc;
}}
.report-controls {{
  display: flex;
  flex-wrap: wrap;
  gap: 6px;
  padding: 10px 12px;
  border-bottom: 1px solid var(--line);
}}
.report-controls input {{
  width: 180px;
  height: 30px;
  min-height: 30px;
  border: 1px solid rgba(100, 116, 139, 0.62);
  background: #ffffff;
  padding: 0 8px;
  text-align: left;
}}
.report-controls input[type="number"] {{
  width: 112px;
}}
.report-options {{
  display: grid;
  grid-template-columns: repeat(auto-fill, minmax(178px, 1fr));
  gap: 6px;
  max-height: 360px;
  overflow: auto;
  padding: 10px 12px;
}}
.report-option {{
  display: flex;
  gap: 6px;
  align-items: flex-start;
  border: 1px solid #e2e8f0;
  background: #f8fafc;
  padding: 6px;
  font-size: 12px;
  line-height: 1.25;
}}
.report-option input {{
  width: auto;
  height: auto;
  min-height: 0;
  margin-top: 2px;
}}
.report-option span {{
  display: block;
}}
.ai-panel {{
  position: fixed;
  inset: 64px 18px auto auto;
  z-index: 6000;
  width: min(520px, calc(100vw - 36px));
  max-height: calc(100vh - 82px);
  isolation: isolate;
}}
.ai-card {{
  position: relative;
  z-index: 1;
  border: 1px solid rgba(100, 116, 139, 0.45);
  border-radius: 8px;
  background: #ffffff;
  box-shadow: 0 24px 48px rgba(15, 23, 42, 0.22);
  overflow: hidden;
  display: flex;
  flex-direction: column;
  max-height: calc(100vh - 82px);
}}
.ai-head {{
  display: flex;
  justify-content: space-between;
  align-items: center;
  gap: 10px;
  padding: 10px 12px;
  border-bottom: 1px solid var(--line);
  background: #f8fafc;
}}
.ai-messages {{
  display: flex;
  flex-direction: column;
  gap: 8px;
  max-height: 340px;
  flex: 1 1 auto;
  overflow: auto;
  padding: 12px;
  background: #f8fafc;
}}
.ai-message {{
  border: 1px solid #e2e8f0;
  background: #ffffff;
  padding: 8px 10px;
  line-height: 1.45;
  white-space: pre-wrap;
}}
.ai-message.user {{
  border-color: rgba(15, 118, 110, 0.28);
  background: #f0fdfa;
}}
.ai-status {{
  border-top: 1px solid var(--line);
  padding: 8px 12px;
  color: #475569;
  background: #ffffff;
  font-size: 12px;
}}
.ai-status.ready {{
  color: #047857;
  background: #ecfdf5;
}}
.ai-status.error {{
  color: #b91c1c;
  background: #fff1f2;
}}
.ai-controls {{
  display: flex;
  gap: 8px;
  padding: 10px 12px;
  border-top: 1px solid var(--line);
  background: #ffffff;
}}
.ai-controls textarea {{
  flex: 1 1 auto;
  min-height: 66px;
  min-width: 0;
  resize: vertical;
  border: 1px solid rgba(100, 116, 139, 0.62);
  padding: 8px;
  font-family: Arial, Helvetica, sans-serif;
  font-size: 13px;
}}
.ai-controls button {{
  height: 34px;
  flex: 0 0 auto;
  align-self: flex-end;
}}
body.ai-webgis-assistant .ai-panel {{
  inset: 86px 18px auto auto;
  z-index: 6000;
  width: min(430px, calc(100vw - 118px));
  max-height: calc(100vh - 108px);
}}
body.ai-webgis-assistant .ai-card {{
  border-radius: 12px;
  background: #ffffff;
  max-height: calc(100vh - 108px);
  box-shadow: 0 18px 42px rgba(15, 23, 42, 0.2);
}}
body.ai-webgis-assistant .ai-messages {{
  max-height: 210px;
}}
body.ai-webgis-assistant .ai-controls textarea {{
  min-height: 54px;
}}
@media (max-width: 900px) {{
  body.ai-webgis-assistant .ai-panel {{
    inset: 86px 12px auto 88px;
    width: auto;
  }}
}}
.startup-screen {{
  position: fixed;
  inset: 0;
  z-index: 9000;
  display: grid;
  place-items: center;
  overflow: hidden;
  color: #e5f4ff;
  background:
    radial-gradient(circle at 72% 42%, rgba(124, 58, 237, 0.28), transparent 0 22%, transparent 42%),
    radial-gradient(circle at 24% 78%, rgba(14, 165, 233, 0.2), transparent 0 24%, transparent 44%),
    linear-gradient(135deg, #07111f 0%, #0a1630 48%, #111827 100%);
  transition: opacity 0.36s ease, visibility 0.36s ease;
}}
.startup-screen[hidden] {{
  display: none;
}}
.startup-screen.is-hidden {{
  opacity: 0;
  visibility: hidden;
  pointer-events: none;
}}
.startup-grid {{
  position: absolute;
  inset: 0;
  background-image:
    linear-gradient(rgba(148, 163, 184, 0.12) 1px, transparent 1px),
    linear-gradient(90deg, rgba(148, 163, 184, 0.12) 1px, transparent 1px);
  background-size: 88px 88px;
  mask-image: linear-gradient(90deg, transparent 0%, #000 32%, #000 100%);
  animation: startupGridDrift 12s linear infinite;
}}
.startup-card {{
  position: relative;
  z-index: 1;
  width: min(560px, calc(100vw - 36px));
  padding: 30px;
  border: 1px solid rgba(125, 211, 252, 0.24);
  border-radius: 22px;
  background: rgba(8, 18, 34, 0.82);
  box-shadow: 0 28px 80px rgba(0, 0, 0, 0.42), inset 0 1px 0 rgba(255,255,255,0.08);
  backdrop-filter: blur(18px);
}}
.startup-brand {{
  display: inline-flex;
  align-items: center;
  gap: 10px;
  margin-bottom: 22px;
  color: #93c5fd;
  font-weight: 800;
  letter-spacing: 0.02em;
  text-transform: uppercase;
}}
.startup-brand-mark {{
  width: 34px;
  height: 34px;
  display: grid;
  place-items: center;
  border-radius: 11px;
  background: linear-gradient(135deg, #2563eb, #7c3aed);
  color: #fff;
  box-shadow: 0 12px 28px rgba(59, 130, 246, 0.35);
}}
.startup-title {{
  margin: 0;
  color: #f8fafc;
  font-size: clamp(26px, 4vw, 42px);
  line-height: 1.1;
  letter-spacing: 0;
}}
.startup-description {{
  max-width: 460px;
  margin: 14px 0 26px;
  color: #bfdbfe;
  font-size: 15px;
  line-height: 1.65;
}}
.startup-loader {{
  display: flex;
  align-items: center;
  gap: 12px;
  color: #c4b5fd;
  font-size: 13px;
  font-weight: 700;
  letter-spacing: 0.05em;
  text-transform: uppercase;
}}
.startup-spinner {{
  width: 22px;
  height: 22px;
  border: 3px solid rgba(147, 197, 253, 0.25);
  border-top-color: #60a5fa;
  border-right-color: #a78bfa;
  border-radius: 999px;
  animation: startupSpin 0.88s linear infinite;
}}
.startup-pulse {{
  position: absolute;
  right: -70px;
  bottom: -70px;
  width: 210px;
  height: 210px;
  border-radius: 999px;
  background: rgba(14, 165, 233, 0.18);
  filter: blur(4px);
  animation: startupPulse 2.8s ease-in-out infinite;
}}
body.startup-active {{
  overflow: hidden;
}}
@keyframes startupSpin {{
  to {{ transform: rotate(360deg); }}
}}
@keyframes startupPulse {{
  0%, 100% {{ transform: scale(0.92); opacity: 0.48; }}
  50% {{ transform: scale(1.08); opacity: 0.82; }}
}}
@keyframes startupGridDrift {{
  from {{ transform: translate3d(0, 0, 0); }}
  to {{ transform: translate3d(-88px, -88px, 0); }}
}}
@media print {{
  .appbar, .import-log {{ display: none; }}
  .table-wrap {{ height: auto; overflow: visible; }}
  tr:nth-child(2) td, tr:nth-child(3) td, td:nth-child(1), td:nth-child(2), td:nth-child(3), td:nth-child(4) {{
    position: static;
  }}
}}
{chr(10).join(css_rules)}
.table-wrap td {{
  font-size: 12px !important;
  min-height: 28px;
}}
.table-wrap tr {{
  height: 30px;
}}
.table-wrap td[data-row="2"],
.table-wrap td[data-row="3"] {{
  background: var(--header) !important;
  color: #0f172a !important;
  font-weight: 700 !important;
}}
.table-wrap td[data-col="1"],
.table-wrap td[data-col="2"],
.table-wrap td[data-col="3"],
.table-wrap td[data-col="4"] {{
  background-clip: padding-box !important;
}}
.table-wrap td[data-input="1"] {{
  background: var(--input) !important;
}}
.table-wrap td.diagonal,
.table-wrap td.diagonal input {{
  background: var(--diagonal) !important;
}}
.table-wrap td[data-auto="1"]:not(.diagonal) {{
  background-color: var(--locked) !important;
}}
@media (max-width: 980px) {{
  .appbar {{
    padding-left: 14px;
  }}
  .app-sidebar {{
    top: auto;
    left: 10px;
    right: 10px;
    bottom: 10px;
    width: auto;
    height: 76px;
    flex-direction: row;
    align-items: center;
    justify-content: center;
    border-radius: 18px;
  }}
  .side-logo {{
    display: none;
  }}
  .side-nav {{
    display: flex;
    justify-content: center;
    gap: 8px;
  }}
  .home-page,
  .docs-page,
  .webgis-page {{
    margin: 10px 10px 96px;
  }}
  body.webgis-mode .webgis-page {{
    height: calc(100vh - 168px);
    min-height: 0;
  }}
  .table-toolbar {{
    margin: 10px 10px -2px;
  }}
  .table-wrap {{
    height: calc(100vh - 248px);
    margin: 12px 10px 96px;
  }}
  .module-grid {{
    grid-template-columns: 1fr;
  }}
  .dashboard {{
    padding: 22px;
  }}
}}
@media (max-width: 700px) {{
  html,
  body {{
    width: 100%;
    overflow-x: hidden;
  }}
  body:not(.webgis-mode) {{
    overflow-y: auto;
  }}
  .appbar {{
    min-height: 64px;
    padding: 8px 10px;
    gap: 8px;
  }}
  body:not(.home-mode) .appbar,
  body.webgis-mode .appbar {{
    height: auto;
    min-height: 56px;
    overflow: visible;
  }}
  .brand {{
    flex: 1 1 auto;
    min-width: 0;
  }}
  .brand-logo {{
    width: 34px;
    height: 34px;
  }}
  .title {{
    font-size: 13px;
    line-height: 1.12;
  }}
  .subtitle,
  .designer {{
    font-size: 10px;
  }}
  .module-label {{
    max-width: calc(100vw - 66px);
    font-size: 11px;
    padding: 3px 8px;
  }}
  .app-sidebar {{
    left: 8px;
    right: 8px;
    bottom: 8px;
    height: 64px;
    padding: 7px;
    border-radius: 16px;
  }}
  body.webgis-mode .app-sidebar {{
    right: auto;
    width: 64px;
  }}
  .side-nav button {{
    width: 64px;
    min-height: 50px;
    font-size: 10px;
  }}
  .home-page,
  .docs-page,
  .webgis-page {{
    margin: 8px 8px 80px;
    border-radius: 14px;
  }}
  .dashboard {{
    min-height: calc(100dvh - 152px);
    padding: 18px;
    gap: 18px;
  }}
  .dashboard h1 {{
    font-size: 30px;
  }}
  .dashboard-subtitle {{
    font-size: 14px;
  }}
  .dashboard-metrics {{
    gap: 6px;
  }}
  .dashboard-metric {{
    min-height: 32px;
    font-size: 11px;
  }}
  .module-grid {{
    grid-template-columns: 1fr;
    gap: 12px;
  }}
  .module-card {{
    min-height: 188px;
    padding: 16px;
  }}
  .module-card h2 {{
    font-size: 18px;
  }}
  .module-card p {{
    font-size: 13px;
  }}
  .actions {{
    width: 100%;
    flex-wrap: nowrap;
    overflow-x: auto;
    padding-bottom: 4px;
  }}
  .tool-group {{
    flex: 0 0 auto;
  }}
  .tool-items {{
    position: fixed;
    left: 8px !important;
    right: 8px !important;
    top: 66px;
    max-height: calc(100dvh - 154px);
    overflow: auto;
    min-width: 0;
    max-width: none;
  }}
  .status {{
    flex: 1 1 100%;
    min-width: 0;
    overflow-x: auto;
    flex-wrap: nowrap;
    padding-bottom: 2px;
  }}
  .quick-save {{
    margin-left: 0;
  }}
  .table-toolbar {{
    margin: 8px 8px 6px;
    align-items: stretch;
  }}
  .legend {{
    width: 100%;
    flex-wrap: nowrap;
    overflow-x: auto;
    padding-bottom: 4px;
  }}
  .view-options {{
    width: 100%;
    justify-content: space-between;
  }}
  .table-wrap {{
    height: calc(100dvh - 214px);
    min-height: 320px;
    margin: 8px 8px 80px;
    border-radius: 10px;
  }}
  .table-wrap td {{
    font-size: 11px !important;
  }}
  .docs-page {{
    min-height: calc(100dvh - 152px);
  }}
  .library-shell {{
    padding: 10px;
  }}
  .library-head {{
    align-items: stretch;
    gap: 10px;
  }}
  .library-head h1 {{
    font-size: 20px;
  }}
  .library-head-actions {{
    justify-content: flex-start;
    overflow-x: auto;
    flex-wrap: nowrap;
    padding-bottom: 4px;
  }}
  .library-controls {{
    grid-template-columns: 1fr;
    gap: 8px;
  }}
  .library-grid {{
    grid-template-columns: 1fr;
  }}
  .library-card {{
    min-height: 0;
  }}
  .library-admin,
  .library-access,
  .pdf-reader {{
    inset: 64px 8px 78px;
  }}
  .library-admin-toolbar,
  .reader-toolbar {{
    overflow-x: auto;
    flex-wrap: nowrap;
    justify-content: flex-start;
  }}
  .library-admin-table {{
    min-width: 760px;
  }}
  body.webgis-mode {{
    height: 100dvh;
    overflow: hidden;
  }}
  body.webgis-mode .webgis-page {{
    height: calc(100dvh - 138px);
    min-height: 0;
  }}
  .webgis-shell {{
    height: 100%;
  }}
  .webgis-topbar {{
    grid-template-columns: 1fr;
    gap: 8px;
    padding: 9px 10px;
    max-height: 190px;
    overflow: auto;
  }}
  .webgis-title strong {{
    font-size: 15px;
  }}
  .webgis-title span {{
    font-size: 11px;
  }}
  .webgis-stats {{
    margin-top: 6px;
    flex-wrap: nowrap;
    overflow-x: auto;
    padding-bottom: 2px;
  }}
  .webgis-stat {{
    flex: 0 0 auto;
  }}
  .webgis-search {{
    gap: 6px;
  }}
  .webgis-search input {{
    min-width: 0;
  }}
  .webgis-search button {{
    flex: 0 0 auto;
  }}
  .webgis-actions {{
    justify-content: flex-start;
    flex-wrap: nowrap;
    overflow-x: auto;
    padding-bottom: 3px;
  }}
  .webgis-actions button {{
    flex: 0 0 auto;
  }}
  .webgis-workspace,
  .webgis-page.layers-collapsed .webgis-workspace,
  .webgis-page.info-collapsed .webgis-workspace,
  .webgis-page.layers-collapsed.info-collapsed .webgis-workspace {{
    display: flex;
    flex-direction: column;
    gap: 8px;
    padding: 8px;
    overflow: auto;
  }}
  .webgis-sidebar,
  .webgis-info {{
    display: block;
    height: auto;
    max-height: none;
    overflow: visible;
  }}
  .webgis-page.layers-collapsed .webgis-sidebar .webgis-panel-body,
  .webgis-page.layers-collapsed .webgis-sidebar .webgis-admin-panel,
  .webgis-page.info-collapsed .webgis-info .webgis-panel-body,
  .webgis-page.info-collapsed .webgis-info .webgis-panel:not(:first-child) {{
    display: block;
  }}
  .webgis-page.layers-collapsed .webgis-sidebar h2,
  .webgis-page.info-collapsed .webgis-info h2 {{
    writing-mode: horizontal-tb;
  }}
  .webgis-map-panel {{
    order: 1;
    flex: 0 0 auto;
    height: min(58dvh, 430px);
    min-height: 330px;
  }}
  .webgis-map {{
    height: 100%;
    min-height: 330px;
  }}
  .webgis-sidebar {{
    order: 2;
  }}
  .webgis-info {{
    order: 3;
  }}
  .webgis-sidebar > .webgis-panel:first-child .webgis-panel-body,
  .webgis-info > .webgis-panel:first-child .webgis-panel-body {{
    max-height: 210px;
  }}
  .webgis-map-tools {{
    top: 8px;
    left: 8px;
    right: 8px;
    transform: none;
    justify-content: flex-start;
    flex-wrap: nowrap;
    overflow-x: auto;
    padding: 8px;
  }}
  .webgis-map-tools button {{
    flex: 0 0 auto;
  }}
  .webgis-attr-panel {{
    position: fixed;
    z-index: 760;
    inset: 64px 8px 78px;
    margin: 0;
  }}
  .webgis-attr-tools {{
    overflow-x: auto;
    flex-wrap: nowrap;
    justify-content: flex-start;
  }}
  .webgis-attr-wrap {{
    max-height: calc(100dvh - 176px);
  }}
  body.ai-webgis-assistant .ai-panel,
  .ai-panel {{
    inset: 70px 8px auto 8px;
    width: auto;
    max-height: calc(100dvh - 154px);
  }}
  body.webgis-mode .appbar,
  body.webgis-mode .app-sidebar,
  body.webgis-mode .webgis-topbar,
  body.webgis-mode .webgis-info,
  body.webgis-mode .webgis-attr-panel,
  body.webgis-mode .webgis-admin-panel {{
    display: none !important;
  }}
  body.webgis-mode .webgis-page {{
    position: fixed;
    inset: 0;
    z-index: 20;
    height: 100dvh;
    width: 100vw;
    margin: 0;
    border: 0;
    border-radius: 0;
    background: #dbeafe;
  }}
  body.webgis-mode .webgis-shell {{
    height: 100dvh;
    border-radius: 0;
  }}
  body.webgis-mode .webgis-workspace,
  body.webgis-mode .webgis-page.layers-collapsed .webgis-workspace,
  body.webgis-mode .webgis-page.info-collapsed .webgis-workspace,
  body.webgis-mode .webgis-page.layers-collapsed.info-collapsed .webgis-workspace {{
    position: relative;
    display: block;
    height: 100%;
    padding: 0;
    overflow: hidden;
  }}
  body.webgis-mode .webgis-map-panel {{
    position: absolute;
    inset: 0;
    height: 100%;
    min-height: 0;
    border: 0;
    border-radius: 0;
    box-shadow: none;
  }}
  body.webgis-mode .webgis-map {{
    height: 100%;
    min-height: 0;
  }}
  body.webgis-mode .webgis-sidebar {{
    position: absolute;
    z-index: 650;
    top: 10px;
    right: auto;
    left: 8px;
    width: min(218px, calc(100vw - 58px));
    max-height: min(34dvh, 238px);
    overflow: auto;
    display: block;
    border-radius: 10px;
    box-shadow: 0 12px 30px rgba(15, 47, 87, 0.20);
  }}
  body.webgis-mode .webgis-page.layers-collapsed .webgis-sidebar {{
    width: 38px;
    max-height: 38px;
    overflow: hidden;
  }}
  body.webgis-mode .webgis-sidebar > .webgis-panel:first-child {{
    display: block;
    min-height: 0;
    border-radius: 10px;
    box-shadow: none;
  }}
  body.webgis-mode .webgis-sidebar .webgis-panel-head {{
    min-height: 0;
    padding: 6px 7px;
    border-bottom-color: #dbe7f3;
  }}
  body.webgis-mode .webgis-sidebar .webgis-panel-head h2 {{
    writing-mode: horizontal-tb;
    font-size: 11px;
  }}
  body.webgis-mode .webgis-sidebar .webgis-panel-actions {{
    display: inline-flex !important;
    gap: 4px;
  }}
  body.webgis-mode #webgisFitAllBtn,
  body.webgis-mode #webgisToggleSidebarBtn {{
    display: inline-grid !important;
    width: 28px;
    min-width: 28px;
    height: 28px !important;
    padding: 0 !important;
    place-items: center;
    border-radius: 7px !important;
    font-size: 0;
  }}
  body.webgis-mode #webgisFitAllBtn::before {{
    content: "⌖";
    font-size: 16px;
    line-height: 1;
  }}
  body.webgis-mode #webgisToggleSidebarBtn {{
    font-size: 16px;
  }}
  body.webgis-mode .webgis-page.layers-collapsed .webgis-sidebar .webgis-panel-head {{
    min-height: 38px;
    padding: 4px;
    justify-content: center;
  }}
  body.webgis-mode .webgis-page.layers-collapsed .webgis-sidebar h2,
  body.webgis-mode .webgis-page.layers-collapsed #webgisFitAllBtn,
  body.webgis-mode .webgis-page.layers-collapsed .webgis-sidebar > .webgis-panel:first-child .webgis-panel-body {{
    display: none !important;
  }}
  body.webgis-mode .webgis-sidebar > .webgis-panel:first-child .webgis-panel-body {{
    display: block;
    max-height: none;
    padding: 6px;
    overflow: visible;
  }}
  body.webgis-mode .webgis-layer-list {{
    gap: 4px;
  }}
  body.webgis-mode .webgis-layer-item {{
    grid-template-columns: auto 1fr auto;
    gap: 5px;
    padding: 5px;
    border-radius: 8px;
    box-shadow: none;
    background: rgba(255, 255, 255, 0.96);
  }}
  body.webgis-mode .webgis-symbol {{
    width: 13px;
    height: 13px;
    border-radius: 4px;
  }}
  body.webgis-mode .webgis-layer-main label {{
    font-size: 10.5px;
    line-height: 1.2;
  }}
  body.webgis-mode .webgis-layer-count {{
    font-size: 9.5px;
  }}
  body.webgis-mode .webgis-layer-actions {{
    display: inline-flex;
  }}
  body.webgis-mode .webgis-layer-tools {{
    display: none;
  }}
  body.webgis-mode .webgis-layer-actions .webgis-icon-btn {{
    width: 26px;
    min-width: 26px;
    height: 26px !important;
    border-radius: 7px !important;
    font-size: 12px !important;
  }}
  body.webgis-mode .webgis-map-tools {{
    top: auto;
    bottom: 76px;
    left: 8px;
    right: auto;
    max-width: 44px;
    padding: 0;
    border: 0;
    border-radius: 8px;
    background: transparent;
    box-shadow: none;
    overflow: visible;
  }}
  body.webgis-mode .webgis-map-tools .webgis-tool-divider,
  body.webgis-mode .webgis-map-tools .webgis-tool-group:not(:first-child) {{
    display: none !important;
  }}
  body.webgis-mode #webgisLocateBtn {{
    width: 34px;
    height: 34px;
    padding: 0;
    border-radius: 6px;
    background: #fff;
    color: #0f2f57;
    font-size: 0;
    box-shadow: 0 4px 12px rgba(15, 47, 87, 0.22);
  }}
  body.webgis-mode #webgisLocateBtn::before {{
    content: "◎";
    font-size: 18px;
    line-height: 1;
  }}
  body.webgis-mode .webgis-measure-badge {{
    display: none;
  }}
  body.webgis-mode .webgis-coordinate-bar {{
    right: 8px;
    bottom: 8px;
    max-width: calc(100vw - 92px);
    padding: 5px 7px;
    font-size: 11px;
  }}
  body.webgis-mode .leaflet-control-zoom {{
    margin-top: 12px !important;
    margin-left: 8px !important;
  }}
  body.webgis-mode .leaflet-control-layers {{
    margin-top: 12px !important;
    margin-right: 8px !important;
  }}
}}
</style>
</head>
<body class="home-mode">
<section id="startupScreen" class="startup-screen" role="status" aria-live="polite" aria-label="Đang khởi động WebGIS">
  <div class="startup-grid" aria-hidden="true"></div>
  <div class="startup-card">
    <div class="startup-pulse" aria-hidden="true"></div>
    <div class="startup-brand"><span class="startup-brand-mark">DD</span> WebGIS</div>
    <h1 class="startup-title">Đang khởi động WebGIS</h1>
    <p class="startup-description">Máy chủ đang được đánh thức, vui lòng chờ trong giây lát...</p>
    <div class="startup-loader">
      <span class="startup-spinner" aria-hidden="true"></span>
      <span id="startupStatus">Đang kiểm tra máy chủ</span>
    </div>
  </div>
</section>
<header class="appbar">
  <div class="brand">
    <img class="brand-logo" src="{logo_data_url}" alt="Logo Nguyễn Quang Huy">
    <div class="brand-text">
      <div class="title">PHẦN MỀM QUẢN LÝ ĐẤT ĐAI</div>
      <div class="subtitle">Công cụ hỗ trợ lập, kiểm tra và xuất biểu chu chuyển sử dụng đất</div>
      <div id="activeModuleLabel" class="module-label">Trang chủ</div>
      <div class="designer">Designed by Nguyễn Quang Huy</div>
    </div>
  </div>
  <div class="status module-only">
    <span id="statusTotal" class="badge">Đang tính</span>
    <span id="statusRows" class="badge">0 lệch hàng</span>
    <span id="statusMissing" class="badge"></span>
  </div>
  <button class="primary quick-save module-only" id="saveBtn" type="button">Lưu</button>
  <div class="actions module-only">
    <div class="tool-group">
      <button class="tool-group-title" type="button">Dự án</button>
      <div class="tool-items project-items">
        <div class="project-section">
          <strong>Thiết lập đơn vị hành chính</strong>
          <label class="project-field">
            <span>Tên xã</span>
            <input id="projectCommune" type="text" placeholder="Ví dụ: xã An Bình">
          </label>
          <label class="project-field">
            <span>Tỉnh/thành</span>
            <input id="projectProvince" type="text" placeholder="Ví dụ: tỉnh Bắc Ninh">
          </label>
        </div>
        <div class="project-section">
          <strong>Thông tin quy hoạch</strong>
          <label class="project-field">
            <span>Quy hoạch kỳ trước</span>
            <input id="projectPreviousPlanYear" type="text" placeholder="Ví dụ: 2015-2025">
          </label>
          <label class="project-field">
            <span>Năm hiện trạng</span>
            <input id="projectCurrentYear" type="number" min="1900" max="2200" value="2020">
          </label>
          <label class="project-field">
            <span>Kỳ quy hoạch</span>
            <input id="projectPlanYear" type="text" placeholder="Ví dụ: 2025-2035" value="2020-2030">
          </label>
        </div>
        <div class="project-section">
          <strong>Cơ sở dữ liệu dự án (*.gtp)</strong>
          <button id="gtpOpenBtn" type="button">Add file GTP</button>
          <button id="gtpSetupBtn" type="button">Thiết lập nơi lưu GTP</button>
          <button id="gtpSaveBtn" type="button">Lưu vào file GTP</button>
          <span id="gtpStatus" class="project-db-status">Chưa thiết lập file GTP</span>
        </div>
        <div class="project-actions">
          <button id="projectConfirmBtn" class="primary" type="button">Xác nhận</button>
        </div>
      </div>
    </div>
    <div class="tool-group">
      <button class="tool-group-title" type="button">Nhập dữ liệu</button>
      <div class="tool-items">
        <button id="importCurrentBtn" type="button">Nhập hiện trạng XLSX</button>
        <button id="importGisBtn" type="button">Import bảng chồng xếp GIS</button>
        <button id="importPreviousPlanBtn" type="button">Import quy hoạch kỳ trước</button>
        <button id="loadBtn" type="button">Nhập JSON</button>
      </div>
    </div>
    <div class="tool-group">
      <button class="tool-group-title" type="button">Xử lý</button>
      <div class="tool-items">
        <button id="reportBtn" type="button">Xuất tăng/giảm</button>
        <select id="gisImportMode" title="Chế độ xử lý mã đất lạ">
          <option value="add" selected>Tự thêm mã mới</option>
          <option value="known">Chỉ mã đã có</option>
        </select>
        <label class="import-options" title="Áp dụng khi cột diện tích là m2">
          <input id="gisM2ToHa" type="checkbox" checked>
          m2 -> ha
        </label>
        <button id="clearBtn" type="button">Xóa nhập</button>
      </div>
    </div>
    <div class="tool-group">
      <button class="tool-group-title" type="button">Xuất file</button>
      <div class="tool-items">
        <button id="jsonBtn" type="button">Xuất JSON</button>
        <button id="xlsxBtn" type="button">Xuất XLSX</button>
        <button id="csvBtn" type="button">Xuất CSV</button>
        <button id="printBtn" type="button">In</button>
      </div>
    </div>
    <div class="tool-group">
      <button class="tool-group-title" type="button">Công cụ</button>
      <div class="tool-items">
        <button id="homeBtn" type="button">Màn chính</button>
        <div class="search-box">
          <input id="codeSearch" type="search" placeholder="Tìm mã" aria-label="Tìm mã đất">
          <button id="codeSearchBtn" type="button">Tìm</button>
        </div>
      </div>
    </div>
    <div class="tool-group">
      <button class="tool-group-title" type="button">AI</button>
      <div class="tool-items">
        <button id="aiBtn" type="button">Trợ lý AI</button>
      </div>
    </div>
    <div class="sample-downloads">
      <span>Tải file mẫu</span>
      <div class="sample-items">
        {sample_links_html}
      </div>
    </div>
  </div>
  <input id="fileInput" class="hidden-input" type="file" accept="application/json">
  <input id="currentXlsxInput" class="hidden-input" type="file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet">
  <input id="previousPlanXlsxInput" class="hidden-input" type="file" accept=".xlsx,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet">
  <input id="gisXlsxInput" class="hidden-input" type="file" accept=".xlsx,.xls,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.ms-excel">
  <input id="gtpInput" class="hidden-input" type="file" accept=".gtp,application/json">
</header>
<aside class="app-sidebar" aria-label="Điều hướng module">
  <div class="side-logo" aria-hidden="true">ĐĐ</div>
  <nav class="side-nav">
    <button id="sideHomeBtn" type="button" title="Trang chủ"><span class="side-icon"><svg viewBox="0 0 24 24" aria-hidden="true"><path d="M3 10.5 12 3l9 7.5"></path><path d="M5 10v10h14V10"></path><path d="M9 20v-6h6v6"></path></svg></span><span>Trang chủ</span></button>
    <button id="sideLandTransferBtn" type="button" title="Chu chuyển đất đai"><span class="side-icon"><svg viewBox="0 0 24 24" aria-hidden="true"><path d="M4 5h16"></path><path d="M4 12h16"></path><path d="M4 19h16"></path><path d="M8 5v14"></path><path d="M16 5v14"></path></svg></span><span>Chu chuyển</span></button>
    <button id="sideLibraryBtn" type="button" title="Thư viện tài liệu PDF"><span class="side-icon"><svg viewBox="0 0 24 24" aria-hidden="true"><path d="M6 3h9l3 3v15H6z"></path><path d="M14 3v4h4"></path><path d="M8.5 12h7"></path><path d="M8.5 16h5"></path></svg></span><span>Thư viện</span></button>
    <button id="sideWebGisBtn" type="button" title="WebGIS quản lý dữ liệu đất đai"><span class="side-icon"><svg viewBox="0 0 24 24" aria-hidden="true"><path d="M9 18 3 20V6l6-2 6 2 6-2v14l-6 2z"></path><path d="M9 4v14"></path><path d="M15 6v14"></path></svg></span><span>WebGIS</span></button>
  </nav>
</aside>
<main id="homePage" class="home-page" aria-label="Trang chính">
  <section class="dashboard">
    <div class="dashboard-hero">
      <span class="dashboard-eyebrow">Phần mềm đất đai</span>
      <h1>PHẦN MỀM QUẢN LÝ ĐẤT ĐAI</h1>
      <p class="dashboard-subtitle">Công cụ hỗ trợ lập, kiểm tra và xuất biểu chu chuyển sử dụng đất với các module chuyên biệt cho số liệu, tài liệu và bản đồ.</p>
      <div class="dashboard-metrics" aria-label="Điểm nhận diện chức năng">
        <span class="dashboard-metric">Biểu chu chuyển đất đai</span>
        <span class="dashboard-metric">Thư viện PDF bảo vệ</span>
        <span class="dashboard-metric">WebGIS dữ liệu đất đai</span>
      </div>
    </div>
    <div class="module-grid" aria-label="Danh sách module">
      <article class="module-card">
        <div class="module-card-icon" aria-hidden="true">▦</div>
        <span class="module-card-tag">Ma trận chu chuyển</span>
        <h2>Chu chuyển đất đai</h2>
        <p>Nhập hiện trạng, import bảng chồng xếp GIS, kiểm tra ma trận, xuất CSV/XLSX/Word và quản lý dữ liệu dự án.</p>
        <button id="homeLandTransferBtn" class="primary" type="button">Mở biểu chu chuyển</button>
      </article>
      <article class="module-card">
        <div class="module-card-icon" aria-hidden="true">□</div>
        <span class="module-card-tag">Tài liệu PDF</span>
        <h2>Thư viện tài liệu PDF</h2>
        <p>Quản lý, phân loại và đọc tài liệu PDF trực tuyến theo quyền khách hoặc admin trong không gian gọn gàng.</p>
        <button id="homeLibraryBtn" class="primary" type="button">Mở thư viện</button>
      </article>
      <article class="module-card">
        <div class="module-card-icon" aria-hidden="true">◎</div>
        <span class="module-card-tag">Bản đồ GIS</span>
        <h2>WebGIS quản lý dữ liệu đất đai</h2>
        <p>Hiển thị lớp bản đồ, tra cứu thửa đất, import GeoJSON và xem thuộc tính không gian trong giao diện GIS chuyên nghiệp.</p>
        <button id="homeWebGisBtn" class="primary" type="button">Mở WebGIS</button>
      </article>
    </div>
    <footer class="dashboard-footer">Designed by Nguyễn Quang Huy</footer>
  </section>
</main>
<main id="documentLibraryPage" class="docs-page" aria-label="Thư viện tài liệu">
  <section class="library-shell">
    <div class="library-head">
      <div>
        <h1>Thư viện số tài liệu PDF</h1>
      </div>
      <div class="library-head-actions">
        <span id="librarySessionBadge" class="library-session-badge" hidden></span>
        <span id="librarySessionHint" class="library-session-hint" hidden></span>
        <button id="libraryHomeBtn" type="button">Màn chính</button>
        <button id="libraryLogoutBtn" type="button" hidden>Đăng xuất</button>
        <button id="libraryAdminOpenBtn" class="primary" type="button">Quản trị</button>
      </div>
    </div>
    <div class="library-controls">
      <input id="librarySearch" type="search" placeholder="Tìm theo tên, tác giả, năm, danh mục">
      <select id="libraryCategoryFilter"><option value="">Tất cả danh mục</option></select>
      <select id="libraryYearFilter"><option value="">Tất cả năm</option></select>
      <button id="libraryRefreshBtn" type="button">Làm mới</button>
    </div>
    <div id="libraryGrid" class="library-grid"></div>
    <div id="libraryEmpty" class="library-empty" hidden>Chưa có tài liệu phù hợp.</div>
  </section>
</main>
{WEBGIS_HTML}
<section id="libraryAccessPanel" class="library-access" hidden>
  <div class="library-access-card">
    <div class="library-admin-actions">
      <h2 style="flex:1">&#272;&#259;ng nh&#7853;p th&#432; vi&#7879;n</h2>
      <button id="libraryAccessCloseBtn" type="button">&#272;&#243;ng</button>
    </div>
    <p>Vui l&#242;ng &#273;&#259;ng nh&#7853;p tr&#432;&#7899;c khi v&#224;o th&#432; vi&#7879;n t&#224;i li&#7879;u.</p>
    <div class="library-access-form">
      <label>T&#224;i kho&#7843;n
        <input id="libraryAccessUser" type="text" autocomplete="username">
      </label>
      <label>M&#7853;t kh&#7849;u
        <input id="libraryAccessPassword" type="password" autocomplete="current-password">
      </label>
      <button id="libraryAccessLoginBtn" class="primary" type="button">&#272;&#259;ng nh&#7853;p</button>
      <div id="libraryAccessMsg" class="library-empty" hidden></div>
    </div>
    <p class="library-access-hint">N&#7871;u ch&#432;a c&#243; t&#224;i kho&#7843;n li&#234;n h&#7879; tr&#7921;c ti&#7871;p admin &#273;&#7875; &#273;&#432;&#7907;c cung c&#7845;p!</p>
  </div>
</section>
<section id="libraryAdminPanel" class="library-admin" hidden>
  <div class="library-admin-toolbar">
    <h2>Qu&#7843;n tr&#7883; th&#432; vi&#7879;n PDF</h2>
    <button id="libraryAdminCloseBtn" type="button">&#272;&#243;ng</button>
  </div>
  <div class="library-admin-inner">
    <div id="libraryUploadCard" class="library-admin-card" hidden>
      <h3>Upload t&#224;i li&#7879;u</h3>
      <form id="libraryDocForm" class="library-admin-form" hidden>
        <input id="libraryDocId" type="hidden">
        <label>Tên tài liệu
          <input id="libraryDocTitle" type="text" required>
        </label>
        <label>Tác giả / đơn vị biên soạn
          <input id="libraryDocAuthor" type="text">
        </label>
        <label>Năm xuất bản
          <input id="libraryDocYear" type="number" min="1800" max="2300">
        </label>
        <label>Danh mục tài liệu
          <input id="libraryDocCategory" type="text" list="libraryCategorySuggestions">
        </label>
        <datalist id="libraryCategorySuggestions"></datalist>
        <label>Mô tả ngắn
          <textarea id="libraryDocDescription"></textarea>
        </label>
        <label>File PDF
          <input id="libraryDocPdf" type="file" accept="application/pdf">
        </label>
        <label>Ảnh bìa nếu có
          <input id="libraryDocCover" type="file" accept="image/png,image/jpeg,image/webp,image/svg+xml">
        </label>
        <label class="import-options">
          <input id="libraryDocVisible" type="checkbox" checked>
          Hiển thị tài liệu
        </label>
        <div class="library-admin-actions">
          <button id="libraryDocSaveBtn" class="primary" type="submit">Lưu tài liệu</button>
          <button id="libraryDocNewBtn" type="button">Tạo mới</button>
        </div>
        <div id="libraryAdminMsg" class="library-empty" hidden></div>
      </form>
    </div>
    <div class="library-admin-card">
      <div class="library-admin-actions">
        <h3 style="flex:1">Danh sách tài liệu</h3>
        <button id="libraryAdminReloadBtn" type="button">Tải lại</button>
      </div>
      <div style="overflow:auto">
        <table class="library-admin-table">
          <thead>
            <tr>
              <th>Tên tài liệu</th>
              <th>Danh mục</th>
              <th>Năm</th>
              <th>Trạng thái</th>
              <th>Thao tác</th>
            </tr>
          </thead>
          <tbody id="libraryAdminRows"></tbody>
        </table>
      </div>
    </div>
  </div>
</section>
<section id="pdfReader" class="pdf-reader" hidden>
  <div class="reader-topbar">
    <div id="readerTitle" class="reader-title">Tài liệu PDF</div>
    <div class="reader-tools">
      <button id="readerPrevBtn" type="button">Trang trước</button>
      <input id="readerPageInput" class="reader-page-input" type="number" min="1" value="1">
      <span id="readerPageTotal">/ 1</span>
      <button id="readerNextBtn" type="button">Trang sau</button>
      <button id="readerZoomOutBtn" type="button">Thu nhỏ</button>
      <button id="readerZoomInBtn" type="button">Phóng to</button>
      <button id="readerFullscreenBtn" type="button">Toàn màn hình</button>
      <button id="readerCloseBtn" type="button">Đóng</button>
    </div>
  </div>
  <div class="reader-notice">Tài liệu chỉ được phép đọc trực tuyến, không được sao chép hoặc tải xuống.</div>
  <div id="pdfStage" class="pdf-stage">
    <div id="pdfCanvasWrap" class="pdf-canvas-wrap">
      <canvas id="pdfCanvas"></canvas>
    </div>
  </div>
</section>
<section id="importLog" class="import-log" hidden></section>
<section id="reportPanel" class="report-panel" hidden>
  <div class="report-card">
    <div class="report-head">
      <strong>Xuất thuyết minh cộng tăng/cộng giảm</strong>
      <button id="reportCloseBtn" type="button">Đóng</button>
    </div>
    <div class="report-controls">
      <input id="reportFilter" type="search" placeholder="Lọc mã hoặc tên đất">
      <input id="reportCurrentYear" type="number" min="1900" max="2200" value="2020" title="Năm hiện trạng">
      <input id="reportPlanYear" type="number" min="1900" max="2200" value="2030" title="Năm quy hoạch">
      <button id="reportSelectActiveBtn" type="button">Chọn mã có dữ liệu</button>
      <button id="reportClearBtn" type="button">Bỏ chọn</button>
      <button class="primary" id="reportExportBtn" type="button">Xuất Word</button>
    </div>
    <div id="reportOptions" class="report-options"></div>
  </div>
</section>
<section id="aiPanel" class="ai-panel" hidden>
  <div class="ai-card">
    <div class="ai-head">
      <strong id="aiPanelTitle">Trợ lý AI</strong>
      <button id="aiCloseBtn" type="button">Đóng</button>
    </div>
    <div id="aiMessages" class="ai-messages">
      <div id="aiIntroMessage" class="ai-message">Anh có thể hỏi: “Kiểm tra giúp tôi bảng này có lệch tổng không?”, “LUC tăng giảm thế nào?”, “Nhận xét lớp WebGIS đang bật”, hoặc “Viết nhận xét ngắn về biến động đất”.</div>
    </div>
    <div id="aiStatus" class="ai-status">Đang kiểm tra cấu hình AI...</div>
    <div class="ai-controls">
      <textarea id="aiQuestion" placeholder="Nhập câu hỏi cho AI"></textarea>
      <button id="aiSendBtn" class="primary" type="button">Gửi</button>
    </div>
  </div>
</section>
<section class="table-toolbar module-only">
  <div class="legend" aria-label="Chú giải màu">
    <span class="legend-item"><span class="swatch input"></span>Ô nhập liệu</span>
    <span class="legend-item"><span class="swatch diagonal"></span>Ô giữ nguyên loại đất / đường chéo</span>
    <span class="legend-item"><span class="swatch auto"></span>Ô công thức / tổng hợp</span>
    <span class="legend-item"><span class="swatch locked"></span>Ô khóa không nhập</span>
    <span class="legend-item"><span class="swatch plan-alert"></span>C&#7843;nh b&#225;o ch&#432;a &#273;&#7841;t / v&#432;&#7907;t ch&#7881; ti&#234;u</span>
  </div>
  <label class="view-options">
    <input id="hideZeroToggle" type="checkbox">
    Ẩn ô 0,00
  </label>
  <label class="view-options">
    <input id="compactColumnsToggle" type="checkbox">
    Ẩn cột không phát sinh
  </label>
</section>
<main class="table-wrap">
<table id="landTable">
<colgroup>{''.join(colgroup)}</colgroup>
<tbody>
{''.join(rows_html)}
</tbody>
</table>
</main>
<script id="meta" type="application/json">{meta_json}</script>
<script id="webgisSampleData" type="application/json">{webgis_sample_json}</script>
<script>{jszip_js}</script>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/html2canvas/1.4.1/html2canvas.min.js" referrerpolicy="no-referrer"></script>
<script>
const meta = JSON.parse(document.getElementById('meta').textContent);
const $ = (sel, root = document) => root.querySelector(sel);
const $$ = (sel, root = document) => Array.from(root.querySelectorAll(sel));
const storageKey = 'land-transfer-html-v1';
const hideZeroKey = 'land-transfer-hide-zero';
const compactColumnsKey = 'land-transfer-compact-zero-cols';
const projectId = 'default';
const apiBase = '/api/projects';
const libraryApiBase = '/api/library';
const inputCodes = meta.inputCodes;
const inputSet = new Set(inputCodes);
const rowsByCode = meta.codeRows;
const colsByCode = meta.codeCols;
const rowCodes = Object.fromEntries(Object.entries(rowsByCode).map(([code, row]) => [String(row), code]));
const colCodes = Object.fromEntries(Object.entries(colsByCode).map(([code, col]) => [String(col), code]));
const directChildren = meta.directChildren || {{}};
let matrixCodes = Object.keys(colsByCode);
let calcRowEntries = Object.entries(rowsByCode).filter(([, row]) => row >= meta.dttnRow && row < meta.totalIncreaseRow);
const inputCells = Array.from(document.querySelectorAll('td[data-input="1"]'));
const inputTds = new Map();
const inputEls = new Map();
const cellsByKey = new Map();
const autoSpans = new Map();
const inputKeys = new Set();
const previousWarnCells = new Set();
const previousPlanValues = {{}};
let projectTitlesConfirmed = false;
let gtpFileHandle = null;
let gtpFileName = '';
{WEBGIS_JS}
const StartupScreen = (() => {{
  const intervalMs = 3000;
  let timer = 0;
  let attempts = 0;
  const canCheckBackend = ['http:', 'https:'].includes(window.location.protocol);
  const screen = document.getElementById('startupScreen');
  const status = document.getElementById('startupStatus');

  function setStatus(text) {{
    if (status) status.textContent = text;
  }}

  function hide() {{
    if (!screen) return;
    clearInterval(timer);
    document.body.classList.remove('startup-active');
    screen.classList.add('is-hidden');
    window.setTimeout(() => {{
      screen.hidden = true;
    }}, 420);
  }}

  async function checkHealth() {{
    attempts += 1;
    try {{
      const response = await fetch('/health', {{
        cache: 'no-store',
        headers: {{ Accept: 'application/json' }}
      }});
      if (response.status === 200) {{
        setStatus('Máy chủ đã sẵn sàng');
        hide();
        return;
      }}
      setStatus(`Máy chủ đang khởi động (${{response.status}})`);
    }} catch (error) {{
      setStatus(attempts <= 1 ? 'Đang đánh thức máy chủ' : 'Máy chủ chưa sẵn sàng, thử lại sau 3 giây');
    }}
  }}

  function start() {{
    if (!screen) return;
    if (!canCheckBackend) {{
      hide();
      return;
    }}
    document.body.classList.add('startup-active');
    checkHealth();
    timer = window.setInterval(checkHealth, intervalMs);
  }}

  return {{ start, hide, checkHealth }};
}})();
StartupScreen.start();
let libraryDocuments = [];
const librarySessionTokenKey = 'library-session-token';
const librarySessionRoleKey = 'library-session-role';
let librarySessionToken = localStorage.getItem(librarySessionTokenKey) || localStorage.getItem('library-admin-token') || '';
let librarySessionRole = localStorage.getItem(librarySessionRoleKey) || (librarySessionToken ? 'admin' : '');
let libraryAdminToken = librarySessionRole === 'admin' ? librarySessionToken : '';
let activePdf = null;
let activePdfPage = 1;
let activePdfScale = 1.2;
let activePdfRenderTask = null;
let activePdfRenderSerial = 0;
let nextDynamicRow = meta.planRow + 1;
let nextDynamicCol = (meta.previousPlanStructureCol || meta.previousPlanCol || meta.planCol) + 1;

function isDiagonalMatrixCell(td) {{
  const row = Number(td.dataset.row || 0);
  const col = Number(td.dataset.col || 0);
  if (row < meta.dttnRow || col < meta.matrixStartCol || col > meta.matrixEndCol) return false;
  const rowCode = rowCodes[String(row)];
  const colCode = colCodes[String(col)];
  return Boolean(rowCode && colCode && rowCode === colCode);
}}

function registerCell(td) {{
  const key = `${{td.dataset.row}}:${{td.dataset.col}}`;
  cellsByKey.set(key, td);
  if (td.dataset.input === '1') inputKeys.add(key);
  td.classList.toggle('diagonal', isDiagonalMatrixCell(td));
  const span = td.querySelector('.value');
  if (span) autoSpans.set(key, span);
  const input = td.querySelector('input');
  if (input) {{
    inputTds.set(td.dataset.addr, td);
    inputEls.set(td.dataset.addr, input);
    input.addEventListener('input', scheduleRecalc);
    input.addEventListener('blur', () => {{
      normalizeInputElement(input);
      recalc();
    }});
  }}
}}

document.querySelectorAll('td[data-row][data-col]').forEach(registerCell);

inputCells.forEach(td => {{
  const input = td.querySelector('input');
  inputTds.set(td.dataset.addr, td);
  if (input) inputEls.set(td.dataset.addr, input);
}});

function addr(col, row) {{
  let n = col, s = '';
  while (n > 0) {{
    const m = (n - 1) % 26;
    s = String.fromCharCode(65 + m) + s;
    n = Math.floor((n - m) / 26);
  }}
  return s + row;
}}

function createCell(row, col, content, options = {{}}) {{
  const td = document.createElement('td');
  td.className = options.className || 'xl8';
  td.dataset.addr = addr(col, row);
  td.dataset.row = String(row);
  td.dataset.col = String(col);
  if (options.code) td.dataset.code = options.code;
  if (options.colCode) td.dataset.colCode = options.colCode;
  if (options.input) {{
    td.dataset.input = '1';
    td.innerHTML = `<input inputmode="decimal" value="${{content || ''}}" aria-label="${{td.dataset.addr}}">`;
  }} else if (options.auto) {{
    td.dataset.auto = '1';
    td.innerHTML = '<span class="value"></span>';
  }} else {{
    td.textContent = content || '';
  }}
  registerCell(td);
  return td;
}}

function refreshCalcEntries() {{
  matrixCodes = Object.keys(colsByCode);
  calcRowEntries = Object.entries(rowsByCode).filter(([, row]) => row >= meta.dttnRow);
}}

function addMatrixColumn(code) {{
  const col = nextDynamicCol++;
  colsByCode[code] = col;
  matrixCodes.push(code);
  const colgroup = document.querySelector('#landTable colgroup');
  const colEl = document.createElement('col');
  colEl.style.width = '64px';
  colEl.style.minWidth = '64px';
  colgroup.appendChild(colEl);

  document.querySelectorAll('#landTable tbody tr').forEach(tr => {{
    const row = Number(tr.querySelector('td[data-row]')?.dataset.row || 0);
    let cell;
    if (row === 3) {{
      cell = createCell(row, col, code, {{ className: 'xl3', colCode: code }});
    }} else if (row >= 4) {{
      const rowCode = rowCodes[String(row)];
      const isInputRow = rowCode && inputSet.has(rowCode);
      cell = createCell(row, col, '', {{ input: isInputRow, auto: !isInputRow, colCode: code }});
    }} else {{
      cell = createCell(row, col, '', {{ colCode: code }});
    }}
    tr.appendChild(cell);
  }});
  return col;
}}

function addMatrixRow(code) {{
  const row = nextDynamicRow++;
  rowsByCode[code] = row;
  rowCodes[String(row)] = code;
  inputSet.add(code);
  if (!inputCodes.includes(code)) inputCodes.push(code);
  leavesCache.set(code, [code]);

  const tr = document.createElement('tr');
  tr.style.height = '30px';
  tr.appendChild(createCell(row, 1, '', {{ className: 'xl8' }}));
  tr.appendChild(createCell(row, 2, `Mã mới ${{code}}`, {{ className: 'xl8' }}));
  tr.appendChild(createCell(row, 3, code, {{ className: 'xl8', code }}));
  tr.appendChild(createCell(row, meta.currentCol, '', {{ className: 'xl8', input: true, code }}));
  for (const colCode of matrixCodes) {{
    tr.appendChild(createCell(row, colsByCode[colCode], '', {{
      className: 'xl8',
      input: true,
      code,
      colCode
    }}));
  }}
  for (let col = meta.decreaseCol; col <= (meta.previousPlanStructureCol || meta.previousPlanCol || meta.planCol); col++) {{
    tr.appendChild(createCell(row, col, '', {{ className: 'xl8', auto: true, code }}));
  }}
  document.querySelector('#landTable tbody').appendChild(tr);
  calcRowEntries.push([code, row]);
  return row;
}}

function addMissingLandCode(code) {{
  const normalized = normalizeLandCode(code);
  if (!normalized) return false;
  if (!colsByCode[normalized]) addMatrixColumn(normalized);
  if (!rowsByCode[normalized]) addMatrixRow(normalized);
  refreshCalcEntries();
  return true;
}}

function readProjectSettings() {{
  return {{
    commune: ($('#projectCommune')?.value || '').trim(),
    province: ($('#projectProvince')?.value || '').trim(),
    previousPlanYear: ($('#projectPreviousPlanYear')?.value || '').trim(),
    currentYear: ($('#projectCurrentYear')?.value || '').trim(),
    planYear: ($('#projectPlanYear')?.value || '').trim(),
    confirmed: projectTitlesConfirmed
  }};
}}

function extractYears(text) {{
  return String(text || '').match(/(?:19|20|21|22)\\d{{2}}/g) || [];
}}

function yearFromPlanPeriod(period, fallback = 2030) {{
  const years = extractYears(period);
  return Number(years[years.length - 1]) || fallback;
}}

function syncProjectYearsToReport() {{
  const currentYear = ($('#projectCurrentYear')?.value || '').trim();
  const planYear = yearFromPlanPeriod($('#projectPlanYear')?.value, 2030);
  if ($('#reportCurrentYear') && currentYear) $('#reportCurrentYear').value = currentYear;
  if ($('#reportPlanYear') && planYear) $('#reportPlanYear').value = planYear;
}}

function syncReportYearsToProject() {{
  const currentYear = ($('#reportCurrentYear')?.value || '').trim();
  const planYear = ($('#reportPlanYear')?.value || '').trim();
  if ($('#projectCurrentYear') && currentYear) $('#projectCurrentYear').value = currentYear;
  if ($('#projectPlanYear') && planYear && !extractYears($('#projectPlanYear').value).length) $('#projectPlanYear').value = `${{currentYear || '2020'}}-${{planYear}}`;
}}

function planningPeriodLabel(period) {{
  const years = extractYears(period);
  if (years.length >= 2) return `năm ${{years[0]}} đến năm ${{years[years.length - 1]}}`;
  if (years.length === 1) return `đến năm ${{years[0]}}`;
  return 'theo kỳ quy hoạch đã thiết lập';
}}

function updateProjectTitles() {{
  const settings = readProjectSettings();
  const commune = (settings.commune || '...').replace(/^xã\\s+/i, '').trim() || '...';
  const period = planningPeriodLabel(settings.planYear);
  const titleCell = document.querySelector('[data-row="1"][data-col="1"]');
  const matrixTitleCell = document.querySelector('[data-addr="E2"]');
  if (titleCell) titleCell.textContent = `Chu chuyển đất đai trong kỳ quy hoạch sử dụng đất của xã ${{commune}}`;
  if (matrixTitleCell) matrixTitleCell.textContent = `Chu chuyển đất đai ${{period}}`;
}}

function resetProjectTitles() {{
  const titleCell = document.querySelector('[data-row="1"][data-col="1"]');
  const matrixTitleCell = document.querySelector('[data-addr="E2"]');
  if (titleCell) titleCell.textContent = 'BẢNG CHU CHUYỂN ĐẤT ĐAI';
  if (matrixTitleCell) matrixTitleCell.textContent = 'Chu chuyển các loại đất';
}}

function applyProjectSettings(settings = {{}}) {{
  const safe = settings && typeof settings === 'object' ? settings : {{}};
  projectTitlesConfirmed = Boolean(safe.confirmed);
  if ($('#projectCommune')) $('#projectCommune').value = safe.commune || '';
  if ($('#projectProvince')) $('#projectProvince').value = safe.province || '';
  if ($('#projectPreviousPlanYear')) $('#projectPreviousPlanYear').value = safe.previousPlanYear || '';
  if ($('#projectCurrentYear')) $('#projectCurrentYear').value = safe.currentYear || $('#projectCurrentYear').value || '2020';
  if ($('#projectPlanYear')) $('#projectPlanYear').value = safe.planYear || $('#projectPlanYear').value || '2020-2030';
  syncProjectYearsToReport();
  if (projectTitlesConfirmed) updateProjectTitles();
  else resetProjectTitles();
}}

function readInputs() {{
  normalizeAllInputs();
  collectPreviousPlanValues();
  const data = {{}};
  inputEls.forEach((input, address) => {{
    data[address] = input.value.trim();
  }});
  data.__previousPlan = {{ ...previousPlanValues }};
  data.__projectSettings = readProjectSettings();
  return data;
}}

function applyInputs(data) {{
  if (!data || typeof data !== 'object') return;
  if (data && data.__previousPlan && typeof data.__previousPlan === 'object') {{
    applyPreviousPlanValues(data.__previousPlan);
  }}
  applyProjectSettings(data.__projectSettings);
  inputEls.forEach((input, address) => {{
    if (Object.prototype.hasOwnProperty.call(data, address)) {{
      input.value = data[address];
      normalizeInputElement(input);
    }}
  }});
}}

function gtpPayload() {{
  return {{
    format: 'gtp-land-transfer',
    version: 1,
    savedAt: new Date().toISOString(),
    data: readInputs()
  }};
}}

function gtpDataFromPayload(payload) {{
  if (!payload || typeof payload !== 'object') throw new Error('File GTP không hợp lệ.');
  if (payload.format === 'gtp-land-transfer' && payload.data && typeof payload.data === 'object') return payload.data;
  return payload;
}}

function updateGtpStatus(message = '') {{
  const label = gtpFileName || 'chưa chọn file';
  $('#gtpStatus').textContent = message || `File GTP: ${{label}}`;
}}

function applyProjectData(data) {{
  applyInputs(data);
  localStorage.setItem(storageKey, JSON.stringify(readInputs()));
  normalizeAllInputs();
  recalc();
}}

async function saveGtpFile({{ choose = false, silent = false }} = {{}}) {{
  const text = JSON.stringify(gtpPayload(), null, 2);
  if (window.showSaveFilePicker) {{
    if (choose || !gtpFileHandle) {{
      gtpFileHandle = await window.showSaveFilePicker({{
        suggestedName: gtpFileName || 'du_an_chu_chuyen_dat_dai.gtp',
        types: [{{
          description: 'Dữ liệu dự án GTP',
          accept: {{ 'application/json': ['.gtp'] }}
        }}]
      }});
      gtpFileName = gtpFileHandle.name || gtpFileName || 'du_an_chu_chuyen_dat_dai.gtp';
    }}
    const writable = await gtpFileHandle.createWritable();
    await writable.write(text);
    await writable.close();
    updateGtpStatus(silent ? '' : `Đã lưu: ${{gtpFileName}}`);
    return true;
  }}
  download(gtpFileName || 'du_an_chu_chuyen_dat_dai.gtp', 'application/json;charset=utf-8', text);
  updateGtpStatus('Trình duyệt tải xuống file GTP mới');
  return true;
}}

async function openGtpProjectFile(file) {{
  const payload = JSON.parse(await file.text());
  const data = gtpDataFromPayload(payload);
  gtpFileHandle = null;
  gtpFileName = file.name || 'du_an_chu_chuyen_dat_dai.gtp';
  applyProjectData(data);
  updateGtpStatus(`Đã nạp: ${{gtpFileName}}`);
}}

function xmlText(node) {{
  return node ? node.textContent || '' : '';
}}

function columnIndexFromCellRef(ref) {{
  const letters = String(ref || '').replace(/[0-9]/g, '');
  let n = 0;
  for (const ch of letters) n = n * 26 + ch.charCodeAt(0) - 64;
  return n;
}}

function rowIndexFromCellRef(ref) {{
  const match = String(ref || '').match(/\\d+/);
  return match ? Number(match[0]) : 0;
}}

async function parseXlsxRows(file) {{
  const zip = await JSZip.loadAsync(await file.arrayBuffer());
  const parser = new DOMParser();
  const workbookXml = parser.parseFromString(await zip.file('xl/workbook.xml').async('text'), 'application/xml');
  const firstSheet = workbookXml.querySelector('sheet');
  const relId = firstSheet?.getAttribute('r:id');
  let sheetPath = 'xl/worksheets/sheet1.xml';
  const relsFile = zip.file('xl/_rels/workbook.xml.rels');
  if (relId && relsFile) {{
    const relsXml = parser.parseFromString(await relsFile.async('text'), 'application/xml');
    const rel = Array.from(relsXml.querySelectorAll('Relationship')).find(item => item.getAttribute('Id') === relId);
    const target = rel?.getAttribute('Target');
    if (target) sheetPath = target.startsWith('/') ? target.slice(1) : 'xl/' + target.replace(/^\\.\\.\\//, '');
  }}

  const sharedStrings = [];
  const sharedFile = zip.file('xl/sharedStrings.xml');
  if (sharedFile) {{
    const sharedXml = parser.parseFromString(await sharedFile.async('text'), 'application/xml');
    Array.from(sharedXml.querySelectorAll('si')).forEach(si => {{
      sharedStrings.push(Array.from(si.querySelectorAll('t')).map(t => t.textContent || '').join(''));
    }});
  }}

  const sheetXml = parser.parseFromString(await zip.file(sheetPath).async('text'), 'application/xml');
  const rows = [];
  Array.from(sheetXml.querySelectorAll('row')).forEach(rowNode => {{
    const row = {{ number: Number(rowNode.getAttribute('r') || 0), cells: {{}} }};
    Array.from(rowNode.querySelectorAll('c')).forEach(cell => {{
      const ref = cell.getAttribute('r') || '';
      const col = columnIndexFromCellRef(ref);
      const type = cell.getAttribute('t');
      let value = '';
      if (type === 's') value = sharedStrings[Number(xmlText(cell.querySelector('v')))] || '';
      else if (type === 'inlineStr') value = xmlText(cell.querySelector('is t'));
      else value = xmlText(cell.querySelector('v'));
      row.cells[col] = value;
    }});
    rows.push(row);
  }});
  return rows;
}}

function normalizeHeader(text) {{
  return String(text || '').trim().toLowerCase();
}}

function normalizeNumber(text) {{
  const value = parseNumericText(text);
  return Number.isFinite(value) ? String(roundNumber(value)) : '';
}}

function setPreviousPlanCell(code, value) {{
  const row = code === 'DTTN' ? meta.dttnRow : rowsByCode[code];
  if (!row || !meta.previousPlanCol) return false;
  const td = cellsByKey.get(`${{row}}:${{meta.previousPlanCol}}`);
  const span = td?.querySelector('.value');
  if (!span) return false;
  const numeric = parseNumericText(value);
  if (!Number.isFinite(numeric)) return false;
  previousPlanValues[code] = formatNumber(numeric);
  span.textContent = formatNumber(numeric);
  return true;
}}

function applyPreviousPlanValues(values) {{
  Object.keys(previousPlanValues).forEach(code => delete previousPlanValues[code]);
  document.querySelectorAll('td[data-previous-plan="1"] .value').forEach(span => {{
    span.textContent = '';
  }});
  Object.entries(values || {{}}).forEach(([code, value]) => {{
    setPreviousPlanCell(normalizeLandCode(code), value);
  }});
}}

function collectPreviousPlanValues() {{
  document.querySelectorAll('td[data-previous-plan="1"]').forEach(td => {{
    const row = Number(td.dataset.row || 0);
    const code = row === meta.dttnRow ? 'DTTN' : rowCodes[String(row)];
    const text = td.querySelector('.value')?.textContent || '';
    const value = parseNumericText(text);
    if (code && Number.isFinite(value)) previousPlanValues[code] = formatNumber(value);
  }});
}}

function previousPlanValueFor(code, row) {{
  const normalized = code || (row === meta.dttnRow ? 'DTTN' : rowCodes[String(row)]);
  const stored = previousPlanValues[normalized];
  let value = parseNumericText(stored);
  if (Number.isFinite(value)) return value;
  const td = cellsByKey.get(`${{row}}:${{meta.previousPlanCol}}`);
  value = parseNumericText(td?.textContent || '');
  return Number.isFinite(value) ? value : NaN;
}}

function updatePreviousPlanIndicator(row, structure) {{
  const planCell = cellsByKey.get(`${{row}}:${{meta.previousPlanCol}}`);
  if (!planCell) return;
  const rounded = roundNumber(structure);
  const shouldWarn = Number.isFinite(structure) && Math.abs(rounded - 100) > meta.tolerance;
  planCell.classList.toggle('plan-target-alert', shouldWarn);
  if (!shouldWarn) {{
    planCell.removeAttribute('title');
    return;
  }}
  planCell.title = rounded < 100
    ? 'ChÆ°a Ä‘áº¡t chá»‰ tiĂªu so vá»›i quy hoáº¡ch ká»³ trÆ°á»›c'
    : 'VÆ°á»£t chá»‰ tiĂªu so vá»›i quy hoáº¡ch ká»³ trÆ°á»›c';
}}

function setPreviousPlanChange(row, code, current) {{
  if (!meta.previousPlanChangeCol || !meta.previousPlanStructureCol) return;
  const previous = previousPlanValueFor(code, row);
  if (!Number.isFinite(previous) || Math.abs(previous) <= meta.tolerance) {{
    setAuto(row, meta.previousPlanChangeCol, NaN);
    setAuto(row, meta.previousPlanStructureCol, NaN);
    updatePreviousPlanIndicator(row, NaN);
    return;
  }}
  const structure = (current / previous) * 100;
  setAuto(row, meta.previousPlanChangeCol, current - previous);
  setAuto(row, meta.previousPlanStructureCol, structure);
  updatePreviousPlanIndicator(row, structure);
}}

function detectPreviousPlanColumns(rows) {{
  let codeCol = null;
  let areaCol = null;
  let headerRow = 0;
  const codeNames = new Set(['ma', 'ma_dat', 'code', 'land_code', 'ma_loai_dat']);
  const areaNames = new Set(['dien_tich', 'dien_tich_ha', 'area', 'area_ha', 'quy_hoach', 'quy_hoach_ky_truoc']);
  for (const row of rows.slice(0, 30)) {{
    for (const [colText, value] of Object.entries(row.cells)) {{
      const key = normalizeHeaderKey(value);
      const col = Number(colText);
      if (codeNames.has(key)) {{
        codeCol = col;
        headerRow = Math.max(headerRow, row.number);
      }}
      if ((areaNames.has(key) || key.includes('dien_tich')) && (!codeCol || col > codeCol)) {{
        areaCol = col;
        headerRow = Math.max(headerRow, row.number);
      }}
    }}
    if (codeCol && areaCol) break;
  }}
  if (!areaCol && codeCol) {{
    for (const row of rows.slice(0, 30)) {{
      for (const [colText, value] of Object.entries(row.cells)) {{
        const key = normalizeHeaderKey(value);
        const col = Number(colText);
        if (col > codeCol && key.includes('quy_hoach')) {{
          areaCol = col;
          headerRow = Math.max(headerRow, row.number);
          break;
        }}
      }}
      if (areaCol) break;
    }}
  }}
  if (!codeCol || !areaCol) {{
    throw new Error('Không nhận diện được cột Mã đất và cột Diện tích quy hoạch kỳ trước.');
  }}
  return {{ codeCol, areaCol, headerRow }};
}}

async function importPreviousPlanExcel(file) {{
  if (/\\.xls$/i.test(file.name) && !/\\.xlsx$/i.test(file.name)) {{
    throw new Error('File .xls đời cũ chưa được trình đọc tích hợp hỗ trợ. Vui lòng lưu lại thành .xlsx rồi import.');
  }}
  const rows = await parseXlsxRows(file);
  const columns = detectPreviousPlanColumns(rows);
  const imported = {{}};
  const unknownCodes = new Set();
  let readRows = 0;
  let validRows = 0;
  let skippedRows = 0;
  for (const row of rows) {{
    if (row.number <= columns.headerRow) continue;
    readRows++;
    let code = normalizeLandCode(row.cells[columns.codeCol]);
    const nameKey = normalizeHeaderKey(row.cells[columns.codeCol - 1]);
    if (!code && nameKey.includes('tong_dien_tich_tu_nhien')) code = 'DTTN';
    const value = parseNumericText(row.cells[columns.areaCol]);
    if (!code || !Number.isFinite(value)) {{
      skippedRows++;
      continue;
    }}
    if (code !== 'DTTN' && !rowsByCode[code]) {{
      unknownCodes.add(code);
      skippedRows++;
      continue;
    }}
    imported[code] = formatNumber(value);
    validRows++;
  }}
  applyPreviousPlanValues(imported);
  recalc();
  localStorage.setItem(storageKey, JSON.stringify(readInputs()));
  const el = $('#importLog');
  el.hidden = false;
  el.innerHTML = `
    <strong>Log import quy hoạch kỳ trước</strong>
    <ul>
      <li>Tổng số dòng đã đọc: ${{readRows}}</li>
      <li>Số dòng hợp lệ: ${{validRows}}</li>
      <li>Số dòng bị bỏ qua: ${{skippedRows}}</li>
      <li>Mã đất lạ: ${{unknownCodes.size ? Array.from(unknownCodes).sort().join(', ') : 'Không có'}}</li>
    </ul>`;
  return {{ readRows, validRows, skippedRows, unknownCodes: Array.from(unknownCodes) }};
}}

async function importCurrentAreasFromXlsx(file) {{
  const rows = await parseXlsxRows(file);
  let codeCol = null;
  let areaCol = null;
  for (const row of rows) {{
    for (const [colText, value] of Object.entries(row.cells)) {{
      const col = Number(colText);
      const header = normalizeHeaderKey(value);
      if (['ma', 'ma_dat', 'code', 'land_code'].includes(header)) codeCol = col;
      if (header.includes('dien_tich') || header.includes('area')) areaCol = col;
    }}
    if (codeCol && areaCol) break;
  }}
  if (!codeCol || !areaCol) throw new Error('Không tìm thấy cột Mã và cột Diện tích trong file Excel.');

  let imported = 0;
  let matchedNoValue = 0;
  const unmatched = [];
  const currentAreasByCode = new Map();
  for (const row of rows) {{
    const code = normalizeLandCode(row.cells[codeCol]);
    if (!code || !rowsByCode[code]) continue;
    const value = normalizeNumber(row.cells[areaCol]);
    if (value !== '') currentAreasByCode.set(code, Number(value));
    const input = inputEls.get(`D${{rowsByCode[code]}}`);
    if (!input) {{
      if (!directChildren[code]) unmatched.push(code);
      continue;
    }}
    if (value === '') {{
      matchedNoValue++;
      continue;
    }}
    setInputNumber(`D${{rowsByCode[code]}}`, Number(value));
    imported++;
  }}
  const adjustments = reconcileCurrentAreaRounding(currentAreasByCode);
  recalc();
  return {{ imported, matchedNoValue, unmatched: Array.from(new Set(unmatched)), adjustments }};
}}

function reconcileCurrentAreaRounding(currentAreasByCode) {{
  const adjustments = [];
  const parentCodes = Object.keys(directChildren)
    .filter(code => currentAreasByCode.has(code) && !inputEls.has(`D${{rowsByCode[code]}}`))
    .sort((a, b) => leaves(a).length - leaves(b).length);

  parentCodes.forEach(parentCode => {{
    const leafCodes = leaves(parentCode).filter(code => inputEls.has(`D${{rowsByCode[code]}}`));
    if (!leafCodes.length) return;
    const parentValue = roundNumber(currentAreasByCode.get(parentCode));
    const childSum = roundNumber(leafCodes.reduce((sum, code) => sum + numberFromInputByAddr(`D${{rowsByCode[code]}}`), 0));
    const diff = roundNumber(parentValue - childSum);
    if (!diff || Math.abs(diff) > 0.05) return;

    const targetCode = leafCodes.reduce((best, code) => {{
      const bestValue = numberFromInputByAddr(`D${{rowsByCode[best]}}`);
      const codeValue = numberFromInputByAddr(`D${{rowsByCode[code]}}`);
      return codeValue > bestValue ? code : best;
    }}, leafCodes[0]);
    const targetAddress = `D${{rowsByCode[targetCode]}}`;
    setInputNumber(targetAddress, numberFromInputByAddr(targetAddress) + diff);
    adjustments.push({{ parentCode, targetCode, diff }});
  }});
  return adjustments;
}}

function normalizeHeaderKey(text) {{
  return String(text || '')
    .normalize('NFD')
    .replace(/[̀-ͯ]/g, '')
    .toLowerCase()
    .trim()
    .replace(/[^a-z0-9]+/g, '_')
    .replace(/^_+|_+$/g, '');
}}

function normalizeLandCode(value) {{
  return String(value ?? '').trim().toUpperCase();
}}

function detectGISColumns(rows) {{
  const fromNames = new Set(['ma_hien_trang', 'ma_ht', 'hien_trang', 'from', 'from_code']);
  const toNames = new Set(['ma_quy_hoach', 'ma_qh', 'quy_hoach', 'to', 'to_code']);
  const areaNames = new Set(['dien_tich', 'area', 'area_ha', 'shape_area', 'shape_area_ha', 'area_m2', 'shape_area_m2', 'dt']);

  for (const row of rows.slice(0, 20)) {{
    const found = {{ headerRow: row.number, fromCode: null, toCode: null, area: null, areaHeader: '' }};
    for (const [colText, value] of Object.entries(row.cells)) {{
      const key = normalizeHeaderKey(value);
      const col = Number(colText);
      if (fromNames.has(key)) found.fromCode = col;
      if (toNames.has(key)) found.toCode = col;
      if (areaNames.has(key)) {{
        found.area = col;
        found.areaHeader = key;
      }}
    }}
    if (found.fromCode && found.toCode && found.area) return found;
  }}
  throw new Error('Không nhận diện được các cột Mã hiện trạng, Mã quy hoạch và Diện tích.');
}}

function areaUnitInfo(areaHeader) {{
  const key = normalizeHeaderKey(areaHeader);
  if (['shape_area', 'area_m2', 'shape_area_m2'].includes(key)) return {{ unit: 'm2', uncertain: false }};
  if (['dien_tich', 'area', 'area_ha', 'shape_area_ha', 'dt'].includes(key)) return {{ unit: 'ha', uncertain: false }};
  return {{ unit: 'unknown', uncertain: true }};
}}

function normalizeAreaValue(value) {{
  const raw = String(value ?? '').trim();
  if (!raw) return {{ value: 0, empty: true }};
  return {{ value: parseNumericText(raw), empty: false }};
}}

function aggregateOverlayRows(rows, columns, options = {{}}) {{
  const log = {{
    totalRows: 0,
    validRows: 0,
    skippedRows: 0,
    unknownCodes: new Set(),
    negativeRows: 0,
    totalArea: 0,
    warnings: [],
    addedCodes: new Set(),
    skippedUnknownRows: 0
  }};
  const matrix = {{}};
  const unit = areaUnitInfo(columns.areaHeader);
  const convertM2ToHa = unit.unit === 'm2' && options.convertM2ToHa;

  if (unit.unit === 'm2' && !convertM2ToHa) {{
    log.warnings.push('Cột diện tích có vẻ là m2; đang import nguyên giá trị. Hãy bật m2 -> ha nếu cần.');
  }}
  if (unit.uncertain) {{
    log.warnings.push('Không chắc đơn vị diện tích; vui lòng kiểm tra lại đơn vị sau khi import.');
  }}

  for (const row of rows) {{
    if (row.number <= columns.headerRow) continue;
    log.totalRows++;
    const fromCode = normalizeLandCode(row.cells[columns.fromCode]);
    const toCode = normalizeLandCode(row.cells[columns.toCode]);
    const parsedArea = normalizeAreaValue(row.cells[columns.area]);
    if (!fromCode || !toCode || !Number.isFinite(parsedArea.value)) {{
      log.skippedRows++;
      continue;
    }}
    if (parsedArea.value < 0) {{
      log.skippedRows++;
      log.negativeRows++;
      continue;
    }}
    const missing = [fromCode, toCode].filter(code => !rowsByCode[code] || !colsByCode[code]);
    if (missing.length) {{
      missing.forEach(code => log.unknownCodes.add(code));
      if (options.mode === 'add') {{
        missing.forEach(code => {{
          if (addMissingLandCode(code)) log.addedCodes.add(code);
        }});
      }} else {{
        log.skippedRows++;
        log.skippedUnknownRows++;
        continue;
      }}
    }}
    const area = convertM2ToHa ? parsedArea.value / 10000 : parsedArea.value;
    matrix[fromCode] ||= {{}};
    matrix[fromCode][toCode] = (matrix[fromCode][toCode] || 0) + area;
    log.validRows++;
    log.totalArea += area;
  }}

  log.unknownCodes = Array.from(log.unknownCodes).sort();
  log.addedCodes = Array.from(log.addedCodes).sort();
  return {{ matrix, log }};
}}

function clearGISMatrixInputs() {{
  inputCodes.forEach(code => {{
    matrixCodes.forEach(colCode => {{
      const input = inputEls.get(addr(colsByCode[colCode], rowsByCode[code]));
      if (input) input.value = '';
    }});
  }});
}}

function applyGISMatrix(matrix) {{
  clearGISMatrixInputs();
  let filledCells = 0;
  let skippedCells = 0;
  for (const [fromCode, rowValues] of Object.entries(matrix)) {{
    for (const [toCode, area] of Object.entries(rowValues)) {{
      const input = inputEls.get(addr(colsByCode[toCode], rowsByCode[fromCode]));
      if (input) {{
        setInputNumber(addr(colsByCode[toCode], rowsByCode[fromCode]), area);
        filledCells++;
      }} else {{
        skippedCells++;
      }}
    }}
  }}
  return {{ filledCells, skippedCells }};
}}

function calculateCurrentArea() {{}}
function calculateMatrixTotals() {{}}
function calculateDecrease() {{}}
function calculateIncrease() {{}}
function calculatePlanningArea() {{}}
function calculateChange() {{}}
function validateTable() {{}}
function renderTable() {{ recalc(); }}

function recalculateAfterImport() {{
  calculateCurrentArea();
  calculateMatrixTotals();
  calculateDecrease();
  calculateIncrease();
  calculatePlanningArea();
  calculateChange();
  validateTable();
  renderTable();
}}

function showImportLog(log) {{
  const el = $('#importLog');
  const warnings = [];
  if (log.negativeRows) warnings.push(`${{log.negativeRows}} dòng diện tích âm đã bị bỏ qua`);
  if (log.skippedUnknownRows) warnings.push(`${{log.skippedUnknownRows}} dòng có mã lạ đã bị bỏ qua`);
  warnings.push(...log.warnings);
  el.hidden = false;
  el.innerHTML = `
    <strong>Log import GIS</strong>
    <ul>
      <li>Tổng số dòng đã đọc: ${{log.totalRows}}</li>
      <li>Số dòng hợp lệ: ${{log.validRows}}</li>
      <li>Số dòng bị bỏ qua: ${{log.skippedRows}}</li>
      <li>Tổng diện tích đã import: ${{formatNumber(log.totalArea)}}</li>
      <li>Số ô vàng ma trận đã điền: ${{log.filledCells || 0}}</li>
      <li>Mã đất lạ: ${{log.unknownCodes.length ? log.unknownCodes.join(', ') : 'Không có'}}</li>
      <li>Mã đất đã tự thêm: ${{log.addedCodes.length ? log.addedCodes.join(', ') : 'Không có'}}</li>
      ${{warnings.length ? `<li>Cảnh báo: ${{warnings.join('; ')}}</li>` : ''}}
    </ul>`;
}}

async function importGISOverlayExcel(file) {{
  if (/\\.xls$/i.test(file.name) && !/\\.xlsx$/i.test(file.name)) {{
    throw new Error('File .xls đời cũ chưa được trình đọc tích hợp hỗ trợ. Vui lòng lưu lại thành .xlsx rồi import.');
  }}
  const rows = await parseXlsxRows(file);
  const columns = detectGISColumns(rows);
  const {{ matrix, log }} = aggregateOverlayRows(rows, columns, {{
    mode: $('#gisImportMode').value,
    convertM2ToHa: $('#gisM2ToHa').checked
  }});
  Object.assign(log, applyGISMatrix(matrix));
  recalculateAfterImport();
  showImportLog(log);
  localStorage.setItem(storageKey, JSON.stringify(readInputs()));
  return log;
}}



const displayDecimals = 2;
const displayFactor = 10 ** displayDecimals;

function roundNumber(value) {{
  if (!Number.isFinite(value)) return 0;
  const rounded = Math.round((value + Number.EPSILON) * displayFactor) / displayFactor;
  return Math.abs(rounded) < 0.0000001 ? 0 : rounded;
}}

function parseNumericText(text) {{
  const raw = String(text ?? '').trim();
  if (!raw) return NaN;
  let cleaned = raw.replace(/\\s/g, '');
  if (cleaned.includes(',') && cleaned.includes('.')) {{
    const lastComma = cleaned.lastIndexOf(',');
    const lastDot = cleaned.lastIndexOf('.');
    if (lastComma > lastDot) cleaned = cleaned.replace(/\\./g, '').replace(',', '.');
    else cleaned = cleaned.replace(/,/g, '');
  }} else if (cleaned.includes(',')) {{
    cleaned = cleaned.replace(',', '.');
  }}
  const value = Number(cleaned);
  return Number.isFinite(value) ? value : NaN;
}}

function formatInputValue(value) {{
  if (!Number.isFinite(value)) return '';
  const rounded = roundNumber(value);
  if (Math.abs(rounded) < 0.0000001) return '';
  return rounded.toFixed(displayDecimals).replace('.', ',');
}}

function normalizeInputElement(input) {{
  if (!input) return;
  const value = parseNumericText(input.value);
  input.value = Number.isFinite(value) ? formatInputValue(value) : '';
  updateInputZeroState(input);
}}

function normalizeAllInputs() {{
  inputEls.forEach(input => normalizeInputElement(input));
}}

function parseInputNumber(input) {{
  if (!input) return 0;
  const value = parseNumericText(input.value);
  return Number.isFinite(value) ? roundNumber(value) : 0;
}}

function setInputNumber(address, value) {{
  const input = inputEls.get(address);
  if (!input) return;
  input.value = formatInputValue(value);
  updateInputZeroState(input);
}}

function updateInputZeroState(input) {{
  if (!input) return;
  const td = input.closest('td');
  if (!td) return;
  const value = parseInputNumber(input);
  td.classList.toggle('zero-cell', Math.abs(value) <= meta.tolerance);
}}

function setDiagonalValue(code, value) {{
  const row = rowsByCode[code];
  const col = colsByCode[code];
  if (!row || !col) return;
  const address = addr(col, row);
  if (inputEls.has(address)) setInputNumber(address, value);
  else setAuto(row, col, value);
}}

function numberFromInputByAddr(address) {{
  return parseInputNumber(inputEls.get(address));
}}

function formatNumber(value) {{
  if (!Number.isFinite(value)) return '';
  const rounded = roundNumber(value);
  return rounded.toLocaleString('vi-VN', {{ maximumFractionDigits: displayDecimals, minimumFractionDigits: displayDecimals }});
}}

function setAuto(row, col, value) {{
  const key = `${{row}}:${{col}}`;
  const td = cellsByKey.get(key);
  if (!td || td.dataset.input === '1') return;
  const text = formatNumber(value);
  const span = autoSpans.get(key);
  if (span && span.textContent !== text) span.textContent = text;
  const raw = String(roundNumber(value));
  if (td.dataset.value !== raw) td.dataset.value = raw;
  td.classList.toggle('zero-cell', Math.abs(roundNumber(value)) <= meta.tolerance);
}}

function getAuto(row, col) {{
  const td = cellsByKey.get(`${{row}}:${{col}}`);
  if (!td) return 0;
  if (td.dataset.input === '1') return numberFromInputByAddr(td.dataset.addr);
  const value = Number(td.dataset.value);
  return Number.isFinite(value) ? roundNumber(value) : 0;
}}

const leavesCache = new Map();

function leaves(code) {{
  if (leavesCache.has(code)) return leavesCache.get(code);
  let result;
  if (code === 'DTTN') result = inputCodes.slice();
  else if (directChildren[code]) result = directChildren[code].flatMap(child => leaves(child));
  else result = inputSet.has(code) ? [code] : [];
  leavesCache.set(code, result);
  return result;
}}

leavesCache.set('DTTN', inputCodes.slice());

function diagonalCodes() {{
  return matrixCodes.filter(code => rowsByCode[code] && colsByCode[code]);
}}

function diagonalOutflowTotal(code, matrixValue) {{
  const ownLeafCodes = new Set(leaves(code));
  return inputCodes.reduce((sum, colCode) => {{
    if (ownLeafCodes.has(colCode)) return sum;
    return sum + matrixValue(code, colCode);
  }}, 0);
}}

function createCalcContext() {{
  const inputValues = new Map();
  inputEls.forEach((input, address) => inputValues.set(address, parseInputNumber(input)));
  const currentCache = new Map();
  const matrixCache = new Map();

  function currentArea(code) {{
    if (currentCache.has(code)) return currentCache.get(code);
    const value = inputSet.has(code)
      ? (inputValues.get('D' + rowsByCode[code]) || 0)
      : leaves(code).reduce((sum, leaf) => sum + currentArea(leaf), 0);
    currentCache.set(code, value);
    return value;
  }}

  function matrixLeaf(rowCode, colCode) {{
    if (!inputSet.has(rowCode) || !inputSet.has(colCode)) return 0;
    return inputValues.get(addr(colsByCode[colCode], rowsByCode[rowCode])) || 0;
  }}

  function matrixValue(rowCode, colCode) {{
    const key = rowCode + ':' + colCode;
    if (matrixCache.has(key)) return matrixCache.get(key);
    const rLeaves = leaves(rowCode);
    const cLeaves = leaves(colCode);
    let sum = 0;
    rLeaves.forEach(r => cLeaves.forEach(c => sum += matrixLeaf(r, c)));
    matrixCache.set(key, sum);
    return sum;
  }}

  return {{ currentArea, matrixLeaf, matrixValue }};
}}

function recalc() {{
  collectPreviousPlanValues();
  const {{ currentArea, matrixLeaf, matrixValue }} = createCalcContext();
  for (const code of Object.keys(rowsByCode)) {{
    setAuto(rowsByCode[code], meta.currentCol, currentArea(code));
  }}
  setAuto(meta.dttnRow, meta.currentCol, currentArea('DTTN'));

  for (const [code, row] of calcRowEntries) {{
    for (const colCode of matrixCodes) {{
      const col = colsByCode[colCode];
      if (inputKeys.has(`${{row}}:${{col}}`)) continue;
      setAuto(row, col, matrixValue(code, colCode));
    }}
  }}
  for (const colCode of matrixCodes) {{
    setAuto(meta.dttnRow, colsByCode[colCode], matrixValue('DTTN', colCode));
  }}

  for (const code of diagonalCodes()) {{
    const current = currentArea(code);
    const outflowTotal = diagonalOutflowTotal(code, matrixValue);
    if (current > meta.tolerance || outflowTotal > meta.tolerance) {{
      setDiagonalValue(code, Math.max(0, current - outflowTotal));
    }}
  }}

  const refreshedCalc = createCalcContext();
  for (const [code, row] of calcRowEntries) {{
    const current = row === meta.dttnRow ? refreshedCalc.currentArea('DTTN') : refreshedCalc.currentArea(code);
    const diagonal = refreshedCalc.matrixValue(code, code);
    const plan = refreshedCalc.matrixValue('DTTN', code);
    setAuto(row, meta.decreaseCol, current - diagonal);
    setAuto(row, meta.planCol, plan);
    setAuto(row, meta.changeCol, plan - current);
    setPreviousPlanChange(row, code, current);
  }}
  setAuto(meta.dttnRow, meta.decreaseCol,
    ['NNP', 'PNN', 'CSD'].reduce((sum, code) => sum + getAuto(rowsByCode[code] || 0, meta.decreaseCol), 0)
  );
  setAuto(meta.dttnRow, meta.planCol, ['NNP', 'PNN', 'CSD'].reduce((sum, code) => sum + refreshedCalc.matrixValue('DTTN', code), 0));
  setAuto(meta.dttnRow, meta.changeCol, getAuto(meta.dttnRow, meta.planCol) - getAuto(meta.dttnRow, meta.currentCol));
  setPreviousPlanChange(meta.dttnRow, 'DTTN', getAuto(meta.dttnRow, meta.currentCol));

  for (const colCode of matrixCodes) {{
    const col = colsByCode[colCode];
    const plan = refreshedCalc.matrixValue('DTTN', colCode);
    const diagonal = refreshedCalc.matrixValue(colCode, colCode);
    setAuto(meta.totalIncreaseRow, col, plan - diagonal);
    setAuto(meta.planRow, col, plan);
  }}
  setAuto(meta.totalIncreaseRow, meta.decreaseCol,
    ['NNP', 'PNN', 'CSD'].reduce((sum, code) => sum + getAuto(rowsByCode[code] || 0, meta.decreaseCol), 0)
  );

  updateWarnings({{ currentArea: refreshedCalc.currentArea, matrixLeaf: refreshedCalc.matrixLeaf }});
  updateCompactColumns();
}}

function updateWarnings(calc) {{
  const tol = meta.tolerance;
  let rowErrors = 0;
  const nextWarnCells = new Set();
  for (const code of inputCodes) {{
    const row = rowsByCode[code];
    const rowSum = inputCodes.reduce((sum, colCode) => sum + calc.matrixLeaf(code, colCode), 0);
    const current = calc.currentArea(code);
    if (Math.abs(rowSum - current) > tol) {{
      rowErrors++;
      for (let col = 1; col <= (meta.previousPlanStructureCol || meta.previousPlanCol || meta.planCol); col++) {{
        const td = cellsByKey.get(`${{row}}:${{col}}`);
        if (td) nextWarnCells.add(td);
      }}
    }}
  }}
  previousWarnCells.forEach(td => {{
    if (!nextWarnCells.has(td)) td.classList.remove('warn');
  }});
  nextWarnCells.forEach(td => {{
    if (!previousWarnCells.has(td)) td.classList.add('warn');
  }});
  previousWarnCells.clear();
  nextWarnCells.forEach(td => previousWarnCells.add(td));
  const totalDiff = Math.abs(getAuto(meta.dttnRow, meta.currentCol) - getAuto(meta.dttnRow, meta.planCol));
  const totalBadge = $('#statusTotal');
  totalBadge.textContent = totalDiff > tol ? `DTTN lệch ${{formatNumber(totalDiff)}}` : 'DTTN cân bằng';
  totalBadge.classList.toggle('warn', totalDiff > tol);
  const rowBadge = $('#statusRows');
  rowBadge.textContent = `${{rowErrors}} lệch hàng`;
  rowBadge.classList.toggle('warn', rowErrors > 0);
}}

function compactColumnIsActive(code, calc) {{
  const tol = meta.tolerance;
  if (Math.abs(calc.currentArea(code)) > tol) return true;
  if (Math.abs(calc.matrixValue('DTTN', code)) > tol) return true;
  if (Math.abs(calc.matrixValue(code, 'DTTN')) > tol) return true;
  for (const otherCode of matrixCodes) {{
    if (Math.abs(calc.matrixValue(code, otherCode)) > tol) return true;
    if (Math.abs(calc.matrixValue(otherCode, code)) > tol) return true;
  }}
  return false;
}}

function updateCompactColumns() {{
  const enabled = document.body.classList.contains('compact-zero-cols');
  const calc = enabled ? createCalcContext() : null;
  matrixCodes.forEach(code => {{
    const col = colsByCode[code];
    if (!col) return;
    const hide = enabled && !compactColumnIsActive(code, calc);
    const colEl = document.querySelector(`#landTable col:nth-child(${{col}})`);
    if (colEl) colEl.classList.toggle('compact-hidden', hide);
    document.querySelectorAll(`#landTable td[data-col="${{col}}"]`).forEach(td => td.classList.toggle('compact-hidden', hide));
  }});
}}

function applyCompactColumnsState(enabled) {{
  document.body.classList.toggle('compact-zero-cols', enabled);
  const checkbox = $('#compactColumnsToggle');
  if (checkbox) checkbox.checked = enabled;
  localStorage.setItem(compactColumnsKey, enabled ? '1' : '0');
  updateCompactColumns();
}}

function download(name, type, text) {{
  const blob = new Blob([text], {{ type }});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = name;
  a.click();
  URL.revokeObjectURL(a.href);
}}

function csvText(row, col) {{
  const td = cellsByKey.get(`${{row}}:${{col}}`);
  if (!td) return '';
  return td.dataset.input === '1' ? (inputEls.get(td.dataset.addr)?.value || '') : (td.textContent || '').trim();
}}

function exportCellText(row, col, renumberMap = new Map()) {{
  if (col === 1 && renumberMap.has(row)) return renumberMap.get(row);
  return csvText(row, col);
}}

function originalSttForCode(code) {{
  const row = rowsByCode[code];
  return row ? csvText(row, 1) : '';
}}

function csvEscape(text) {{
  return '"' + String(text).replaceAll('"', '""') + '"';
}}

function xmlEscape(text) {{
  return String(text ?? '')
    .replaceAll('&', '&amp;')
    .replaceAll('<', '&lt;')
    .replaceAll('>', '&gt;')
    .replaceAll('"', '&quot;');
}}

function exportActiveMatrixCodes(calc) {{
  const tol = meta.tolerance;
  return matrixCodes.filter(code => {{
    if (!rowsByCode[code] && !colsByCode[code]) return false;
    if (Math.abs(calc.currentArea(code)) > tol) return true;
    for (const otherCode of matrixCodes) {{
      if (Math.abs(calc.matrixValue(code, otherCode)) > tol) return true;
      if (Math.abs(calc.matrixValue(otherCode, code)) > tol) return true;
    }}
    return false;
  }});
}}

function exportCsv() {{
  normalizeAllInputs();
  recalc();
  const {{ exportCols, exportRows, renumberMap }} = exportMatrixShape();
  const rows = [];
  for (const row of exportRows) {{
    rows.push(exportCols.map(col => csvEscape(exportCellText(row, col, renumberMap))).join(','));
  }}
  download('chu_chuyen_dat_dai.csv', 'text/csv;charset=utf-8', '\\ufeff' + rows.join('\\n'));
}}

function exportMatrixShape() {{
  const calc = createCalcContext();
  const activeCodes = exportActiveMatrixCodes(calc);
  const activeSet = new Set(activeCodes);
  const exportCols = [
    1,
    2,
    3,
    meta.currentCol,
    ...activeCodes.map(code => colsByCode[code]).filter(Boolean),
    meta.decreaseCol,
    meta.changeCol,
    meta.planCol
  ].filter(Boolean);
  const exportDataRows = [];
  for (const [code, row] of calcRowEntries) {{
    if (row === meta.dttnRow || activeSet.has(code)) exportDataRows.push(row);
  }}
  exportDataRows.sort((a, b) => a - b);
  const exportRows = Array.from(new Set([1, 2, 3, meta.dttnRow, ...exportDataRows, meta.totalIncreaseRow, meta.planRow]))
    .sort((a, b) => a - b);
  const renumberMap = new Map();
  exportRows.forEach(row => {{
    if (row === meta.dttnRow || row === meta.totalIncreaseRow || row === meta.planRow) return;
    const code = rowCodes[String(row)];
    const stt = code ? originalSttForCode(code) : '';
    if (stt) renumberMap.set(row, stt);
  }});
  return {{ exportCols, exportRows, renumberMap }};
}}

function xlsxCellXml(cellRef, text, styleId = 0, forceText = false) {{
  const raw = String(text ?? '').trim();
  const numeric = parseNumericText(raw);
  if (!forceText && raw && Number.isFinite(numeric) && !/[A-Za-zÀ-ỹ]/.test(raw)) {{
    return `<c r="${{cellRef}}" s="${{styleId}}"><v>${{String(numeric)}}</v></c>`;
  }}
  return `<c r="${{cellRef}}" s="${{styleId}}" t="inlineStr"><is><t>${{xmlEscape(raw)}}</t></is></c>`;
}}

function xlsxStyleFor(row) {{
  const code = rowCodes[String(row)];
  if (row === 1) return 2;
  if (row <= 3 || row === meta.dttnRow || row === meta.totalIncreaseRow || row === meta.planRow) return 1;
  if (['NNP', 'PNN', 'CSD'].includes(code)) return 1;
  return 0;
}}

function xlsxCellStyleFor(row, col) {{
  if (col === 2 && row > 3) {{
    const baseStyle = xlsxStyleFor(row);
    return baseStyle === 1 ? 4 : 3;
  }}
  return xlsxStyleFor(row);
}}

function exportRowHeight(row) {{
  const name = csvText(row, 2);
  if (name.length > 48) return 36;
  if (name.length > 32) return 28;
  return row === 2 ? 31.2 : 18;
}}

function exportXlsx() {{
  normalizeAllInputs();
  recalc();
  const {{ exportCols, exportRows, renumberMap }} = exportMatrixShape();
  const sheetRows = exportRows.map((row, rowIndex) => {{
    const cells = exportCols.map((col, colIndex) => xlsxCellXml(addr(colIndex + 1, rowIndex + 1), exportCellText(row, col, renumberMap), xlsxCellStyleFor(row, col), col === 1)).join('');
    const height = exportRowHeight(row);
    return `<row r="${{rowIndex + 1}}" ht="${{height}}" customHeight="1">${{cells}}</row>`;
  }}).join('');
  const widths = exportCols.map((col, index) => {{
    const width = col === 2 ? 34 : (col === 1 || col === 3 ? 10 : 15);
    return `<col min="${{index + 1}}" max="${{index + 1}}" width="${{width}}" customWidth="1"/>`;
  }}).join('');
  const lastRef = addr(exportCols.length, 1);
  const matrixEnd = 4 + exportCols.filter(col => col >= meta.matrixStartCol && col <= meta.matrixEndCol).length;
  const mergeRefs = [`<mergeCell ref="A1:${{lastRef}}"/>`];
  if (matrixEnd > 5) mergeRefs.push(`<mergeCell ref="E2:${{addr(matrixEnd, 2)}}"/>`);
  const previousChangeIndex = exportCols.indexOf(meta.previousPlanChangeCol) + 1;
  const previousStructureIndex = exportCols.indexOf(meta.previousPlanStructureCol) + 1;
  if (previousChangeIndex > 0 && previousStructureIndex > previousChangeIndex) {{
    mergeRefs.push(`<mergeCell ref="${{addr(previousChangeIndex, 2)}}:${{addr(previousStructureIndex, 2)}}"/>`);
  }}
  const mergeXml = `<mergeCells count="${{mergeRefs.length}}">${{mergeRefs.join('')}}</mergeCells>`;
  const zip = new JSZip();
  zip.file('[Content_Types].xml', `<?xml version="1.0" encoding="UTF-8"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>`);
  zip.file('_rels/.rels', `<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>`);
  zip.file('xl/_rels/workbook.xml.rels', `<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>`);
  zip.file('xl/workbook.xml', `<?xml version="1.0" encoding="UTF-8"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Chu chuyển đất đai" sheetId="1" r:id="rId1"/></sheets>
</workbook>`);
  zip.file('xl/styles.xml', `<?xml version="1.0" encoding="UTF-8"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3">
    <font><sz val="12"/><name val="Times New Roman"/></font>
    <font><b/><sz val="12"/><name val="Times New Roman"/></font>
    <font><b/><sz val="12"/><name val="Times New Roman"/></font>
  </fonts>
  <fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>
  <borders count="1"><border><left style="thin"><color auto="1"/></left><right style="thin"><color auto="1"/></right><top style="thin"><color auto="1"/></top><bottom style="thin"><color auto="1"/></bottom><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="5">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="left" vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyFont="1" applyBorder="1" applyAlignment="1"><alignment horizontal="left" vertical="center" wrapText="1"/></xf>
  </cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>`);
  zip.file('xl/worksheets/sheet1.xml', `<?xml version="1.0" encoding="UTF-8"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetViews><sheetView workbookViewId="0"><pane xSplit="4" ySplit="3" topLeftCell="E4" activePane="bottomRight" state="frozen"/></sheetView></sheetViews>
  <cols>${{widths}}</cols>
  <sheetData>${{sheetRows}}</sheetData>
  ${{mergeXml}}
</worksheet>`);
  zip.generateAsync({{ type: 'blob', mimeType: 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' }})
    .then(blob => {{
      const a = document.createElement('a');
      a.href = URL.createObjectURL(blob);
      a.download = 'chu_chuyen_dat_dai.xlsx';
      a.click();
      URL.revokeObjectURL(a.href);
    }});
}}

let searchHitTimer = 0;
function jumpToLandCode(rawCode) {{
  const code = normalizeLandCode(rawCode);
  if (!code) return;
  const row = rowsByCode[code];
  const col = colsByCode[code];
  if (!row || !col) {{
    alert(`Không tìm thấy mã đất: ${{code}}`);
    return;
  }}
  const td = cellsByKey.get(`${{row}}:${{col}}`);
  if (!td) return;
  td.scrollIntoView({{ block: 'center', inline: 'center', behavior: 'smooth' }});
  clearTimeout(searchHitTimer);
  document.querySelectorAll('.search-hit').forEach(el => el.classList.remove('search-hit'));
  td.classList.add('search-hit');
  searchHitTimer = setTimeout(() => td.classList.remove('search-hit'), 2600);
}}

function landName(code) {{
  const row = rowsByCode[code];
  const name = cellsByKey.get(`${{row}}:2`)?.textContent?.trim();
  return name || code;
}}

function ownLeafSet(code) {{
  return new Set(leaves(code));
}}

function reportIncomingEntries(code, calc) {{
  const ownLeaves = ownLeafSet(code);
  return inputCodes
    .filter(sourceCode => !ownLeaves.has(sourceCode))
    .map(sourceCode => [sourceCode, calc.matrixValue(sourceCode, code)])
    .filter(([, value]) => Math.abs(value) > meta.tolerance)
    .sort((a, b) => Math.abs(b[1]) - Math.abs(a[1]));
}}

function reportOutgoingEntries(code, calc) {{
  const ownLeaves = ownLeafSet(code);
  return inputCodes
    .filter(targetCode => !ownLeaves.has(targetCode))
    .map(targetCode => [targetCode, calc.matrixValue(code, targetCode)])
    .filter(([, value]) => Math.abs(value) > meta.tolerance)
    .sort((a, b) => Math.abs(b[1]) - Math.abs(a[1]));
}}

function hasReportData(code, calc) {{
  return Math.abs(calc.currentArea(code)) > meta.tolerance ||
    Math.abs(calc.matrixValue('DTTN', code)) > meta.tolerance ||
    reportIncomingEntries(code, calc).length > 0 ||
    reportOutgoingEntries(code, calc).length > 0;
}}

function reportCodeOptions() {{
  return matrixCodes
    .filter(code => rowsByCode[code] && colsByCode[code])
    .map(code => ({{ code, name: landName(code) }}));
}}

function renderReportOptions(filter = '') {{
  const q = normalizeHeaderKey(filter);
  const selected = new Set(Array.from(document.querySelectorAll('#reportOptions input:checked')).map(input => input.value));
  const options = reportCodeOptions().filter(item => {{
    if (!q) return true;
    return normalizeHeaderKey(item.code).includes(q) || normalizeHeaderKey(item.name).includes(q);
  }});
  $('#reportOptions').innerHTML = options.map(item => `
    <label class="report-option">
      <input type="checkbox" value="${{item.code}}" ${{selected.has(item.code) ? 'checked' : ''}}>
      <span><strong>${{item.code}}</strong><br>${{item.name}}</span>
    </label>
  `).join('');
}}

function selectedReportCodes() {{
  return Array.from(document.querySelectorAll('#reportOptions input:checked')).map(input => input.value);
}}

function reportLine(prefix, text, value, end = ';') {{
  return `<div class="line">${{prefix}} ${{text}}<span class="amount">: ${{formatNumber(value)}} ha${{end}}</span></div>`;
}}

function reportBlock(code, calc, years) {{
  const name = landName(code);
  const current = calc.currentArea(code);
  const plan = calc.matrixValue('DTTN', code);
  const natural = calc.currentArea('DTTN');
  const share = natural > meta.tolerance ? (plan / natural) * 100 : 0;
  const change = plan - current;
  const direction = change >= -meta.tolerance ? 'tăng' : 'giảm';
  const incoming = reportIncomingEntries(code, calc);
  const outgoing = reportOutgoingEntries(code, calc);
  const incomingTotal = incoming.reduce((sum, [, value]) => sum + value, 0);
  const outgoingTotal = outgoing.reduce((sum, [, value]) => sum + value, 0);
  if (Math.abs(change) <= meta.tolerance) {{
    return `
    <div class="block">
      <div class="title-line">* <strong><em>${{name}}:</em></strong></div>
      <div>Quy hoạch sử dụng đất đến năm ${{years.planYear}} là ${{formatNumber(plan)}} ha, chiếm ${{formatNumber(share)}}% tổng diện tích tự nhiên, không biến động so với năm ${{years.currentYear}}.</div>
    </div>`;
  }}
  const incomingLines = incoming.length
    ? incoming.map(([sourceCode, value], index) => reportLine('-', landName(sourceCode), value, index === incoming.length - 1 ? '.' : ';')).join('')
    : '';
  const outgoingLines = outgoing.length
    ? outgoing.map(([targetCode, value], index) => reportLine('-', landName(targetCode), value, index === outgoing.length - 1 ? '.' : ';')).join('')
    : '';
  const incomingSection = Math.abs(incomingTotal) > meta.tolerance
    ? `<div class="section">+ Cộng tăng ${{formatNumber(incomingTotal)}} ha do chuyển sang từ các loại đất sau:</div>${{incomingLines}}`
    : '';
  const outgoingSection = Math.abs(outgoingTotal) > meta.tolerance
    ? `<div class="section">+ Cộng giảm ${{formatNumber(outgoingTotal)}} ha, do chuyển sang các loại đất sau:</div>${{outgoingLines}}`
    : '';
  return `
    <div class="block">
      <div class="title-line">* <strong><em>${{name}}:</em></strong></div>
      <div>Diện tích năm ${{years.currentYear}} là ${{formatNumber(current)}} ha, quy hoạch sử dụng đất đến năm ${{years.planYear}} là ${{formatNumber(plan)}} ha, chiếm ${{formatNumber(share)}}% tổng diện tích tự nhiên, ${{direction}} ${{formatNumber(Math.abs(change))}} ha so với năm ${{years.currentYear}}, chi tiết như sau:</div>
      ${{incomingSection}}
      ${{outgoingSection}}
    </div>`;
}}

function reportYears() {{
  const currentYear = Number(($('#projectCurrentYear')?.value || $('#reportCurrentYear').value || '').trim()) || 2020;
  const planYear = yearFromPlanPeriod($('#projectPlanYear')?.value || $('#reportPlanYear').value, 2030);
  $('#reportCurrentYear').value = currentYear;
  $('#reportPlanYear').value = planYear;
  return {{ currentYear, planYear }};
}}

function exportReportWord() {{
  normalizeAllInputs();
  recalc();
  const codes = selectedReportCodes();
  if (!codes.length) {{
    alert('Hãy chọn ít nhất một loại đất để xuất.');
    return;
  }}
  const calc = createCalcContext();
  const years = reportYears();
  const body = codes.map(code => reportBlock(code, calc, years)).join('');
  const html = `<!doctype html>
<html>
<head>
<meta charset="utf-8">
<style>
@page {{ margin: 2cm; }}
body {{ font-family: "Times New Roman", serif; font-size: 13pt; line-height: 1.35; }}
.block {{ margin: 0 0 14pt 0; }}
.title-line {{ font-weight: 700; }}
.section {{ margin-top: 6pt; }}
.line {{ white-space: nowrap; }}
.amount {{ display: inline-block; min-width: 150px; text-align: left; }}
</style>
</head>
<body>${{body}}</body>
</html>`;
  download('thuyet_minh_cong_tang_cong_giam.doc', 'application/msword;charset=utf-8', '\\ufeff' + html);
}}

let pendingRecalc = 0;
function scheduleRecalc() {{
  if (pendingRecalc) return;
  pendingRecalc = requestAnimationFrame(() => {{
    pendingRecalc = 0;
    recalc();
  }});
}}

async function saveProjectToServer() {{
  const response = await fetch(`${{apiBase}}/${{encodeURIComponent(projectId)}}`, {{
    method: 'PUT',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify({{ data: readInputs() }})
  }});
  if (!response.ok) {{
    const payload = await response.json().catch(() => ({{ error: response.statusText }}));
    throw new Error(payload.error || 'Không lưu được dữ liệu lên server');
  }}
  return response.json();
}}

async function loadProjectFromServer() {{
  try {{
    const response = await fetch(`${{apiBase}}/${{encodeURIComponent(projectId)}}`);
    if (response.status === 404) return false;
    if (!response.ok) throw new Error('Không đọc được dữ liệu từ server');
    const payload = await response.json();
    if (payload.data && typeof payload.data === 'object') {{
      applyInputs(payload.data);
      localStorage.setItem(storageKey, JSON.stringify(payload.data));
      normalizeAllInputs();
      recalc();
      return true;
    }}
  }} catch (error) {{
    console.warn(error.message || error);
  }}
  return false;
}}

function showLibraryMessage(target, message, isError = false) {{
  const box = $(target);
  if (!box) return;
  box.hidden = !message;
  box.textContent = message || '';
  box.style.borderColor = isError ? '#f4b0a1' : '#bbf7d0';
  box.style.background = isError ? '#fff1ed' : '#f0fdf4';
  box.style.color = isError ? '#7a271a' : '#166534';
}}

function libraryAuthHeaders() {{
  return librarySessionToken ? {{ Authorization: `Bearer ${{librarySessionToken}}` }} : {{}};
}}

function libraryAdminHeaders() {{
  return librarySessionRole === 'admin' && librarySessionToken ? {{ Authorization: `Bearer ${{librarySessionToken}}` }} : {{}};
}}

function updateLibrarySessionUi() {{
  const logged = Boolean(librarySessionToken);
  const isAdmin = logged && librarySessionRole === 'admin';
  const badge = $('#librarySessionBadge');
  if (badge) {{
    badge.hidden = !logged;
    badge.textContent = isAdmin ? 'Vai trò: Admin' : 'Vai trò: Khách';
    badge.classList.toggle('admin', isAdmin);
    badge.classList.toggle('guest', logged && !isAdmin);
  }}
  const hint = $('#librarySessionHint');
  if (hint) {{
    hint.hidden = !logged;
    hint.textContent = isAdmin ? 'Được upload, sửa, ẩn/hiện và xóa tài liệu' : 'Chỉ được đọc tài liệu trực tuyến';
  }}
  const logoutBtn = $('#libraryLogoutBtn');
  if (logoutBtn) logoutBtn.hidden = !logged;
  const adminBtn = $('#libraryAdminOpenBtn');
  if (adminBtn) adminBtn.hidden = !isAdmin;
}}

function setLibrarySession(payload) {{
  librarySessionToken = payload.token || '';
  librarySessionRole = payload.role || 'guest';
  libraryAdminToken = librarySessionRole === 'admin' ? librarySessionToken : '';
  localStorage.setItem(librarySessionTokenKey, librarySessionToken);
  localStorage.setItem(librarySessionRoleKey, librarySessionRole);
  if (libraryAdminToken) localStorage.setItem('library-admin-token', libraryAdminToken);
  else localStorage.removeItem('library-admin-token');
  updateLibrarySessionUi();
}}

function clearLibrarySession() {{
  librarySessionToken = '';
  librarySessionRole = '';
  libraryAdminToken = '';
  localStorage.removeItem(librarySessionTokenKey);
  localStorage.removeItem(librarySessionRoleKey);
  localStorage.removeItem('library-admin-token');
  $('#libraryAdminPanel').hidden = true;
  updateLibrarySessionUi();
}}

function showLibraryAccessPanel(message = '') {{
  closeMainMenu();
  $('#libraryAccessPanel').hidden = false;
  showLibraryMessage('#libraryAccessMsg', message);
  setTimeout(() => $('#libraryAccessUser')?.focus(), 0);
}}

function hideLibraryAccessPanel() {{
  $('#libraryAccessPanel').hidden = true;
  showLibraryMessage('#libraryAccessMsg', '');
}}

function escapeHtml(value) {{
  return String(value ?? '').replace(/[&<>"']/g, ch => ({{
    '&': '&amp;',
    '<': '&lt;',
    '>': '&gt;',
    '"': '&quot;',
    "'": '&#39;'
  }}[ch]));
}}

function fileToDataUrl(file) {{
  return new Promise((resolve, reject) => {{
    if (!file) {{
      resolve('');
      return;
    }}
    const reader = new FileReader();
    reader.onload = () => resolve(String(reader.result || ''));
    reader.onerror = () => reject(reader.error || new Error('Không đọc được file.'));
    reader.readAsDataURL(file);
  }});
}}

async function loadPdfJs() {{
  if (window.pdfjsLib) return window.pdfjsLib;
  await new Promise((resolve, reject) => {{
    const script = document.createElement('script');
    script.src = 'https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.min.js';
    script.onload = resolve;
    script.onerror = () => reject(new Error('Không tải được PDF.js. Hãy kiểm tra kết nối mạng.'));
    document.head.appendChild(script);
  }});
  window.pdfjsLib.GlobalWorkerOptions.workerSrc = 'https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.worker.min.js';
  return window.pdfjsLib;
}}

function libraryQueryString(includeHidden = false) {{
  const params = new URLSearchParams();
  const q = $('#librarySearch')?.value.trim();
  const category = $('#libraryCategoryFilter')?.value;
  const year = $('#libraryYearFilter')?.value;
  if (q) params.set('q', q);
  if (category) params.set('category', category);
  if (year) params.set('year', year);
  if (includeHidden) params.set('includeHidden', '1');
  const text = params.toString();
  return text ? `?${{text}}` : '';
}}

async function fetchLibraryDocuments(includeHidden = false) {{
  if (!librarySessionToken) {{
    showLibraryAccessPanel();
    throw new Error('Bạn cần đăng nhập để vào thư viện tài liệu.');
  }}
  const response = await fetch(`${{libraryApiBase}}/documents${{libraryQueryString(includeHidden)}}`, {{
    headers: includeHidden ? libraryAdminHeaders() : libraryAuthHeaders()
  }});
  const payload = await response.json().catch(() => ({{}}));
  if (response.status === 401) {{
    clearLibrarySession();
    showLibraryAccessPanel(payload.error || 'Phiên đăng nhập thư viện đã hết hạn. Vui lòng đăng nhập lại.');
  }}
  if (!response.ok) throw new Error(payload.error || 'Không tải được thư viện tài liệu.');
  libraryDocuments = payload.documents || [];
  renderLibraryFilters(payload);
  renderLibraryGrid(libraryDocuments);
  if (includeHidden) renderLibraryAdminRows(libraryDocuments);
  return payload;
}}

function renderLibraryFilters(payload = {{}}) {{
  const categorySelect = $('#libraryCategoryFilter');
  const yearSelect = $('#libraryYearFilter');
  const currentCategory = categorySelect.value;
  const currentYear = yearSelect.value;
  categorySelect.innerHTML = '<option value="">Tất cả danh mục</option>' +
    (payload.categories || []).map(item => `<option value="${{escapeHtml(item.category)}}">${{escapeHtml(item.category)}} (${{item.count}})</option>`).join('');
  yearSelect.innerHTML = '<option value="">Tất cả năm</option>' +
    (payload.years || []).map(year => `<option value="${{year}}">${{year}}</option>`).join('');
  categorySelect.value = currentCategory;
  yearSelect.value = currentYear;
  const datalist = $('#libraryCategorySuggestions');
  if (datalist) {{
    datalist.innerHTML = (payload.categories || []).map(item => `<option value="${{escapeHtml(item.category)}}"></option>`).join('');
  }}
}}

function renderLibraryGrid(documents) {{
  const grid = $('#libraryGrid');
  const empty = $('#libraryEmpty');
  grid.innerHTML = documents.map(doc => `
    <article class="library-card">
      <div class="library-cover">
        ${{doc.coverUrl ? `<img src="${{doc.coverUrl}}" alt="Bìa tài liệu ${{escapeHtml(doc.title)}}">` : `<div class="library-cover-placeholder">${{escapeHtml(doc.title)}}</div>`}}
      </div>
      <div class="library-card-body">
        <h3>${{escapeHtml(doc.title)}}</h3>
        <span class="library-card-status">Chỉ đọc trực tuyến</span>
        <div class="library-meta">
          <span class="library-pill">${{escapeHtml(doc.category || 'Chưa phân loại')}}</span>
          <span class="library-pill">${{doc.year || 'Không rõ năm'}}</span>
        </div>
        <div class="library-meta library-author">${{escapeHtml(doc.author || 'Chưa rõ tác giả')}}</div>
        <div class="library-description">${{escapeHtml(doc.description || '')}}</div>
        <button class="primary library-read-btn" type="button" data-id="${{doc.id}}">Đọc trực tuyến</button>
      </div>
    </article>
  `).join('');
  empty.hidden = documents.length > 0;
}}

function renderLibraryAdminRows(documents) {{
  const tbody = $('#libraryAdminRows');
  tbody.innerHTML = documents.map(doc => `
    <tr>
      <td><strong>${{escapeHtml(doc.title)}}</strong><br><span class="library-meta">${{escapeHtml(doc.author || '')}}</span></td>
      <td>${{escapeHtml(doc.category || '')}}</td>
      <td>${{doc.year || ''}}</td>
      <td>${{doc.visible ? 'Hiển thị' : 'Đang ẩn'}}</td>
      <td>
        <button type="button" data-action="edit" data-id="${{doc.id}}">Sửa</button>
        <button type="button" data-action="toggle" data-id="${{doc.id}}">${{doc.visible ? 'Ẩn' : 'Hiện'}}</button>
        <button type="button" data-action="delete" data-id="${{doc.id}}">Xóa</button>
      </td>
    </tr>
  `).join('');
}}

function resetLibraryDocForm() {{
  $('#libraryDocForm').reset();
  $('#libraryDocId').value = '';
  $('#libraryDocVisible').checked = true;
  showLibraryMessage('#libraryAdminMsg', '');
}}

function fillLibraryDocForm(doc) {{
  $('#libraryDocId').value = doc.id;
  $('#libraryDocTitle').value = doc.title || '';
  $('#libraryDocAuthor').value = doc.author || '';
  $('#libraryDocYear').value = doc.year || '';
  $('#libraryDocCategory').value = doc.category || '';
  $('#libraryDocDescription').value = doc.description || '';
  $('#libraryDocVisible').checked = Boolean(doc.visible);
  $('#libraryDocPdf').value = '';
  $('#libraryDocCover').value = '';
  showLibraryMessage('#libraryAdminMsg', `Đang sửa: ${{doc.title}}`);
}}

async function openLibraryAdminPanel() {{
  if (librarySessionRole !== 'admin') {{
    if (!librarySessionToken) showLibraryAccessPanel('Vui lòng đăng nhập bằng tài khoản admin để quản trị thư viện.');
    else alert('Tài khoản khách chỉ được đọc tài liệu. Vui lòng đăng nhập admin để upload hoặc chỉnh sửa.');
    return;
  }}
  $('#libraryAdminPanel').hidden = false;
  showLibraryMessage('#libraryAdminMsg', '');
  $('#libraryUploadCard').hidden = false;
  $('#libraryDocForm').hidden = false;
  try {{
    await fetchLibraryDocuments(true);
  }} catch (error) {{
    $('#libraryAdminPanel').hidden = true;
    showLibraryAccessPanel(error.message || String(error));
  }}
}}

async function libraryAccessLogin() {{
  const username = $('#libraryAccessUser').value.trim();
  const password = $('#libraryAccessPassword').value;
  const response = await fetch(`${{libraryApiBase}}/login`, {{
    method: 'POST',
    headers: {{ 'Content-Type': 'application/json' }},
    body: JSON.stringify({{ username, password }})
  }});
  const payload = await response.json().catch(() => ({{}}));
  if (!response.ok) throw new Error(payload.error || 'Không đăng nhập được thư viện.');
  setLibrarySession(payload);
  hideLibraryAccessPanel();
  showDocumentLibraryPage();
}}

async function saveLibraryDocument(event) {{
  event.preventDefault();
  const id = $('#libraryDocId').value;
  const pdfFile = $('#libraryDocPdf').files[0];
  if (!id && !pdfFile) {{
    showLibraryMessage('#libraryAdminMsg', 'Tài liệu mới cần có file PDF.', true);
    return;
  }}
  const coverFile = $('#libraryDocCover').files[0];
  const payload = {{
    title: $('#libraryDocTitle').value,
    author: $('#libraryDocAuthor').value,
    year: $('#libraryDocYear').value,
    category: $('#libraryDocCategory').value,
    description: $('#libraryDocDescription').value,
    visible: $('#libraryDocVisible').checked,
    pdfName: pdfFile?.name || '',
    coverName: coverFile?.name || '',
    pdfDataUrl: await fileToDataUrl(pdfFile),
    coverDataUrl: await fileToDataUrl(coverFile)
  }};
  const response = await fetch(`${{libraryApiBase}}/documents${{id ? `/${{id}}` : ''}}`, {{
    method: id ? 'PUT' : 'POST',
    headers: {{ 'Content-Type': 'application/json', ...libraryAdminHeaders() }},
    body: JSON.stringify(payload)
  }});
  const result = await response.json().catch(() => ({{}}));
  if (!response.ok) throw new Error(result.error || 'Không lưu được tài liệu.');
  showLibraryMessage('#libraryAdminMsg', 'Đã lưu tài liệu.');
  resetLibraryDocForm();
  await fetchLibraryDocuments(true);
}}

async function handleLibraryAdminAction(event) {{
  const button = event.target.closest('button[data-action]');
  if (!button) return;
  const id = Number(button.dataset.id);
  const doc = libraryDocuments.find(item => Number(item.id) === id);
  if (!doc) return;
  const action = button.dataset.action;
  if (action === 'edit') {{
    fillLibraryDocForm(doc);
    return;
  }}
  if (action === 'toggle') {{
    const response = await fetch(`${{libraryApiBase}}/documents/${{id}}/visibility`, {{
      method: 'PATCH',
      headers: {{ 'Content-Type': 'application/json', ...libraryAdminHeaders() }},
      body: JSON.stringify({{ visible: !doc.visible }})
    }});
    const payload = await response.json().catch(() => ({{}}));
    if (!response.ok) throw new Error(payload.error || 'Không đổi được trạng thái tài liệu.');
    await fetchLibraryDocuments(true);
    return;
  }}
  if (action === 'delete') {{
    if (!confirm(`Xóa tài liệu "${{doc.title}}"?`)) return;
    const response = await fetch(`${{libraryApiBase}}/documents/${{id}}`, {{
      method: 'DELETE',
      headers: libraryAdminHeaders()
    }});
    const payload = await response.json().catch(() => ({{}}));
    if (!response.ok) throw new Error(payload.error || 'Không xóa được tài liệu.');
    await fetchLibraryDocuments(true);
  }}
}}

function drawPdfWatermark(ctx, width, height) {{
  const text = 'Thư viện số - Chỉ đọc trực tuyến';
  ctx.save();
  ctx.globalAlpha = 0.09;
  ctx.fillStyle = '#0f766e';
  ctx.font = `${{Math.max(22, Math.round(width / 28))}}px Arial`;
  ctx.textAlign = 'center';
  ctx.translate(width / 2, height / 2);
  ctx.rotate(-Math.PI / 6);
  for (let y = -height; y <= height; y += 170) {{
    for (let x = -width; x <= width; x += 420) {{
      ctx.fillText(text, x, y);
    }}
  }}
  ctx.restore();
}}

async function renderPdfPage() {{
  if (!activePdf) return;
  const serial = ++activePdfRenderSerial;
  try {{
    if (activePdfRenderTask) activePdfRenderTask.cancel();
  }} catch (error) {{}}
  const page = await activePdf.getPage(activePdfPage);
  if (serial !== activePdfRenderSerial) return;
  const canvas = $('#pdfCanvas');
  const ctx = canvas.getContext('2d', {{ alpha: false }});
  const cssViewport = page.getViewport({{ scale: activePdfScale }});
  const pixelRatio = Math.min(window.devicePixelRatio || 1, 2.5);
  const renderViewport = page.getViewport({{ scale: activePdfScale * pixelRatio }});
  canvas.width = Math.floor(renderViewport.width);
  canvas.height = Math.floor(renderViewport.height);
  canvas.style.width = `${{Math.floor(cssViewport.width)}}px`;
  canvas.style.height = `${{Math.floor(cssViewport.height)}}px`;
  ctx.fillStyle = '#ffffff';
  ctx.fillRect(0, 0, canvas.width, canvas.height);
  activePdfRenderTask = page.render({{ canvasContext: ctx, viewport: renderViewport }});
  await activePdfRenderTask.promise.catch(error => {{
    if (error?.name !== 'RenderingCancelledException') throw error;
  }});
  if (serial !== activePdfRenderSerial) return;
  drawPdfWatermark(ctx, canvas.width, canvas.height);
  $('#readerPageInput').value = activePdfPage;
  $('#readerPageTotal').textContent = `/ ${{activePdf.numPages}}`;
  $('#readerPrevBtn').disabled = activePdfPage <= 1;
  $('#readerNextBtn').disabled = activePdfPage >= activePdf.numPages;
}}

async function openPdfReader(doc) {{
  $('#pdfReader').hidden = false;
  $('#readerTitle').textContent = doc.title;
  const tokenResponse = await fetch(`${{libraryApiBase}}/documents/${{doc.id}}/view-token`, {{
    method: 'POST',
    headers: libraryAuthHeaders()
  }});
  const tokenPayload = await tokenResponse.json().catch(() => ({{}}));
  if (!tokenResponse.ok) throw new Error(tokenPayload.error || 'Không tạo được phiên đọc tài liệu.');
  const pdfjs = await loadPdfJs();
  const url = `${{libraryApiBase}}/documents/${{doc.id}}/pdf?token=${{encodeURIComponent(tokenPayload.token)}}`;
  activePdf = await pdfjs.getDocument({{ url, disableAutoFetch: true, disableStream: true }}).promise;
  activePdfPage = 1;
  activePdfScale = 1.2;
  await renderPdfPage();
}}

function closePdfReader() {{
  $('#pdfReader').hidden = true;
  activePdf = null;
  activePdfRenderSerial++;
}}

async function changePdfPage(delta) {{
  if (!activePdf) return;
  activePdfPage = Math.min(activePdf.numPages, Math.max(1, activePdfPage + delta));
  await renderPdfPage();
}}

async function setPdfPage(value) {{
  if (!activePdf) return;
  const next = Number(value);
  if (!Number.isFinite(next)) return;
  activePdfPage = Math.min(activePdf.numPages, Math.max(1, Math.trunc(next)));
  await renderPdfPage();
}}

async function zoomPdf(delta) {{
  activePdfScale = Math.min(3, Math.max(0.6, activePdfScale + delta));
  await renderPdfPage();
}}

function closeMainMenu() {{
  const menu = $('#menuList');
  const button = $('#menuBtn');
  if (menu) menu.hidden = true;
  if (button) button.setAttribute('aria-expanded', 'false');
}}

function closeToolDropdowns(except = null) {{
  $$('.tool-group.open, .sample-downloads.open').forEach(group => {{
    if (group !== except) group.classList.remove('open');
  }});
}}

function setActiveModuleLabel(label) {{
  const el = $('#activeModuleLabel');
  if (el) el.textContent = label || 'Trang chủ';
}}

function showHomePage() {{
  setActiveModuleLabel('Trang chủ');
  document.body.classList.add('home-mode');
  document.body.classList.remove('module-mode');
  document.body.classList.remove('docs-mode');
  document.body.classList.remove('webgis-mode');
  document.body.classList.remove('ai-webgis-assistant');
  $('#reportPanel').hidden = true;
  $('#aiPanel').hidden = true;
  $('#libraryAccessPanel').hidden = true;
  $('#importLog').hidden = true;
  closeMainMenu();
}}

function showLandTransferPage() {{
  setActiveModuleLabel('Chu chuyển đất đai');
  document.body.classList.add('module-mode');
  document.body.classList.remove('home-mode');
  document.body.classList.remove('docs-mode');
  document.body.classList.remove('webgis-mode');
  document.body.classList.remove('ai-webgis-assistant');
  $('#libraryAccessPanel').hidden = true;
  closeMainMenu();
  recalc();
}}

function showDocumentLibraryPage() {{
  if (!librarySessionToken) {{
    showLibraryAccessPanel();
    return;
  }}
  setActiveModuleLabel('Thư viện tài liệu PDF');
  document.body.classList.add('docs-mode');
  document.body.classList.remove('home-mode');
  document.body.classList.remove('module-mode');
  document.body.classList.remove('webgis-mode');
  document.body.classList.remove('ai-webgis-assistant');
  $('#reportPanel').hidden = true;
  $('#aiPanel').hidden = true;
  $('#importLog').hidden = true;
  closeMainMenu();
  updateLibrarySessionUi();
  fetchLibraryDocuments().catch(error => alert(error.message || String(error)));
}}

function showWebGisPage() {{
  setActiveModuleLabel('WebGIS quản lý dữ liệu đất đai');
  document.body.classList.add('webgis-mode');
  document.body.classList.remove('home-mode');
  document.body.classList.remove('module-mode');
  document.body.classList.remove('docs-mode');
  document.body.classList.remove('ai-webgis-assistant');
  $('#reportPanel').hidden = true;
  $('#aiPanel').hidden = true;
  $('#libraryAccessPanel').hidden = true;
  $('#libraryAdminPanel').hidden = true;
  $('#pdfReader').hidden = true;
  $('#importLog').hidden = true;
  closeMainMenu();
  webgisUpdateAdminUi();
  initializeWebGIS().catch(error => {{
    webgisSetSaveStatus('Không khởi động được WebGIS', true);
    alert(error.message || String(error));
  }});
}}

$('#sideHomeBtn').addEventListener('click', showHomePage);
$('#sideLandTransferBtn').addEventListener('click', showLandTransferPage);
$('#sideLibraryBtn').addEventListener('click', showDocumentLibraryPage);
$('#sideWebGisBtn').addEventListener('click', showWebGisPage);
$('#homeLandTransferBtn').addEventListener('click', showLandTransferPage);
$('#homeLibraryBtn').addEventListener('click', showDocumentLibraryPage);
$('#homeWebGisBtn').addEventListener('click', showWebGisPage);
$('#homeBtn').addEventListener('click', showHomePage);
$('#libraryHomeBtn').addEventListener('click', showHomePage);
$('#libraryLogoutBtn').addEventListener('click', () => {{
  clearLibrarySession();
  libraryDocuments = [];
  renderLibraryGrid([]);
  showLibraryAccessPanel('Đã đăng xuất. Vui lòng đăng nhập lại để vào thư viện.');
}});
$('#libraryAccessCloseBtn').addEventListener('click', () => $('#libraryAccessPanel').hidden = true);
$('#libraryAccessLoginBtn').addEventListener('click', () => {{
  libraryAccessLogin().catch(error => showLibraryMessage('#libraryAccessMsg', error.message || String(error), true));
}});
['libraryAccessUser', 'libraryAccessPassword'].forEach(id => {{
  $(`#${{id}}`).addEventListener('keydown', event => {{
    if (event.key === 'Enter') libraryAccessLogin().catch(error => showLibraryMessage('#libraryAccessMsg', error.message || String(error), true));
  }});
}});
$('#libraryAdminOpenBtn').addEventListener('click', openLibraryAdminPanel);
$('#libraryAdminCloseBtn').addEventListener('click', () => $('#libraryAdminPanel').hidden = true);
$('#libraryDocForm').addEventListener('submit', event => {{
  saveLibraryDocument(event).catch(error => showLibraryMessage('#libraryAdminMsg', error.message || String(error), true));
}});
$('#libraryDocNewBtn').addEventListener('click', resetLibraryDocForm);
$('#libraryAdminReloadBtn').addEventListener('click', () => {{
  fetchLibraryDocuments(true).catch(error => showLibraryMessage('#libraryAdminMsg', error.message || String(error), true));
}});
$('#libraryAdminRows').addEventListener('click', event => {{
  handleLibraryAdminAction(event).catch(error => showLibraryMessage('#libraryAdminMsg', error.message || String(error), true));
}});
$('#libraryGrid').addEventListener('click', event => {{
  const button = event.target.closest('.library-read-btn');
  if (!button) return;
  const doc = libraryDocuments.find(item => Number(item.id) === Number(button.dataset.id));
  if (doc) openPdfReader(doc).catch(error => alert(error.message || String(error)));
}});
['librarySearch', 'libraryCategoryFilter', 'libraryYearFilter'].forEach(id => {{
  const input = $(`#${{id}}`);
  input.addEventListener(id === 'librarySearch' ? 'input' : 'change', () => fetchLibraryDocuments().catch(error => alert(error.message || String(error))));
}});
$('#libraryRefreshBtn').addEventListener('click', () => fetchLibraryDocuments().catch(error => alert(error.message || String(error))));
$('#readerCloseBtn').addEventListener('click', closePdfReader);
$('#readerPrevBtn').addEventListener('click', () => changePdfPage(-1));
$('#readerNextBtn').addEventListener('click', () => changePdfPage(1));
$('#readerPageInput').addEventListener('change', event => setPdfPage(event.currentTarget.value));
$('#readerZoomOutBtn').addEventListener('click', () => zoomPdf(-0.15));
$('#readerZoomInBtn').addEventListener('click', () => zoomPdf(0.15));
$('#readerFullscreenBtn').addEventListener('click', () => {{
  const reader = $('#pdfReader');
  if (!document.fullscreenElement) reader.requestFullscreen?.();
  else document.exitFullscreen?.();
}});
$('#pdfReader').addEventListener('contextmenu', event => event.preventDefault());
$('#pdfReader').addEventListener('selectstart', event => event.preventDefault());
$('#pdfReader').addEventListener('dragstart', event => event.preventDefault());
document.addEventListener('keydown', event => {{
  if ($('#pdfReader').hidden) return;
  const key = event.key.toLowerCase();
  const blocked = event.key === 'F12' ||
    event.key === 'PrintScreen' ||
    ((event.ctrlKey || event.metaKey) && ['c', 'v', 's', 'p', 'a', 'u'].includes(key));
  if (blocked) {{
    event.preventDefault();
    event.stopPropagation();
  }}
}});
$$('.tool-group-title').forEach(button => {{
  button.addEventListener('click', event => {{
    event.stopPropagation();
    const group = event.currentTarget.closest('.tool-group');
    const willOpen = !group.classList.contains('open');
    closeToolDropdowns(group);
    group.classList.toggle('open', willOpen);
  }});
}});
$('.sample-downloads > span').addEventListener('click', event => {{
  event.stopPropagation();
  const group = event.currentTarget.closest('.sample-downloads');
  const willOpen = !group.classList.contains('open');
  closeToolDropdowns(group);
  group.classList.toggle('open', willOpen);
}});
$$('.tool-items, .sample-items').forEach(panel => {{
  panel.addEventListener('click', event => event.stopPropagation());
}});
document.addEventListener('click', event => {{
  if (!event.target.closest('.main-menu')) closeMainMenu();
  closeToolDropdowns();
}});
document.addEventListener('keydown', event => {{
  if (event.key === 'Escape') {{
    closeMainMenu();
    closeToolDropdowns();
  }}
}});

function landName(code) {{
  const row = rowsByCode[code];
  const cell = row ? cellsByKey.get(`${{row}}:2`) : null;
  return cell ? cell.textContent.trim() : code;
}}

function buildWebgisAiContext() {{
  if (typeof webgisState === 'undefined') return null;
  const layers = (webgisState.layerDefs || []);
  const visibleLayers = layers
    .filter(layer => layer.visible === true || (webgisState.overlayLayers && webgisState.overlayLayers.has(layer.id)))
    .map(layer => ({{
      id: layer.id,
      label: layer.label,
      category: layer.category,
      opacity: layer.opacity,
      featureCount: typeof webgisLayerFeatureCount === 'function' ? webgisLayerFeatureCount(layer.id) : 0
    }}));
  let selectedFeature = null;
  if (webgisState.selectedFeatureId) {{
    const feature = (webgisState.features || []).find(item => item?.properties?.__id === webgisState.selectedFeatureId);
    if (feature) {{
      const entries = typeof webgisVisiblePropertyEntries === 'function'
        ? webgisVisiblePropertyEntries(feature)
        : Object.entries(feature.properties || {{}}).filter(([key]) => !['__id', 'layer'].includes(key));
      selectedFeature = {{
        layer: feature.properties?.layer || '',
        layerLabel: typeof webgisLayerLabel === 'function' ? webgisLayerLabel(feature.properties?.layer) : feature.properties?.layer || '',
        title: typeof webgisFeatureTitle === 'function' ? webgisFeatureTitle(feature) : '',
        geometryType: feature.geometry?.type || '',
        properties: Object.fromEntries(entries.slice(0, 30))
      }};
    }}
  }}
  return {{
    module: document.body.classList.contains('webgis-mode') ? 'webgis' : document.body.classList.contains('docs-mode') ? 'library' : document.body.classList.contains('module-mode') ? 'land-transfer' : 'home',
    layerCount: layers.length,
    featureCount: (webgisState.features || []).length,
    loadedFeatureCount: typeof webgisAllCachedFeatures === 'function' ? webgisAllCachedFeatures().length : (webgisState.features || []).length,
    publicLayerCount: layers.filter(layer => layer.is_public !== false).length,
    visibleLayers,
    selectedFeature
  }};
}}

function buildAiContext() {{
  normalizeAllInputs();
  recalc();
  const calc = createCalcContext();
  const codes = matrixCodes.filter(code => rowsByCode[code]);
  const landTypes = codes.map(code => {{
    const current = calc.currentArea(code);
    const plan = calc.matrixValue('DTTN', code);
    const diagonal = calc.matrixValue(code, code);
    return {{
      code,
      name: landName(code),
      current: roundNumber(current),
      planning: roundNumber(plan),
      decrease: roundNumber(current - diagonal),
      increase: roundNumber(plan - diagonal),
      change: roundNumber(plan - current)
    }};
  }}).filter(item =>
    Math.abs(item.current) > meta.tolerance ||
    Math.abs(item.planning) > meta.tolerance ||
    Math.abs(item.decrease) > meta.tolerance ||
    Math.abs(item.increase) > meta.tolerance ||
    Math.abs(item.change) > meta.tolerance
  );

  const transfers = [];
  inputCodes.forEach(fromCode => {{
    inputCodes.forEach(toCode => {{
      const value = calc.matrixLeaf(fromCode, toCode);
      if (Math.abs(value) > meta.tolerance) {{
        transfers.push({{
          fromCode,
          fromName: landName(fromCode),
          toCode,
          toName: landName(toCode),
          area: roundNumber(value)
        }});
      }}
    }});
  }});
  transfers.sort((a, b) => Math.abs(b.area) - Math.abs(a.area));

  const totalCurrent = calc.currentArea('DTTN');
  const totalPlanning = calc.matrixValue('DTTN', 'DTTN');
  const activeModule = document.body.classList.contains('webgis-mode') ? 'webgis' : document.body.classList.contains('docs-mode') ? 'library' : document.body.classList.contains('module-mode') ? 'land-transfer' : 'home';
  return {{
    activeModule,
    unit: 'ha',
    decimals: displayDecimals,
    tolerance: meta.tolerance,
    totals: {{
      current: roundNumber(totalCurrent),
      planning: roundNumber(totalPlanning),
      difference: roundNumber(totalPlanning - totalCurrent)
    }},
    project: typeof readProjectSettings === 'function' ? readProjectSettings() : {{}},
    webgis: buildWebgisAiContext(),
    landTypes,
    topTransfers: transfers.slice(0, 40)
  }};
}}

function appendAiMessage(type, text) {{
  const el = document.createElement('div');
  el.className = `ai-message ${{type || ''}}`.trim();
  el.textContent = text;
  $('#aiMessages').appendChild(el);
  $('#aiMessages').scrollTop = $('#aiMessages').scrollHeight;
  return el;
}}

async function refreshAiStatus() {{
  const status = $('#aiStatus');
  if (!status) return;
  status.className = 'ai-status';
  status.textContent = 'Đang kiểm tra cấu hình AI...';
  try {{
    const response = await fetch('/api/ai/status');
    const payload = await response.json().catch(() => ({{}}));
    if (!response.ok) throw new Error(payload.error || 'Không kiểm tra được cấu hình AI.');
    if (payload.enabled) {{
      const providerLabel = payload.provider === 'gemini' ? 'Gemini' : 'OpenAI';
      status.classList.add('ready');
      status.textContent = 'AI đã sẵn sàng: ' + providerLabel + ' - ' + payload.model + '.';
    }} else {{
      status.classList.add('error');
      status.textContent = 'AI chưa được cấu hình trên server hiện tại. Nếu chạy local, tạo/cập nhật file .env cùng cấp package.json rồi npm start lại. Nếu chạy Render, thêm OPENAI_API_KEY hoặc GEMINI_API_KEY trong Environment rồi Save, rebuild, and deploy.';
    }}
  }} catch (error) {{
    status.classList.add('error');
    status.textContent = error.message || 'Không kiểm tra được trạng thái AI.';
  }}
}}

function openAiAssistant(mode = 'land-transfer') {{
  const isWebgis = mode === 'webgis';
  document.body.classList.toggle('ai-webgis-assistant', isWebgis);
  $('#aiPanelTitle').textContent = isWebgis ? 'Trợ lý AI WebGIS' : 'Trợ lý AI';
  $('#aiIntroMessage').textContent = isWebgis
    ? 'Anh có thể hỏi: “Layer nào đang bật?”, “Tóm tắt đối tượng đang chọn”, “Thuộc tính nào đang được phép hiển thị?”, hoặc “Nhận xét nhanh dữ liệu hiện trạng/quy hoạch trên bản đồ”.'
    : 'Anh có thể hỏi: “Kiểm tra giúp tôi bảng này có lệch tổng không?”, “LUC tăng giảm thế nào?”, hoặc “Viết nhận xét ngắn về biến động đất”.';
  $('#aiQuestion').placeholder = isWebgis
    ? 'Nhập câu hỏi về WebGIS, layer, thuộc tính hoặc đối tượng đang chọn'
    : 'Nhập câu hỏi cho AI';
  $('#aiPanel').hidden = false;
  refreshAiStatus();
  $('#aiQuestion').focus();
}}

async function sendAiQuestion() {{
  const input = $('#aiQuestion');
  const question = input.value.trim();
  if (!question) return;
  input.value = '';
  appendAiMessage('user', question);
  const waiting = appendAiMessage('', 'AI đang phân tích dữ liệu...');
  $('#aiSendBtn').disabled = true;
  try {{
    const response = await fetch('/api/ai', {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify({{ question, context: buildAiContext() }})
    }});
    const payload = await response.json().catch(() => ({{ error: response.statusText }}));
    if (!response.ok) throw new Error(payload.error || 'Không gọi được AI.');
    waiting.textContent = payload.answer || 'AI không trả về nội dung.';
  }} catch (error) {{
    const message = error.message || String(error);
    waiting.textContent = message.includes('fetch')
      ? 'Không kết nối được server AI. Hãy chạy npm start và mở phần mềm tại http://127.0.0.1:3000.'
      : message;
  }} finally {{
    $('#aiSendBtn').disabled = false;
  }}
}}

inputEls.forEach(input => {{
  input.addEventListener('input', () => {{
    updateInputZeroState(input);
    scheduleRecalc();
  }});
}});

function applyHideZeroState(enabled) {{
  document.body.classList.toggle('hide-zero', enabled);
  $('#hideZeroToggle').checked = enabled;
  localStorage.setItem(hideZeroKey, enabled ? '1' : '0');
}}

$('#hideZeroToggle').addEventListener('change', event => {{
  applyHideZeroState(event.currentTarget.checked);
}});
$('#compactColumnsToggle').addEventListener('change', event => {{
  applyCompactColumnsState(event.currentTarget.checked);
}});

let hoverRow = null;
let hoverCol = null;
let hoverCell = null;
function clearTableHover() {{
  if (hoverRow !== null) $$(`td[data-row="${{hoverRow}}"]`).forEach(td => td.classList.remove('hover-row'));
  if (hoverCol !== null) $$(`td[data-col="${{hoverCol}}"]`).forEach(td => td.classList.remove('hover-col'));
  if (hoverCell) hoverCell.classList.remove('hover-cell');
  hoverRow = null;
  hoverCol = null;
  hoverCell = null;
}}
$('#landTable').addEventListener('mouseover', event => {{
  const td = event.target.closest('td');
  if (!td || hoverCell === td) return;
  clearTableHover();
  hoverRow = td.dataset.row;
  hoverCol = td.dataset.col;
  hoverCell = td;
  $$(`td[data-row="${{hoverRow}}"]`).forEach(cell => cell.classList.add('hover-row'));
  $$(`td[data-col="${{hoverCol}}"]`).forEach(cell => cell.classList.add('hover-col'));
  td.classList.add('hover-cell');
}});
$('#landTable').addEventListener('mouseleave', clearTableHover);
$('#codeSearchBtn').addEventListener('click', () => jumpToLandCode($('#codeSearch').value));
$('#codeSearch').addEventListener('keydown', event => {{
  if (event.key === 'Enter') {{
    event.preventDefault();
    jumpToLandCode(event.currentTarget.value);
  }}
}});
$('#reportBtn').addEventListener('click', () => {{
  syncProjectYearsToReport();
  renderReportOptions($('#reportFilter').value);
  $('#reportPanel').hidden = false;
}});
$('#reportCloseBtn').addEventListener('click', () => $('#reportPanel').hidden = true);
$('#aiBtn').addEventListener('click', () => {{
  openAiAssistant('land-transfer');
}});
$('#aiCloseBtn').addEventListener('click', () => {{
  $('#aiPanel').hidden = true;
  document.body.classList.remove('ai-webgis-assistant');
}});
$('#aiSendBtn').addEventListener('click', sendAiQuestion);
$('#aiQuestion').addEventListener('keydown', event => {{
  if (event.key === 'Enter' && !event.shiftKey) {{
    event.preventDefault();
    sendAiQuestion();
  }}
}});
$('#reportFilter').addEventListener('input', event => renderReportOptions(event.currentTarget.value));
['projectCommune', 'projectProvince', 'projectPreviousPlanYear', 'projectCurrentYear', 'projectPlanYear'].forEach(id => {{
  const input = $(`#${{id}}`);
  if (!input) return;
  input.addEventListener('input', () => {{
    projectTitlesConfirmed = false;
    syncProjectYearsToReport();
  }});
  input.addEventListener('change', () => localStorage.setItem(storageKey, JSON.stringify(readInputs())));
}});
['reportCurrentYear', 'reportPlanYear'].forEach(id => {{
  const input = $(`#${{id}}`);
  if (!input) return;
  input.addEventListener('input', syncReportYearsToProject);
  input.addEventListener('change', () => localStorage.setItem(storageKey, JSON.stringify(readInputs())));
}});
$('#projectConfirmBtn').addEventListener('click', () => {{
  projectTitlesConfirmed = true;
  syncProjectYearsToReport();
  updateProjectTitles();
  localStorage.setItem(storageKey, JSON.stringify(readInputs()));
  $('#projectConfirmBtn').textContent = 'Đã xác nhận';
  setTimeout(() => $('#projectConfirmBtn').textContent = 'Xác nhận', 900);
}});
$('#gtpOpenBtn').addEventListener('click', () => $('#gtpInput').click());
$('#gtpInput').addEventListener('change', async event => {{
  const file = event.target.files[0];
  if (!file) return;
  try {{
    await openGtpProjectFile(file);
  }} catch (error) {{
    alert(error.message || String(error));
  }} finally {{
    event.target.value = '';
  }}
}});
$('#gtpSetupBtn').addEventListener('click', async () => {{
  try {{
    await saveGtpFile({{ choose: true }});
  }} catch (error) {{
    if (error && error.name === 'AbortError') return;
    alert(error.message || String(error));
  }}
}});
$('#gtpSaveBtn').addEventListener('click', async () => {{
  try {{
    await saveGtpFile({{ choose: !gtpFileHandle }});
  }} catch (error) {{
    if (error && error.name === 'AbortError') return;
    alert(error.message || String(error));
  }}
}});
$('#reportSelectActiveBtn').addEventListener('click', () => {{
  normalizeAllInputs();
  recalc();
  const calc = createCalcContext();
  renderReportOptions($('#reportFilter').value);
  document.querySelectorAll('#reportOptions input').forEach(input => {{
    input.checked = hasReportData(input.value, calc);
  }});
}});
$('#reportClearBtn').addEventListener('click', () => {{
  document.querySelectorAll('#reportOptions input').forEach(input => input.checked = false);
}});
$('#reportExportBtn').addEventListener('click', exportReportWord);
$('#saveBtn').addEventListener('click', async () => {{
  const data = readInputs();
  localStorage.setItem(storageKey, JSON.stringify(data));
  $('#saveBtn').disabled = true;
  $('#saveBtn').textContent = 'Đang lưu';
  const failures = [];
  if (gtpFileHandle) {{
    try {{
      await saveGtpFile({{ silent: true }});
    }} catch (error) {{
      failures.push(error.message || String(error));
    }}
  }}
  try {{
    await saveProjectToServer();
  }} catch (error) {{
    failures.push(error.message || String(error));
  }}
  if (failures.length) {{
    $('#saveBtn').textContent = 'Lưu lỗi';
    alert(failures.join('\\n'));
  }} else {{
    $('#saveBtn').textContent = 'Đã lưu';
  }}
  try {{
    setTimeout(() => {{
      $('#saveBtn').disabled = false;
      $('#saveBtn').textContent = 'Lưu';
    }}, 900);
  }} catch (error) {{}}
}});
$('#importGisBtn').addEventListener('click', () => $('#gisXlsxInput').click());
$('#gisXlsxInput').addEventListener('change', async event => {{
  const file = event.target.files[0];
  if (!file) return;
  try {{
    await importGISOverlayExcel(file);
  }} catch (error) {{
    alert(error.message || String(error));
  }} finally {{
    event.target.value = '';
  }}
}});
$('#importCurrentBtn').addEventListener('click', () => $('#currentXlsxInput').click());
$('#currentXlsxInput').addEventListener('change', async event => {{
  const file = event.target.files[0];
  if (!file) return;
  try {{
    const result = await importCurrentAreasFromXlsx(file);
    localStorage.setItem(storageKey, JSON.stringify(readInputs()));
    const msg = `Đã nhập ${{result.imported}} ô hiện trạng từ XLSX` +
      (result.matchedNoValue ? `; ${{result.matchedNoValue}} mã trùng nhưng trống diện tích` : '') +
      (result.adjustments.length ? `; cân sai số làm tròn: ${{result.adjustments.map(item => `${{item.parentCode}} -> ${{item.targetCode}} ${{formatNumber(item.diff)}}`).join(', ')}}` : '') +
      (result.unmatched.length ? `; bỏ qua mã không phải dòng nhập: ${{result.unmatched.slice(0, 8).join(', ')}}` : '');
    alert(msg);
  }} catch (error) {{
    alert(error.message || String(error));
  }} finally {{
    event.target.value = '';
  }}
}});
$('#importPreviousPlanBtn').addEventListener('click', () => $('#previousPlanXlsxInput').click());
$('#previousPlanXlsxInput').addEventListener('change', async event => {{
  const file = event.target.files[0];
  if (!file) return;
  try {{
    const result = await importPreviousPlanExcel(file);
    alert(`Đã nhập ${{result.validRows}} dòng quy hoạch kỳ trước từ XLSX`);
  }} catch (error) {{
    alert(error.message || String(error));
  }} finally {{
    event.target.value = '';
  }}
}});
$('#jsonBtn').addEventListener('click', () => download('du_lieu_chu_chuyen_dat_dai.json', 'application/json;charset=utf-8', JSON.stringify(readInputs(), null, 2)));
$('#xlsxBtn').addEventListener('click', exportXlsx);
$('#csvBtn').addEventListener('click', exportCsv);
$('#printBtn').addEventListener('click', () => window.print());
$('#clearBtn').addEventListener('click', () => {{
  if (!confirm('Xóa toàn bộ dữ liệu nhập trong trang?')) return;
  const projectSettings = readProjectSettings();
  inputEls.forEach(input => input.value = '');
  applyPreviousPlanValues({{}});
  applyProjectSettings(projectSettings);
  localStorage.setItem(storageKey, JSON.stringify(readInputs()));
  recalc();
}});
$('#loadBtn').addEventListener('click', () => $('#fileInput').click());
$('#fileInput').addEventListener('change', async event => {{
  const file = event.target.files[0];
  if (!file) return;
  applyProjectData(gtpDataFromPayload(JSON.parse(await file.text())));
  event.target.value = '';
}});

const saved = localStorage.getItem(storageKey);
if (saved) applyInputs(JSON.parse(saved));
normalizeAllInputs();
applyHideZeroState(localStorage.getItem(hideZeroKey) === '1');
applyCompactColumnsState(localStorage.getItem(compactColumnsKey) === '1');
updateLibrarySessionUi();
$('#statusMissing').textContent = meta.missingCodes.length ? `Thiếu mã: ${{meta.missingCodes.join(', ')}}` : 'Đủ mã nhập';
$('#statusMissing').classList.toggle('warn', meta.missingCodes.length > 0);
recalc();
loadProjectFromServer();
</script>
</body>
</html>
"""
    OUT.write_text(doc, encoding="utf-8")
    print(OUT)
    print("input_codes=", ",".join(input_codes))
    print("missing_codes=", ",".join(missing_codes))


if __name__ == "__main__":
    main()
