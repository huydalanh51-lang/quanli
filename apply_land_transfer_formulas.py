from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Protection
from openpyxl.workbook.properties import CalcProperties
from openpyxl.utils import get_column_letter


INPUT = Path(r"C:\Users\QUANGHUY\Downloads\Bieu_chu_chuyen_dat_dai_mau_cong_thuc.xlsx")
OUTPUT = Path(r"C:\Users\QUANGHUY\Downloads\Bieu_chu_chuyen_dat_dai_da_ap_dung_cong_thuc.xlsx")

HEADER_ROW = 3
LABEL_COL = 2
CODE_COL = 3
CURRENT_COL = 4
MATRIX_START_COL = 5
MATRIX_END_COL = 66
DECREASE_COL = 67
CHANGE_COL = 68
PLAN_COL = 69
TOTAL_INCREASE_ROW = 67
PLAN_ROW = 68
TOLERANCE = 0.0001

INPUT_DETAIL_CODES = ["LUC", "LUK", "HNK", "CLN", "NTS", "DHT", "ONT", "NTD", "DSH", "CSD"]
AGGREGATES = {
    "LUA": ["LUC", "LUK"],
    "NNP": ["LUA", "HNK", "CLN", "NTS"],
    "PNN": ["DHT", "ONT", "NTD", "DSH"],
}
DTTN_CHILDREN = ["NNP", "PNN", "CSD"]


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def as_sum(refs: list[str]) -> str:
    return "=SUM(" + ",".join(refs) + ")" if refs else '=""'


def find_sheet(wb: openpyxl.Workbook):
    candidates = []
    for ws in wb.worksheets:
        labels = [str(ws.cell(r, c).value or "").strip().lower() for r in range(1, min(ws.max_row, 10) + 1) for c in range(1, min(ws.max_column, 10) + 1)]
        row3_codes = {str(ws.cell(HEADER_ROW, c).value or "").strip() for c in range(MATRIX_START_COL, min(ws.max_column, MATRIX_END_COL) + 1)}
        score = 0
        if any("chu chuyển" in v for v in labels):
            score += 2
        if {"NNP", "LUA", "LUC", "PNN", "CSD"} <= row3_codes:
            score += 5
        if str(ws.cell(2, CODE_COL).value or "").strip().startswith("Mã"):
            score += 1
        if score:
            candidates.append((score, ws))
    if not candidates:
        raise RuntimeError("Không nhận diện được sheet chứa bảng chu chuyển đất đai.")
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1]


def main():
    wb = openpyxl.load_workbook(INPUT, data_only=False)
    ws = find_sheet(wb)

    code_to_row: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        code = ws.cell(row, CODE_COL).value
        if code is not None:
            code_to_row[str(code).strip()] = row

    code_to_col: dict[str, int] = {}
    for col in range(MATRIX_START_COL, MATRIX_END_COL + 1):
        code = ws.cell(HEADER_ROW, col).value
        if code is not None:
            code_to_col[str(code).strip()] = col

    dttn_row = None
    for row in range(1, ws.max_row + 1):
        label = str(ws.cell(row, LABEL_COL).value or "").strip().lower()
        if "tổng diện tích tự nhiên" in label:
            dttn_row = row
            break
    if dttn_row is None:
        raise RuntimeError("Không tìm thấy dòng Tổng diện tích tự nhiên.")

    existing_input_codes = [code for code in INPUT_DETAIL_CODES if code in code_to_row]
    missing_input_codes = [code for code in INPUT_DETAIL_CODES if code not in code_to_row]

    # Lock the sheet by default, then unlock only the direct input cells.
    for row in ws.iter_rows():
        for cell in row:
            cell.protection = copy(cell.protection)
            cell.protection = Protection(locked=True, hidden=False)

    for code in existing_input_codes:
        row = code_to_row[code]
        ws.cell(row, CURRENT_COL).protection = Protection(locked=False, hidden=False)
        if is_formula(ws.cell(row, CURRENT_COL).value) and code == "CSD":
            ws.cell(row, CURRENT_COL).value = None

    for row_code in existing_input_codes:
        row = code_to_row[row_code]
        for col_code in existing_input_codes:
            if col_code in code_to_col:
                cell = ws.cell(row, code_to_col[col_code])
                if is_formula(cell.value):
                    cell.value = None
                cell.protection = Protection(locked=False, hidden=False)

    def existing_children(parent: str) -> list[str]:
        return [code for code in AGGREGATES[parent] if code in code_to_row]

    # Current-area formulas.
    for parent in ["LUA", "NNP", "PNN"]:
        if parent in code_to_row:
            children = existing_children(parent)
            ws.cell(code_to_row[parent], CURRENT_COL).value = as_sum([f"D{code_to_row[ch]}" for ch in children])
    ws.cell(dttn_row, CURRENT_COL).value = as_sum([f"D{code_to_row[ch]}" for ch in DTTN_CHILDREN if ch in code_to_row])

    # Aggregate rows in the transfer matrix.
    for parent in ["LUA", "NNP", "PNN"]:
        if parent not in code_to_row:
            continue
        parent_row = code_to_row[parent]
        children = existing_children(parent)
        for col in range(MATRIX_START_COL, MATRIX_END_COL + 1):
            letter = get_column_letter(col)
            ws.cell(parent_row, col).value = as_sum([f"{letter}{code_to_row[ch]}" for ch in children])

    for col in range(MATRIX_START_COL, MATRIX_END_COL + 1):
        letter = get_column_letter(col)
        ws.cell(dttn_row, col).value = as_sum([f"{letter}{code_to_row[ch]}" for ch in DTTN_CHILDREN if ch in code_to_row])

    # Aggregate columns in the transfer matrix.
    for parent in ["LUA", "NNP", "PNN"]:
        if parent not in code_to_col:
            continue
        parent_col = code_to_col[parent]
        child_cols = [code_to_col[ch] for ch in AGGREGATES[parent] if ch in code_to_col]
        for row in range(dttn_row, PLAN_ROW):
            ws.cell(row, parent_col).value = as_sum([f"{get_column_letter(c)}{row}" for c in child_cols])

    # Cộng giảm, biến động, diện tích quy hoạch.
    for code, row in code_to_row.items():
        if not (dttn_row <= row < TOTAL_INCREASE_ROW):
            continue
        ws.cell(row, DECREASE_COL).value = f'=IFERROR($D{row}-INDEX($E{row}:$BN{row},1,MATCH($C{row},$E$3:$BN$3,0)),"")'
        ws.cell(row, CHANGE_COL).value = f'=IFERROR($BQ{row}-$D{row},"")'
        ws.cell(row, PLAN_COL).value = f'=IFERROR(INDEX($E${PLAN_ROW}:$BN${PLAN_ROW},1,MATCH($C{row},$E$3:$BN$3,0)),"")'

    ws.cell(dttn_row, DECREASE_COL).value = f"=SUM(BO5,BO19,BO61)"
    ws.cell(dttn_row, CHANGE_COL).value = f"=$BQ{dttn_row}-$D{dttn_row}"
    ws.cell(dttn_row, PLAN_COL).value = as_sum([f"BQ{code_to_row[ch]}" for ch in DTTN_CHILDREN if ch in code_to_row])

    for col in range(MATRIX_START_COL, MATRIX_END_COL + 1):
        letter = get_column_letter(col)
        ws.cell(TOTAL_INCREASE_ROW, col).value = f'=IFERROR({letter}{PLAN_ROW}-INDEX({letter}${dttn_row}:{letter}${TOTAL_INCREASE_ROW-1},MATCH({letter}$3,$C${dttn_row}:$C${TOTAL_INCREASE_ROW-1},0)),"")'
        ws.cell(PLAN_ROW, col).value = as_sum([f"{letter}{code_to_row[code]}" for code in existing_input_codes])
    ws.cell(TOTAL_INCREASE_ROW, DECREASE_COL).value = f"=SUM(BO5,BO19,BO61)"

    # Conditional warning checks. They do not add visible columns or disturb layout.
    warn_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    leaf_cols = [code_to_col[code] for code in existing_input_codes if code in code_to_col]
    checked_rows = [dttn_row] + [code_to_row[code] for code in ["NNP", "LUA", "PNN"] if code in code_to_row] + [code_to_row[code] for code in existing_input_codes]
    for row in checked_rows:
        refs = ",".join(f"{get_column_letter(col)}{row}" for col in leaf_cols)
        if refs:
            formula = f"ABS(SUM({refs})-$D{row})>{TOLERANCE}"
            ws.conditional_formatting.add(f"D{row}:BQ{row}", FormulaRule(formula=[formula], fill=warn_fill))

    ws.conditional_formatting.add(
        f"D{dttn_row}:BQ{dttn_row}",
        FormulaRule(formula=[f"ABS($D{dttn_row}-$BQ{dttn_row})>{TOLERANCE}"], fill=warn_fill),
    )

    # Keep formulas fresh when the user opens the file.
    if wb.calculation is None:
        wb.calculation = CalcProperties()
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True

    wb.save(OUTPUT)

    # Lightweight diagnostics from direct numeric inputs only.
    row_mismatches = []
    for code in existing_input_codes:
        row = code_to_row[code]
        current = ws.cell(row, CURRENT_COL).value
        if isinstance(current, (int, float)):
            total = 0.0
            has_value = False
            for col in leaf_cols:
                value = ws.cell(row, col).value
                if isinstance(value, (int, float)):
                    total += value
                    has_value = True
            if has_value and abs(total - float(current)) > TOLERANCE:
                row_mismatches.append((code, row, current, total))

    print(f"SHEET={ws.title}")
    print(f"OUTPUT={OUTPUT}")
    print(f"EXISTING_INPUT_CODES={','.join(existing_input_codes)}")
    print(f"MISSING_INPUT_CODES={','.join(missing_input_codes)}")
    print(f"MATRIX=E:BN")
    print(f"DECREASE_COL=BO CHANGE_COL=BP PLAN_COL=BQ")
    print(f"ROW_MISMATCHES={len(row_mismatches)}")
    for item in row_mismatches:
        print("ROW_MISMATCH", item)


if __name__ == "__main__":
    main()
